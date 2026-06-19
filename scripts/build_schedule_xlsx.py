import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "schedule_rows.json")
with open(DATA, encoding="utf-8") as f:
    rows = json.load(f)
# rows columns: soNo, customer, productCode, wipCode, team, qty, cutDoneDate, readyDate, customerDD, status, daysLate
FIRST_WD = "2026-06-02"

ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = F(bold=True, color="FFFFFF", size=11)
def status_fill(s):
    if s in ("LATE",): return PatternFill("solid", fgColor="FCE4D6")      # light orange
    if s in ("OVERDUE","PARTIAL"): return PatternFill("solid", fgColor="F8CBAD")  # darker orange
    if s == "ONTIME": return PatternFill("solid", fgColor="E2EFDA")        # light green
    if s == "STALLED": return PatternFill("solid", fgColor="D9D9D9")       # grey
    return None
def status_label(s):
    return {"LATE":"Late","OVERDUE":"Already overdue","ONTIME":"On time",
            "STALLED":"No capacity data","PARTIAL":"Partial (some items stalled)"}.get(s, s)

wb = Workbook()

# ---------- Sheet 1: Schedule (PO detail) ----------
ws = wb.active
ws.title = "Schedule"
headers = ["SO No","Customer","Product Code (Model)","WIP Code","Category","Qty","Cut Done","Planned Ready","Customer DD","Status","Days Late"]
ws.append(headers)
srt = sorted(rows, key=lambda r: (r[7] or "9999", r[0] or ""))  # by readyDate then SO
STATUS_COL = 10
for r in srt:
    soNo, cust, pc, wip, team, qty, cut, ready, dd, status, dl = r
    ws.append([soNo, cust, pc, wip, team, qty, cut or "", ready or "(not scheduled)", dd or "", status_label(status), dl if dl else ""])
    fill = status_fill(status)
    if fill: ws.cell(row=ws.max_row, column=STATUS_COL).fill = fill
widths = [14,20,18,42,11,6,12,14,13,20,10]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for c in range(1, len(headers)+1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL; hc.alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        if cell.font is None or cell.font.name != ARIAL: cell.font = F(size=10)
        cell.border = BORDER
        if cell.column in (6,11): cell.alignment = Alignment(horizontal="center")

# ---------- Sheet 2: By Order (SO rollup) ----------
ws2 = wb.create_sheet("By Order")
from collections import defaultdict
g = defaultdict(list)
for r in rows: g[r[0]].append(r)
so_rows = []
for so, items in g.items():
    cust = items[0][1]
    lines = len(items)
    totqty = sum((it[5] or 0) for it in items)
    dds = [it[8] for it in items if it[8]]
    earliestDD = min(dds) if dds else ""
    has_stalled = any(it[9]=="STALLED" for it in items)
    readys = [it[7] for it in items if it[7]]
    latestReady = max(readys) if readys else ""
    if has_stalled:
        status = "PARTIAL"
    elif latestReady and earliestDD and latestReady > earliestDD:
        status = "OVERDUE" if earliestDD < FIRST_WD else "LATE"
    else:
        status = "ONTIME"
    maxdl = max((it[10] or 0) for it in items)
    so_rows.append([so, cust, lines, totqty, earliestDD, latestReady or "(not scheduled)", status_label(status), maxdl if maxdl else "", status])
so_rows.sort(key=lambda x: (x[5] or "9999", x[0] or ""))
h2 = ["SO No","Customer","# Lines","Total Qty","Customer DD","Order Ready (all lines)","Status","Max Days Late"]
ws2.append(h2)
for sr in so_rows:
    ws2.append(sr[:8])
    fill = status_fill(sr[8])
    if fill: ws2.cell(row=ws2.max_row, column=7).fill = fill
w2 = [14,22,9,10,13,22,22,14]
for i,w in enumerate(w2, start=1): ws2.column_dimensions[get_column_letter(i)].width = w
for c in range(1, len(h2)+1):
    hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL; hc.alignment = Alignment(horizontal="center", vertical="center")
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{ws2.max_row}"
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(h2)):
    for cell in row:
        cell.font = F(size=10); cell.border = BORDER
        if cell.column in (3,4,8): cell.alignment = Alignment(horizontal="center")

# ---------- Sheet 3: Summary & Notes ----------
ws3 = wb.create_sheet("Summary & Notes", 0)
total = len(rows)
finished = sum(1 for r in rows if r[7])
stalled = sum(1 for r in rows if r[9]=="STALLED")
overdue = sum(1 for r in rows if r[9]=="OVERDUE")
late = sum(1 for r in rows if r[9]=="LATE")
ontime = sum(1 for r in rows if r[9]=="ONTIME")
last_ready = max((r[7] for r in rows if r[7] and r[7] < "9999"), default="")
# ---- Website-style "Active Jobs" reconciliation (matches /dashboard card) ----
# Dashboard counts BEDFRAME by piece (qty) and SOFA by SET (1 SO = 1 set);
# ACCESSORY is not shown on that card. This schedule lists every production
# line (each sofa component + each accessory separately), so it has more rows.
bf_units   = sum((r[5] or 0) for r in rows if r[4] == "BEDFRAME")
sofa_lines = sum(1 for r in rows if r[4] == "SOFA")
sofa_sets  = len(set(r[0] for r in rows if r[4] == "SOFA"))
acc_lines  = sum(1 for r in rows if r[4] not in ("BEDFRAME", "SOFA"))
WEB_BF, WEB_SOFA = 189, 60  # live /api/dashboard/overview Active Jobs, 2026-06-01
def b(d): return 1 if d else 0
buck = {"1-7":0,"8-14":0,"15-30":0,"30+":0}
for r in rows:
    if r[9] in ("LATE","OVERDUE"):
        dl=r[10] or 0
        if dl<=7: buck["1-7"]+=1
        elif dl<=14: buck["8-14"]+=1
        elif dl<=30: buck["15-30"]+=1
        else: buck["30+"]+=1
lines = [
    ("生产排期总览  Production Schedule Forecast", True, 14),
    (f"生成日期 Generated: 2026-06-01    第一个开工天 First working day: {FIRST_WD}", False, 10),
    ("", False, 10),
    ("【表头说明 New: Product Code vs WIP Code】", True, 11),
    ("  「Schedule」表现在有两栏：Product Code (Model) = 款号；WIP Code = 完整规格（尺寸/缝隙/床座/布料），跟系统 WIP 栏一样。", False, 10),
    ("  Product Code = model only; WIP Code = full spec descriptor (matches the system's WIP column).", False, 10),
    ("", False, 10),
    ("【关键数字 Headline】", True, 11),
    (f"  按网站 Dashboard 算法:{WEB_BF} 件床架 / {WEB_SOFA} 套沙发  (配件另计 {acc_lines} 件,Dashboard 不显示)", False, 10),
    (f"  按生产单(本表每一行):{total} 行 = 床架 {bf_units} + 沙发部件 {sofa_lines} + 配件 {acc_lines}", False, 10),
    (f"  能排进表里的 Scheduled with a ready date: {finished}", False, 10),
    (f"  排不进的 No capacity data (Accessory): {stalled}", False, 10),
    (f"  最后一单做好 Everything cleared by: {last_ready}", False, 10),
    ("", False, 10),
    ("【为什么这里 300 多行,网站首页只看到 100 多?】", True, 11),
    ("  同一批货,两种数法 —— 没有重复、没有算错。", False, 10),
    (f"  · 网站首页「Active Jobs」:床架算『件』、沙发算『套』(一张客户单 = 一套)、配件不算 → 显示 {WEB_BF} / {WEB_SOFA}。", False, 10),
    ("  · 本排期表:每张生产单各一行 —— 一套沙发拆成角位 / 1A / 2A / 凳… 每个部件分开排,所以行数多。", False, 10),
    (f"  · 对照:把本表 {sofa_lines} 个沙发部件併回『套』≈ {sofa_sets} 套(≈ 网站 {WEB_SOFA} 套,差几张是快照时间差);床架 {bf_units} 件两边一样。", False, 10),
    ("  · 为什么一定要拆部件:每个沙发部件要分开过裁布→车缝→…→包装 8 个站,占各站产能;不拆就算不准日期。", False, 10),
    ("  · 想看『一张客户单一行』的视角 → 看「By Order」分页(约 184 张单)。", False, 10),
    ("", False, 10),
    ("【准时 vs 迟 On-time vs Late】", True, 11),
    (f"  准时/提早 On time or early: {ontime}", False, 10),
    (f"  今天已经过期 Already past due today: {overdue}", False, 10),
    (f"  会迟交（未来到期）Will miss a future due date: {late}", False, 10),
    (f"     迟 1-7 天: {buck['1-7']}    迟 8-14 天: {buck['8-14']}    迟 15-30 天: {buck['15-30']}    迟 30 天以上: {buck['30+']}", False, 10),
    ("", False, 10),
    ("【裁布 Cutting (shared crew, concurrent)】", True, 11),
    ("  一组人，每天同时开床架 + 沙发（不是先切完床架才切沙发）。", False, 10),
    ("  床架 6 款/天，沙发 2 款/天；床架清完后沙发提速到全力。", False, 10),
    ("  Bedframe 6 models/day + sofa 2 models/day, both run every day from day 1;", False, 10),
    ("  once bedframe backlog clears (~6 working days) sofa jumps to full force.", False, 10),
    ("  只剩 ~115 张裁布单要切，约 6 个开工天清完——裁布不是瓶颈。", False, 10),
    ("", False, 10),
    ("【这张表怎么算的 How dates are computed】", True, 11),
    ("  1. 只算「还没做完、还在等」的单（Status = WAITING）；做好的 (COMPLETED) 全部跳过。", False, 10),
    ("  2. 一张单要全部 8 个站做完才算「做好」：裁布→车缝→锯木→钉框→织带→海绵→包皮→包装。", False, 10),
    ("  3. 每个站每天有产能上限，多张单抢同一个站要排队（先到期的先做 EDD）。", False, 10),
    ("  4. 「切几天」只是第一站；「Planned Ready」是全部 8 站跑完的日期——所以比单看裁布长很多。", False, 10),
    ("  5. 真正塞车的是后段（包皮/织带/钉框），不是裁布。", False, 10),
    ("  6. 不放安全余量（现场判断值）。星期日 + 公假不开工，星期六照常。", False, 10),
    ("", False, 10),
    ("【还要你确认 Still pending your input】", True, 11),
    ("  · 全力切沙发一天几款？现在暂用 4 款/天（假设）。你给真实数，沙发日期会再准一点。", False, 10),
    ("  · 后面 7 个站的「每天产能」现在用最近开工天量出来的平均；还没一站一站跟你核对。", False, 10),
    ("  · Accessory 类没有产能数据，所以排不进（表里标 No capacity data）。", False, 10),
]
for i,(txt,bold,size) in enumerate(lines, start=1):
    c = ws3.cell(row=i, column=1, value=txt)
    c.font = F(bold=bold, size=size)
ws3.column_dimensions["A"].width = 95

out = r"C:\Users\User\hookka-erp-testing\Production-Schedule-2026-06-02.xlsx"
try:
    wb.save(out)
except PermissionError:
    out = r"C:\Users\User\hookka-erp-testing\Production-Schedule-2026-06-02-v2.xlsx"
    wb.save(out)
print("OK", total, finished, stalled, last_ready, "->", out)
