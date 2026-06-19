# =============================================================================
# FABRIC CUTTING — day-by-day cut calendar  (read-only, no ERP writes)
# =============================================================================
# WHAT THIS IS
#   An actual calendar date stamped on every Fab Cut that still has to be done.
#
#   SCOPE FIX (2026-06-01): schedule ONLY cards whose cutting is NOT done yet
#   (cardStatus == WAITING). The earlier version pulled every FAB_CUT card incl.
#   148 already-cut ones — wrong. Real to-cut backlog = 114 cards.
#   On-hold orders are excluded.
#
#   CAPACITY (docs/PRODUCTION-PLANNING-LOGIC.md STEP 4, confirmed):
#       gate = MODELS per day (款/天). BEDFRAME 6/day, SOFA 2/day.
#       a setup = one (model + size) — that's how productCode-level "6/day" was
#       measured, so we batch by model+size (1007-SS, 1007-Q are separate cuts).
#       big batch takes extra slots: slots = ceil(sets / setupCap)  (BF 5 / SOFA 3)
#
#   PRIORITY: by CUSTOMER DD (the real promise to the customer) — owner confirmed
#   2026-06-01. Earliest customer DD cut first; inside the same deadline week,
#   same model + same size batched together (K/Q/SS/S = 6FT/5FT/3.5FT/3FT), so
#   the cutter changes layout as little as possible.
#
#   LANES RUN IN PARALLEL: Bedframe / Sofa / Accessory are different people, so
#   sofa cutting starts day 1 too (sofa sewing is never starved). The calendar is
#   shown interleaved by DATE so you see what each lane cuts on the same day.
#
#   Calendar starts 2026-06-02 (6/1 holiday), skips Sundays + holidays.
#   READ-ONLY. Nothing written to the ERP.
#
#   ⚠️ ACCESSORY cutting rate was never confirmed — provisional 4/day (flagged).
# =============================================================================
import json, os, math
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
# FRESH LIVE PULL (2026-06-01): the old sched_dept_full.json snapshot was stale
# (it marked already-cut pillows as WAITING). This reads a fresh export of the
# live /api/production-orders FAB_CUT WAITING cards, pulled straight from the
# logged-in ERP. Pipe-delimited (' ¦ ') so model/wipLabel spaces survive.
TXT = os.path.join(HERE, "fabcut_live.txt")
COLS = ["po", "soNo", "cust", "model", "code", "cat", "size",
        "fabric", "wipLabel", "wipQty", "cdd", "edd", "orderStatus"]
full = {"FAB_CUT": []}
with open(TXT, encoding="utf-8") as fh:
    for ln, line in enumerate(fh):
        line = line.rstrip("\n")
        if not line or ln == 0:        # skip header
            continue
        parts = [p.strip() for p in line.split("¦")]
        if len(parts) != len(COLS):
            continue
        d = dict(zip(COLS, parts))
        full["FAB_CUT"].append({
            "cardStatus": "WAITING",   # the live export is WAITING-only
            "cat": d["cat"], "custName": d["cust"], "po": d["po"], "soNo": d["soNo"],
            "model": d["model"], "size": d["size"], "colour": d["fabric"],
            "wipLabel": d["wipLabel"], "wipQty": d["wipQty"],
            "cdd": d["cdd"], "edd": d["edd"], "orderStatus": d["orderStatus"],
        })

# ---- pull WAITING, schedulable FAB_CUT cards only ----
#  Drop ON_HOLD and CANCELLED orders (e.g. SO-2605-052 is CANCELLED).
#  Schedule ALL 114 still-to-cut cards: BEDFRAME + SOFA + ACCESSORY (pillows).
#  The fresh live pull shows these 9 pillows genuinely WAITING (the earlier
#  stale snapshot wrongly marked them COMPLETED). A pillow must be cut near its
#  parent SO so the order can ship, so it belongs on this calendar too.
src = [r for r in full.get("FAB_CUT", [])
       if r.get("cardStatus") == "WAITING"
       and r.get("orderStatus") not in ("ON_HOLD", "CANCELLED")
       and r.get("cat") in ("BEDFRAME", "SOFA", "ACCESSORY")]

START = date(2026, 6, 2)            # first working day (6/1 is a holiday)
HOLIDAYS = {date(2026, 6, 1)}
FAR = "9999-99-99"
WINDOW_DAYS = 7

# Bedframe + sofa share ONE pool of cutters (POOL_CAP cuts/day total). Each lane has
# its own per-day ceiling, but their combined cuts can't exceed POOL_CAP. Bedframe is
# scheduled first (priority); sofa then fills whatever pool capacity is left that day —
# so on days bedframe runs light, sofa automatically speeds up instead of starving.
# Accessory (pillows) uses its own separate capacity.
POOL_CAP = 8                                            # bedframe+sofa cuts/day, FULL
LANE_CAP = {"BEDFRAME": 6, "SOFA": 6, "ACCESSORY": 8}   # per-lane ceiling within pool
# Reserve-capacity tiers (Wei Siang 2026-06-01): new orders walk in every day, so the
# near days must be packed full & accurate while the mid days keep spare slots free for
# tomorrow's incoming work, and the far days stay loose (that data isn't firm yet).
# Each tuple = (up to this WORKING-day index from day 1, pool cuts/day allowed).
#   days 1-3  → 7   (firm orders, pack tight & accurate — normal "full" is 6-7, not 8)
#   days 4-7  → 6   (keep ~1 slot/day free for incoming orders)
#   days 8+   → 5   (loose; leave plenty of room, that data isn't firm yet)
# A single cut already holds up to 8 sets, so 6-7 cuts/day is plenty of throughput;
# 8 is only an occasional rush ceiling (POOL_FULL), kept just as buffer headroom.
RESERVE_TIERS = [(3, 7), (7, 6), (10 ** 9, 5)]
POOL_FULL = 8                                           # occasional rush ceiling (buffer calc)
# Sofa must get a guaranteed minimum of cuts/day so its urgent orders aren't starved by
# bedframe hogging the shared pool (Wei Siang: "沙发一定要排一天两三个"). Bedframe is held
# to (pool - SOFA_MIN) so this many slots are always kept for sofa; on days bedframe runs
# light, sofa borrows the freed slots up to its LANE_CAP ceiling and speeds up further.
SOFA_MIN = 3
# max sets in ONE cut (Wei Siang: "cut 一次最多 5-8 个"). Above this the batch is
# automatically broken into 2 / 3 / ... cuts.
SETUP_CAP = {"BEDFRAME": 8, "SOFA": 3, "ACCESSORY": 8}  # sets per cut
LANE_ORDER = ["BEDFRAME", "SOFA", "ACCESSORY"]

# --- planner knobs (Wei Siang confirmed 2026-06-01) ---------------------------
# 1) PULL_WINDOW: only merge same-cut orders whose customer DDs fall within this
#    many days of each other. Orders due further out are NOT pulled forward — they
#    wait for a later batch near their own deadline, so we don't cut too early and
#    flood storage. (Middle ground between "merge everything" and "cut separately".)
PULL_WINDOW = 10
# 2) Shared cutting capacity: bedframe & sofa draw from the same POOL_CAP cutters/day
#    (see above). No fixed sofa rate — sofa dynamically takes whatever the bedframe
#    schedule leaves free each day, up to its own LANE_CAP ceiling.
# 3) CHUNK_LEAD: when a batch is too big for one cut and must be split into 2/3
#    cuts, the FIRST cut (urgent cards) goes ASAP; each LATER cut is placed only
#    this many working days before its own cards' earliest customer DD — so the
#    second cut isn't jammed next to the first and fabric isn't cut too early.
CHUNK_LEAD = 3
# 4) MODEL_LEAD (middle-ground knob, Wei Siang "不能取中间点嘛"): the schedule used
#    to place every cut purely by its OWN deadline, which scattered one model across
#    the whole week. Now cuts of the SAME model are kept together as a block, placed
#    around the model's own earliest deadline. MODEL_LEAD is how many working days
#    before that earliest deadline the block may start — bigger than CHUNK_LEAD so a
#    model's later cuts pull in NEXT TO the first cut (concentrated) instead of
#    waiting near their own far date (scattered). Middle ground = neither cut-all-
#    early (storage flood) nor cut-each-piece-just-in-time (scattered).
MODEL_LEAD = 6
# 5) CLUSTER_LANES: which lanes use the model-block (concentrated) ordering above.
#    Bedframe benefits (many sizes per model, plenty of daily capacity → models pack
#    onto consecutive days). Sofa does NOT: it is capacity-bound (2 cuts/day until
#    bedframe clears) and inherently scattered (each sofa config is its own layout),
#    so clustering there would only delay urgent cards — it stays on pure JIT order.
CLUSTER_LANES = {"BEDFRAME", "SOFA"}
LANE_CONFIRMED = {"BEDFRAME": True, "SOFA": True, "ACCESSORY": False}  # pillow rate TBC
CN = {"BEDFRAME": "床架 Bedframe", "SOFA": "沙发 Sofa", "ACCESSORY": "配件 Pillow"}

# size sort order so a model's sizes sit together K→Q→SS→S
SIZE_RANK = {"6FT": 0, "5FT": 1, "4FT": 2, "3.5FT": 3, "3FT": 4}
def size_rank(s):
    return SIZE_RANK.get((s or "").strip(), 50)

def is_offday(d):  return d.weekday() == 6 or d in HOLIDAYS
def next_workday(d):
    while is_offday(d): d += timedelta(days=1)
    return d
def step_workday(d): return next_workday(d + timedelta(days=1))
def workday_index(d):                    # 1 = first working day (START), off-days skipped
    i, x = 0, next_workday(START)
    while x <= d:
        i += 1; x = step_workday(x)
    return i
def pool_cap(d):                         # that working day's total cutting budget (reserve tiers)
    i = workday_index(d)
    for upto, cap in RESERVE_TIERS:
        if i <= upto: return cap
    return RESERVE_TIERS[-1][1]
def back_workday(d, n):                 # n working days BEFORE d (skips off-days)
    while n > 0:
        d -= timedelta(days=1)
        while is_offday(d): d -= timedelta(days=1)
        n -= 1
    return d
def parse(s):
    try:
        y, m, dd = str(s)[:10].split("-"); return date(int(y), int(m), int(dd))
    except Exception: return None
def base_model(m):
    m = (m or "").strip(); return m.split("-")[0] if m else "(none)"

THIS_MON = START - timedelta(days=START.weekday())
def window_key(cdd):
    dt = parse(cdd)
    if dt is None: return 9999
    if dt < THIS_MON: return 0
    wk_mon = dt - timedelta(days=dt.weekday())
    return ((wk_mon - THIS_MON).days // WINDOW_DAYS) + 1
def window_label(wk):
    if wk == 0:  return "0) overdue / this week"
    if wk == 9999: return "9) no date"
    s = THIS_MON + timedelta(days=(wk - 1) * WINDOW_DAYS)
    return f"{wk}) {s.strftime('%d %b')}-{(s+timedelta(days=6)).strftime('%d %b')}"

def _fmt_cdd(s):
    d = parse(s)
    return d.strftime("%d %b") if d else "no date"

def deadline_label(g):
    # earliest customer DD in this cut; show range only if it spans dates
    lo, hi = g.get("cdd_min", ""), g.get("cdd_max", "")
    if not lo and not hi:
        lbl = "no date"
    elif lo and hi and lo != hi:
        lbl = f"{_fmt_cdd(lo)} → {_fmt_cdd(hi)}"
    else:
        lbl = _fmt_cdd(lo or hi)
    pa, pn = g.get("part", (1, 1))
    if pn > 1: lbl += f"  ({pa}/{pn}刀)"
    return lbl

# the true "cut" identity = full configuration (wipLabel first segment) — this is
# what needs its own cutting layout. 5531-L(LHF)+2A(RHF) vs 5531-2A(LHF)+L(RHF)
# are DIFFERENT cuts even though both are model 5531; 1013-(K) vs 1013-(Q) too.
def config_of(r):
    wl = (r.get("wipLabel") or r.get("model") or "").strip()
    return wl.split("|")[0].strip() or "(none)"

# normalize a card to the fields we use
def norm(r):
    cfg = config_of(r)
    return {"po": r.get("po"), "cust": r.get("custName"), "model": r.get("model"),
            "config": cfg, "bmodel": base_model(cfg), "size": r.get("size") or "",
            "fabric": r.get("colour") or "", "label": r.get("wipLabel") or r.get("model"),
            "sets": max(1, int(r.get("wipQty") or 1)),
            "cdd": r.get("cdd"), "edd": r.get("edd"), "cat": r.get("cat")}

# ---- build cut-groups per lane: merge same cut identity, but only within a
#      customer-DD window so we don't cut far-future orders too early ------------
#  A cut = one cutting layout (same config + same size). We merge every card of
#  that cut whose customer DDs sit within PULL_WINDOW days of each other into ONE
#  batch, so the cutter sets that layout up once for the whole near-term demand.
#  Orders of the same cut due much later start a SEPARATE batch placed near their
#  own deadline — that keeps cut fabric from piling up in storage too early.
def _split_by_window(items):
    """Group a cut-identity's cards into sub-batches by customer-DD PROXIMITY.

    Old logic measured every card's distance from the batch's EARLIEST card, so two
    cards only 2 days apart (06-10 & 06-12) could land in different batches just
    because both were >PULL_WINDOW from the anchor — that wrongly un-matched cards
    that should cut together. New logic keeps all cards in ONE cluster and only
    splits a cluster when it spans more than PULL_WINDOW days, always cutting it at
    its single LARGEST internal gap. So neighbours stay merged and the split lands at
    the natural break (e.g. the 7-day gap between an urgent pair and a later cluster),
    never between two near-dated orders. Undated cards ride in the last batch."""
    dated = sorted([it for it in items if it.get("cdd")], key=lambda it: it["cdd"])
    undated = [it for it in items if not it.get("cdd")]
    if not dated:
        return [undated] if undated else [[]]

    def span_days(cluster):
        return (parse(cluster[-1]["cdd"]) - parse(cluster[0]["cdd"])).days

    clusters = [dated]
    # repeatedly break the widest over-window cluster at its biggest day-gap
    while True:
        wide = next((c for c in clusters if span_days(c) > PULL_WINDOW), None)
        if not wide:
            break
        gaps = [((parse(wide[i + 1]["cdd"]) - parse(wide[i]["cdd"])).days, i)
                for i in range(len(wide) - 1)]
        _, gi = max(gaps)                 # split right after the largest gap
        clusters.remove(wide)
        clusters.append(wide[:gi + 1]); clusters.append(wide[gi + 1:])
    clusters.sort(key=lambda c: c[0]["cdd"])
    clusters[-1].extend(undated)
    return clusters

def _pe(s):                       # parse customer DD to a sortable date (far if none)
    return parse(s) or date(9999, 12, 31)

def build_groups(lane_rows, lane):
    cap = SETUP_CAP[lane]
    buckets = defaultdict(list)   # (config, size) -> items
    for r in lane_rows:
        buckets[(r["config"], r["size"])].append(r)
    subs = []
    for (cfg, sz), items in buckets.items():
        for b in _split_by_window(items):
            e = min([it["cdd"] for it in b if it.get("cdd")] or [FAR])
            subs.append({"cfg": cfg, "sz": sz, "items": b, "bmodel": b[0]["bmodel"], "earliest": e})
    groups = []
    # MIDDLE GROUND (Wei Siang "不能取中间点嘛"): keep each MODEL together as a block.
    # Block order = the model's OWN earliest customer DD, so a far-future model never
    # steals early capacity from a sooner-due one (preserves the JIT fairness). But
    # WITHIN that order all of a model's cuts are processed back-to-back, so the day-
    # pool packs them onto consecutive days instead of sprinkling the model across the
    # whole week. Inside a model, sizes due together sit adjacent → same cutting day.
    model_earliest = {}
    for s in subs:
        m = s["bmodel"]
        if m not in model_earliest or _pe(s["earliest"]) < _pe(model_earliest[m]):
            model_earliest[m] = s["earliest"]
    def _order(s):
        if lane in CLUSTER_LANES:
            # concentrate: model block first (by the model's own earliest date), then
            # the model's cuts back-to-back. Keeps a model on consecutive days.
            return (_pe(model_earliest[s["bmodel"]]), s["bmodel"],
                    _pe(s["earliest"]), size_rank(s["sz"]), s["cfg"])
        # capacity-bound / inherently-scattered lanes (sofa: different configs are
        # different cuts) stay on pure own-deadline JIT priority so an urgent cut is
        # never delayed behind a less-urgent model's block.
        return (_pe(s["earliest"]), s["bmodel"], size_rank(s["sz"]), s["cfg"])
    for s in sorted(subs, key=_order):
        # earliest-due cards first, so a too-big batch splits into cut #1 (urgent
        # half) + cut #2 (later half), each <= cap sets
        items = sorted(s["items"],
                       key=lambda it: (it.get("cdd") or FAR, it.get("fabric") or "", it.get("po") or ""))
        chunks, cur, cursum = [], [], 0
        for it in items:
            if cur and cursum + it["sets"] > cap:
                chunks.append(cur); cur, cursum = [], 0
            cur.append(it); cursum += it["sets"]
        if cur: chunks.append(cur)
        for ci, ch in enumerate(chunks):
            sets = sum(it["sets"] for it in ch)
            cdds = sorted(it["cdd"] for it in ch if it.get("cdd"))
            ch_min = cdds[0] if cdds else ""
            # Every cut is JIT'd to its OWN earliest customer DD: it may start at most
            # `lead` working days before that date. Genuinely-urgent cuts (DD now/past)
            # fall to ASAP automatically (their back-dated floor lands before START).
            # This fixes far-due batches being force-cut on day 1 just because they were
            # the first chunk of their own batch. A too-big batch's split chunks share a
            # similar floor, so they sit on consecutive days near that date — not pulled
            # to day 1. lead is looser for cluster lanes so a model's cuts pull together.
            lead = MODEL_LEAD if lane in CLUSTER_LANES else CHUNK_LEAD
            if not ch_min:
                floor = next_workday(START)
            else:
                floor = max(next_workday(START), back_workday(_pe(ch_min), lead))
            groups.append({"config": s["cfg"], "bmodel": s["bmodel"], "size": s["sz"],
                           "items": ch, "sets": sets, "slots": max(1, math.ceil(sets / cap)),
                           "cdd_min": ch_min, "cdd_max": cdds[-1] if cdds else "",
                           "floor": floor, "part": (ci + 1, len(chunks))})
    return groups

def schedule_lane(groups, cap_fn, used=None):
    """Day-pool scheduler. cap_fn(day) = that day's cut budget for THIS lane. Passing a
    shared `used` counter lets two lanes draw from one pool: bedframe is scheduled
    first into `used`, then sofa's cap_fn reads bedframe's per-day usage so sofa only
    fills the pool capacity bedframe left free. A batch that needs several cuts (>1,
    because it's bigger than one cut can take) is spread over SEPARATE days — never two
    cuts of the same cut on the same day — so we don't cut a big pile all at once.
    Different cuts (incl. different sizes of the same model) still share a day to keep
    the model moving together for wood-cutting / sewing."""
    if used is None:
        used = defaultdict(int)
    # process in priority order; each group grabs its cuts on the earliest days
    # that still have room and that the group isn't already using that day
    for g in groups:
        n = g["slots"]; g["days"] = []; d = g.get("floor") or next_workday(START); seen = set()
        while len(g["days"]) < n:
            if d not in seen and used[d] < cap_fn(d):
                used[d] += 1; g["days"].append((d, 1)); seen.add(d)
            d = step_workday(d)
        g["start"] = g["days"][0][0]; g["end"] = g["days"][-1][0]
    return groups

# Shared cutter pool: bedframe + sofa together can't exceed POOL_CAP cuts/day.
# Bedframe is scheduled FIRST (priority) into the shared day-counter; sofa is then
# scheduled into the SAME counter, so on any day sofa only takes the pool capacity
# bedframe left free (up to sofa's own ceiling). No fixed BF_END handoff — sofa speeds
# up automatically the moment bedframe runs light.
def cap_for(lane, d):
    """That day's cut budget for a lane (used for scheduling + the By-Day display).
    The day's total bedframe+sofa is held to pool_cap(d), which shrinks on the mid/far
    days so spare slots stay open for orders that walk in later."""
    if lane == "SOFA":
        # sofa fills the pool up to that day's budget, never above its own ceiling
        return min(pool_cap(d), bf_used.get(d, 0) + LANE_CAP["SOFA"])
    if lane == "BEDFRAME":
        # hold bedframe back so SOFA_MIN slots are always reserved for sofa; sofa can
        # still borrow back any of these on days bedframe doesn't need them
        return min(LANE_CAP["BEDFRAME"], max(1, pool_cap(d) - SOFA_MIN))
    return LANE_CAP[lane]

pool_used = defaultdict(int)
_bf_rows = [norm(r) for r in src if r.get("cat") == "BEDFRAME"]
_bf_groups = build_groups(_bf_rows, "BEDFRAME")
schedule_lane(_bf_groups, lambda d: cap_for("BEDFRAME", d), used=pool_used)
bf_used = dict(pool_used)                 # bedframe-only cuts per day (snapshot)

lane_groups = {}
for lane in LANE_ORDER:
    if lane == "BEDFRAME":
        gs = _bf_groups
    elif lane == "SOFA":
        lr = [norm(r) for r in src if r.get("cat") == "SOFA"]
        gs = build_groups(lr, "SOFA")
        schedule_lane(gs, lambda d: cap_for("SOFA", d), used=pool_used)  # shares pool
    else:                                  # accessory = its own separate capacity
        lr = [norm(r) for r in src if r.get("cat") == lane]
        gs = build_groups(lr, lane)
        schedule_lane(gs, (lambda L: (lambda d: cap_for(L, d)))(lane))
    for i, g in enumerate(gs, 1): g["cut"] = f"{lane[0]}{i}"
    lane_groups[lane] = gs

BF_END = max([g["end"] for g in _bf_groups], default=next_workday(START))

# ===================== styling =====================
ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
DATE_FILL = PatternFill("solid", fgColor="1F6F54"); DATE_FONT = F(bold=True, color="FFFFFF", size=10)
BF_FILL = PatternFill("solid", fgColor="DDEBF7")
SOFA_FILL = PatternFill("solid", fgColor="FFF2CC")
ACC_FILL = PatternFill("solid", fgColor="FCE4D6")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
LANE_FILL = {"BEDFRAME": BF_FILL, "SOFA": SOFA_FILL, "ACCESSORY": ACC_FILL}

wb = Workbook()

# ===================== Sheet 1: Cut Calendar (interleaved by DATE) ===========
ws = wb.active; ws.title = "Cut Calendar"
headers = ["Cut Date 裁料日", "Lane 队", "Cut #", "Model 款", "Size", "SO / PO",
           "Customer", "Item (one set/row)", "Fabric", "Sets 套", "Slots 款位",
           "Customer DD", "Our Expected DD", "Batch Deadline 这批最早客期"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# flatten all groups, sort by (start date, lane order)
allg = []
for li, lane in enumerate(LANE_ORDER):
    for g in lane_groups[lane]:
        allg.append((g["start"], li, lane, g))
allg.sort(key=lambda t: (t[0], t[1], t[3]["cut"]))

cur_date = None
for sdate, li, lane, g in allg:
    if sdate != cur_date:
        cur_date = sdate
        ws.append([sdate.strftime("%Y-%m-%d  %a") + f"   (Day {(sdate-next_workday(START)).days+1})"])
        rr = ws.max_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
        ws.cell(row=rr, column=1).fill = DATE_FILL; ws.cell(row=rr, column=1).font = DATE_FONT
    dstr = ""
    if g["start"] != g["end"]:
        dstr = "→ " + g["end"].strftime("%m-%d %a")
    for i, r in enumerate(g["items"]):
        ws.append([dstr if i == 0 else "", CN[lane] if i == 0 else "",
                   g["cut"] if i == 0 else "", g["config"] if i == 0 else "",
                   g["size"] if i == 0 else "", r["po"], r["cust"], r["label"], r["fabric"],
                   r["sets"], g["slots"] if i == 0 else "", r["cdd"], r["edd"],
                   (deadline_label(g) if i == 0 else "")])
        rr = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = LANE_FILL[lane]; ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).font = F(size=9)
        if parse(r["cdd"]) and parse(r["cdd"]) < THIS_MON:
            ws.cell(row=rr, column=12).fill = OVERDUE_FILL
widths = [22, 16, 6, 12, 8, 15, 18, 38, 13, 6, 7, 12, 14, 20]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ===================== Sheet 2: By Day (one row per lane-day) ================
ws2 = wb.create_sheet("By Day")
h2 = ["Date 日期", "Day", "Lane 队", "Models that day 当天裁的款 (model-size×slots)",
      "Slots", "Orders", "Cap"]
ws2.append(h2)
perday = defaultdict(lambda: defaultdict(list))
for lane in LANE_ORDER:
    for g in lane_groups[lane]:
        for d, sl in g["days"]: perday[lane][d].append((g, sl))
rowi = 0
for d in sorted({d for lane in LANE_ORDER for d in perday[lane]}):
    for lane in LANE_ORDER:
        if d not in perday[lane]: continue
        entries = perday[lane][d]; rowi += 1
        models = ", ".join(f"{g['config']} ({g['size']})" + (f"×{sl}" if sl > 1 else "")
                           + (" (cont.)" if g["start"] != g["end"] and d != g["start"] else "")
                           for g, sl in entries)
        norders = sum(len(g["items"]) for g, sl in entries if d == g["start"])
        # display cap = the lane's actual available cuts that day (sofa = the pool
        # capacity bedframe left free, so it visibly rises as bedframe runs light)
        disp_cap = cap_for(lane, d)
        if lane == "SOFA":
            disp_cap = min(LANE_CAP["SOFA"], pool_cap(d) - bf_used.get(d, 0))
        ws2.append([d.strftime("%Y-%m-%d %a"), (d-next_workday(START)).days+1, CN[lane],
                    models, sum(sl for g, sl in entries), norders, disp_cap])
        rr = ws2.max_row
        for c in range(1, len(h2)+1):
            ws2.cell(row=rr, column=c).fill = LANE_FILL[lane]; ws2.cell(row=rr, column=c).border = BORDER
            ws2.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h2)+1):
    hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 16, 64, 7, 8, 6], 1): ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ===================== Sheet 0: Summary & Notes =============================
ws0 = wb.create_sheet("Summary & Notes", 0)
spans = {}
for lane in LANE_ORDER:
    gs = lane_groups[lane]
    if gs:
        spans[lane] = (min(g["start"] for g in gs), max(g["end"] for g in gs),
                       len(gs), sum(len(g["items"]) for g in gs))
ncards = len(src); bydist = {l: sum(1 for r in src if r.get("cat") == l) for l in LANE_ORDER}
notes = [
    ("裁料 逐日排程  Fabric Cutting — Day-by-Day Schedule (修正版)", True, 14),
    ("生成 2026-06-01,资料从你登入的 ERP 现场拉的最新版。只排『还没裁(WAITING)』的卡,", False, 10),
    ("不含已裁好、不含 on-hold、不含已取消。只读,没改系统。", False, 10),
    ("", False, 10),
    ("【范围】跟着页面那 114 row", True, 11),
    (f"  还要裁的卡 = {ncards} 张  (床架 {bydist['BEDFRAME']} / 沙发 {bydist['SOFA']} / 配件 {bydist['ACCESSORY']})。", False, 10),
    ("  · 踢掉已裁好(COMPLETED)的卡 + on-hold + 已取消(SO-2605-052)。", False, 10),
    ("  · 配件 pillow 这次全排进去:现场资料显示它们是真的还没裁;", False, 10),
    ("    pillow 跟它那张 SO 的客期一样排,出货才配得齐。", False, 10),
    ("", False, 10),
    ("【什么算『一个 cut』】", True, 11),
    ("  · 看完整配置,不是看型号:5531-L(LHF)+2A(RHF) 和 5531-2A(LHF)+L(RHF) 是两个不同 cut。", False, 10),
    ("  · 床架 1013 的 K/Q/SS/S(6FT/5FT/3.5FT/3FT)也是不同 cut。", False, 10),
    ("  · 同配置 + 同尺寸 才併一起裁(布料不同没关系,排在一批里)。", False, 10),
    ("", False, 10),
    ("【排法】同一个 cut 能併就併,但太远的不硬拉(你要的中间点)", True, 11),
    (f"  · 同配置 + 同尺寸:客期在 {PULL_WINDOW} 天内的单,拉来一起裁。", False, 10),
    (f"    客期超过 {PULL_WINDOW} 天的同款,留到接近它自己客期才裁 —— 不太早裁、不堆库存。", False, 10),
    (f"  · 一次 cut 最多 {SETUP_CAP['BEDFRAME']} 个;数量多到超过就自动分两次、三次裁。", False, 10),
    (f"    按客期分:急的进第一刀(尽早裁),迟的进第二刀,第二刀排在它自己客期前 {CHUNK_LEAD} 天,", False, 10),
    ("    不挤在第一刀隔天 —— 自动隔开,也不会太早裁堆库存。(看『这批最早客期』栏的 1/2刀、2/2刀)", False, 10),
    ("  · priority 看每个 cut 自己最早的 Customer DD —— 客期远的不会因为同款有急单就被拉前、抢产能。", False, 10),
    ("  · 客期相近的同款不同 size 自然排同一天(WC 不看 size,FC/sewing/WC 同款一起出,后段接得顺);", False, 10),
    ("    客期差远的同款就各自靠近自己客期裁,不硬绑同天。", False, 10),
    (f"  · 床架沙发共用裁料人手,每天合起来:前 3 个工作天排满 {RESERVE_TIERS[0][1]} 刀(单子最准),", False, 10),
    (f"    第 4-7 天只排 {RESERVE_TIERS[1][1]} 刀、留位子给临时进来的单,第 8 天起更松({RESERVE_TIERS[2][1]} 刀)。", False, 10),
    ("    床架优先排,床架轻的日子沙发自动补上去,不用等床架全部裁完。", False, 10),
    (f"  · 拉前限度 {PULL_WINDOW} 天可调:你觉得要更短(库存更少)或更长(裁更少次),讲一声我改。", False, 10),
    ("", False, 10),
    ("【怎么看】", True, 11),
    ("  · Cut Calendar:按日期排,绿色横条 = 那一天;同一天床架/沙发各裁什么一眼看到。", False, 10),
    ("  · By Day:一天一队一行,看每队当天裁哪几个 model-size。", False, 10),
    ("", False, 10),
    ("【排出来的跨度】", True, 11),
]
for lane in LANE_ORDER:
    if lane in spans:
        a, b, ng, ni = spans[lane]
        tag = "" if LANE_CONFIRMED[lane] else "  (产能待确认)"
        notes.append((f"  {CN[lane]:<16} {a.strftime('%m-%d %a')} → {b.strftime('%m-%d %a')}"
                      f"   共 {ng} 批 / {ni} 张{tag}", False, 10))
notes += [("", False, 10),
          ("  注:只读线上资料做的 planning,还没写进 ERP。你确认排法 OK,我再写进系统。", False, 10)]
for i, (t, b, s) in enumerate(notes, 1):
    ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws0.column_dimensions["A"].width = 98

out = r"C:\Users\User\hookka-erp-testing\Cutting-Daily-Schedule-2026-06-02.xlsx"
try: wb.save(out)
except PermissionError:
    out = r"C:\Users\User\hookka-erp-testing\Cutting-Daily-Schedule-2026-06-02-v2.xlsx"
    wb.save(out)
print("OK ->", out, " WAITING cards:", ncards, bydist)
for lane in LANE_ORDER:
    if lane in spans:
        a, b, ng, ni = spans[lane]
        print(f"  {lane:<10} {a} -> {b}  ({ng} batches, {ni} cards)")
