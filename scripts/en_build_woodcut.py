# =============================================================================
# WOOD CUTTING — English report (presentation-only translation of
# build_woodcut_daily_xlsx.py). Imports the ORIGINAL module so all dates/sets
# are byte-identical; only labels are English.
# =============================================================================
import os, importlib.util
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = r"C:\Users\User\Desktop\Production-Schedule-2026-06-02-EN"


def _load(name, fname):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, fname))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


wd = _load("wood_src", "build_woodcut_daily_xlsx.py")

LANE_ORDER = wd.LANE_ORDER
lane_units = wd.lane_units
units = wd.units
CAP_SETS = wd.CAP_SETS
WOOD_HANDOFF = wd.WOOD_HANDOFF
day_num = wd.day_num
parse = wd.parse
THIS_MON = wd.THIS_MON

EN = {"BEDFRAME": "Bedframe", "SOFA": "Sofa"}

ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
DATE_FILL = PatternFill("solid", fgColor="1F6F54"); DATE_FONT = F(bold=True, color="FFFFFF", size=10)
BF_FILL = PatternFill("solid", fgColor="DDEBF7")
SOFA_FILL = PatternFill("solid", fgColor="FFF2CC")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
LANE_FILL = {"BEDFRAME": BF_FILL, "SOFA": SOFA_FILL}

wb = Workbook()

# ---- Sheet 1: Wood Cut Calendar (interleaved by DATE) ----------------------
ws = wb.active; ws.title = "Wood Cut Calendar"
headers = ["Cut Date", "Lane", "SO ID", "Model (size-agnostic)", "Sizes",
           "Item (wood part)", "Qty", "Sets", "Customer", "Customer DD", "Sew done"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

allu = []
for li, lane in enumerate(LANE_ORDER):
    for u in lane_units[lane]:
        allu.append((u["day"], li, lane, u))
allu.sort(key=lambda t: (t[0], t[1], t[3]["model_key"], t[3]["so"] or ""))

cur_date = None
for sday, li, lane, u in allu:
    if sday != cur_date:
        cur_date = sday
        ws.append([sday.strftime("%Y-%m-%d  %a") + f"   (Day {day_num(sday)})"])
        rr = ws.max_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
        ws.cell(row=rr, column=1).fill = DATE_FILL; ws.cell(row=rr, column=1).font = DATE_FONT
    models_txt = " + ".join(u["models"])
    sizes_txt = " / ".join(u["sizes"])
    sew_txt = u["sew_done"].strftime("%m-%d %a") if u["sew_done"] else "done / no sewing"
    for i, c in enumerate(u["cards"]):
        ws.append(["", EN[lane] if i == 0 else "", u["so"] if i == 0 else "",
                   models_txt if i == 0 else "", sizes_txt if i == 0 else "",
                   c.get("wipLabel") or c.get("model"), max(1, int(c.get("wipQty") or 1)),
                   u["sets"] if i == 0 else "", c.get("custName") if i == 0 else "",
                   c.get("cdd"), sew_txt if i == 0 else ""])
        rr = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=rr, column=col).fill = LANE_FILL[lane]; ws.cell(row=rr, column=col).border = BORDER
            ws.cell(row=rr, column=col).font = F(size=9)
        if parse(c.get("cdd")) and parse(c.get("cdd")) < THIS_MON:
            ws.cell(row=rr, column=10).fill = OVERDUE_FILL
widths = [20, 16, 15, 18, 14, 28, 5, 6, 16, 12, 16]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---- Sheet 2: By Day -------------------------------------------------------
ws2 = wb.create_sheet("By Day")
h2 = ["Date", "Day", "Lane", "SOs that day (model)", "SO count", "Sets", "Cap"]
ws2.append(h2)
perday = defaultdict(lambda: defaultdict(list))
for lane in LANE_ORDER:
    for u in lane_units[lane]:
        perday[lane][u["day"]].append(u)
for d in sorted({d for lane in LANE_ORDER for d in perday[lane]}):
    for lane in LANE_ORDER:
        if d not in perday[lane]: continue
        us = perday[lane][d]
        txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
        ws2.append([d.strftime("%Y-%m-%d %a"), day_num(d), EN[lane], txt,
                    len(us), sum(u["sets"] for u in us), CAP_SETS[lane]])
        rr = ws2.max_row
        for c in range(1, len(h2)+1):
            ws2.cell(row=rr, column=c).fill = LANE_FILL[lane]; ws2.cell(row=rr, column=c).border = BORDER
            ws2.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h2)+1):
    hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 16, 70, 6, 7, 6], 1): ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ---- Sheet 0: Summary & Notes ----------------------------------------------
ws0 = wb.create_sheet("Summary & Notes", 0)
spans = {}
for lane in LANE_ORDER:
    us = lane_units[lane]
    if us:
        spans[lane] = (min(u["day"] for u in us), max(u["day"] for u in us),
                       len(us), sum(u["sets"] for u in us), sum(len(u["cards"]) for u in us))
bydist = {l: sum(1 for u in units if u["cat"] == l) for l in LANE_ORDER}
notes = [
    ("Wood Cutting - Day-by-Day Schedule", True, 14),
    ("Generated 2026-06-01. Schedules ONLY wood-cut cards still to do (WAITING); excludes done, on-hold, cancelled. Read-only, no ERP writes.", False, 10),
    ("", False, 10),
    ("[The chain: each stage scheduled after the previous one finishes]", True, 11),
    ("  Fabric Cut -> (done date) -> Fabric Sew -> (done date) -> Wood Cut. Wood is cut only after that order's SEWING is done.", False, 10),
    ("  Start floor = that SO's sewing-done day 'cannot cut same day', skip two working days, cut on the third.", False, 10),
    ("  e.g. sewing done 6/2 (Tue) -> 6/3, 6/4 idle -> cut on 6/5 (Fri). Orders already sewn / with no sewing can cut from day 1.", False, 10),
    ("", False, 10),
    ("[Scope]", True, 11),
    (f"  Orders still to cut = {len(units)} SOs  (Bedframe {bydist['BEDFRAME']} / Sofa {bydist['SOFA']}).", False, 10),
    ("", False, 10),
    ("[Same SO together + same model, no size split]", True, 11),
    ("  - One SO is always cut on the same day, never split - a sofa's base + arms + cushions (incl. left/right lines) all together.", False, 10),
    ("  - Bedframe 1013-(Q)/(K)/(SS)/(SK) count as one model '1013', all sizes cut together; same sofa width cut together.", False, 10),
    ("", False, 10),
    ("[Capacity: sets/day] (1 set = one SO-line = one finished bed / one sofa)", True, 11),
    (f"  - Bedframe {CAP_SETS['BEDFRAME']} sets/day, Sofa {CAP_SETS['SOFA']} sets/day (matches historical completion).", False, 10),
    ("  - Bedframe and sofa are two separate lines, run at the same time (different people).", False, 10),
    ("", False, 10),
    ("[Sort] Customer DD first (whoever ships first cuts first), then which order's fabric is sewn (the floor).", True, 11),
    ("", False, 10),
    ("[How to read]", True, 11),
    ("  - Wood Cut Calendar: by date, green band = that day. Rightmost 'Sew done' = this order waits until then to be cut.", False, 10),
    ("  - By Day: one row per lane per day - which SOs and how many sets that day.", False, 10),
    ("", False, 10),
    ("[Resulting spans]", True, 11),
]
for lane in LANE_ORDER:
    if lane in spans:
        a, b, nso, nsets, ncards = spans[lane]
        notes.append((f"  {EN[lane]:<10} {a.strftime('%m-%d %a')} -> {b.strftime('%m-%d %a')}"
                      f"   {nso} SOs / {nsets} sets / {ncards} parts", False, 10))
notes += [("", False, 10),
          ("  Note: wood cut follows sewing; if the sewing schedule changes, the wood floor moves and this is re-run. Read-only planning, not written to the ERP.", False, 10)]
for i, (t, b, s) in enumerate(notes, 1):
    ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws0.column_dimensions["A"].width = 120

out = os.path.join(OUT_DIR, "3-WoodCutting.xlsx")
wb.save(out)
print("OK ->", out)
