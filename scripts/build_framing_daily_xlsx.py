# =============================================================================
# FRAMING — day-by-day schedule  (read-only, no ERP writes)
# =============================================================================
# THE CHAIN (Wei Siang 2026-06-01):
#   Fab Cutting -> Fab Sewing -> Wood Cutting -> FRAMING (+ Webbing + HB-foam same day)
#
# RULES (Wei Siang):
#   - FLOOR per SO: framing starts only AFTER that SO's WOOD CUT is done.
#       floor = wood-cut day + FRAME_HANDOFF working days (隔天就做 = +1).
#       SOs with no wood-cut WAITING (wood already done) start from day 1.
#   - CAPACITY by PRODUCTION HOURS, not set count ("framing 跟着 production hour 的 output").
#       One combined daily man-hour budget for the framing crew (BF + sofa share it).
#       Derived from history (good days 5/28 = 26h, 5/29 = 18.6h) -> CAP_HOURS/day.
#   - SAME SO TOGETHER: a whole SO is framed on the same day, never split.
#   - PRIORITY: earliest Customer DD first, then earliest ready (floor).
#   - SAME DAY: each SO's WEBBING + HB-FOAM-BONDING happen the SAME day as its framing
#       (different people / stations, run in parallel; they do NOT eat the framing hours).
#
#   Calendar starts 2026-06-02 (6/1 holiday), skips Sundays + holidays.
#   READ-ONLY. Nothing written to the ERP.
# =============================================================================
import json, os, importlib.util
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# ---- pull the WOOD CUT schedule (its per-SO cut day is our floor) -------------
_spec = importlib.util.spec_from_file_location("wood", os.path.join(HERE, "build_woodcut_daily_xlsx.py"))
wood = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(wood)

# reuse wood's calendar helpers so every dept shares ONE calendar
START, HOLIDAYS, DAY1 = wood.START, wood.HOLIDAYS, wood.DAY1
next_workday, step_workday, add_workdays = wood.next_workday, wood.step_workday, wood.add_workdays
day_num, parse, _pe = wood.day_num, wood.parse, wood._pe
THIS_MON = START - timedelta(days=START.weekday())

wood_day = {}                                  # (coSOId, cat) -> wood-cut day
for _lane, _us in wood.lane_units.items():
    for _u in _us:
        wood_day[(_u["so"], _u["cat"])] = _u["day"]

# ---- parameters --------------------------------------------------------------
# Separate lanes: bedframe & sofa framed by DIFFERENT teams, own hour budget each
# (Wei Siang: framing 要分开 bedframe 和 sofa). History good days:
#   bedframe 5/28=22.5h 5/29=15.6h ; sofa 5/21=7.3h 5/13=6.7h.
CAP_HOURS = {"BEDFRAME": 22, "SOFA": 8}        # framing man-hours / day per lane (Wei Siang 2026-06-02; BF real avg 20.8h, max 32.3h)
CAP_MIN = {k: v * 60 for k, v in CAP_HOURS.items()}
LANE_ORDER = ["BEDFRAME", "SOFA"]
FRAME_HANDOFF = 1                              # working days after wood cut (隔天就做)
CN = {"BEDFRAME": "床架 Bedframe", "SOFA": "沙发 Sofa"}

def mins_of(c): return (c.get("prodTime") or 0) * max(1, int(c.get("wipQty") or 1))

# ---- load WAITING cards for FRAMING + its same-day riders ---------------------
full = json.load(open(os.path.join(HERE, "sched_dept_full.json"), encoding="utf-8"))
def waiting(dept, cats=("BEDFRAME", "SOFA")):
    return [r for r in full.get(dept, [])
            if r.get("cardStatus") == "WAITING"
            and r.get("orderStatus") not in ("ON_HOLD", "CANCELLED")
            and r.get("cat") in cats]

frame_cards = waiting("FRAMING")
web_by_so = defaultdict(list)
for c in waiting("WEBBING"):
    web_by_so[(c.get("coSOId"), c.get("cat"))].append(c)
hbfoam_by_so = defaultdict(list)               # headboard foam bonding (bedframe HB only)
for c in waiting("FOAM"):
    if c.get("wipType") == "HEADBOARD":
        hbfoam_by_so[(c.get("coSOId"), c.get("cat"))].append(c)

# ---- group framing into SO-units (per coSOId + cat) --------------------------
by_unit = defaultdict(list)
for c in frame_cards:
    by_unit[(c.get("coSOId"), c.get("cat"))].append(c)

def base_set_count(cards, cat):
    wt = "DIVAN" if cat == "BEDFRAME" else "SOFA_BASE"
    base = [c for c in cards if c.get("wipType") == wt]
    if base: return sum(max(1, int(c.get("wipQty") or 1)) for c in base)
    return len({c.get("po") for c in cards})

units = []
for (so, cat), cards in by_unit.items():
    wd = wood_day.get((so, cat))
    floor = add_workdays(wd, FRAME_HANDOFF) if wd else DAY1
    cdds = sorted(c["cdd"] for c in cards if c.get("cdd"))
    models = sorted({(c.get("model") or "").strip() for c in cards if c.get("model")})
    fmin = sum(mins_of(c) for c in cards)
    web = web_by_so.get((so, cat), [])
    hbf = hbfoam_by_so.get((so, cat), [])
    units.append({"so": so, "cat": cat, "cards": cards, "fmin": fmin,
                  "sets": base_set_count(cards, cat),
                  "cdd": cdds[0] if cdds else "", "models": models,
                  "floor": floor, "wood_done": wd,
                  "web": web, "hbf": hbf,
                  "web_min": sum(mins_of(c) for c in web),
                  "hbf_min": sum(mins_of(c) for c in hbf),
                  "model_key": models[0] if models else ""})

# ---- schedule: per-lane hour budget; DD first; same SO whole -----------------
load = {lane: defaultdict(int) for lane in LANE_ORDER}   # minutes used per day, per lane
for lane in LANE_ORDER:
    us = [u for u in units if u["cat"] == lane]
    us.sort(key=lambda u: (_pe(u["cdd"]), u["floor"], u["model_key"], u["so"] or ""))
    cap = CAP_MIN[lane]
    for u in us:
        d = max(u["floor"], DAY1)
        while load[lane][d] + u["fmin"] > cap and load[lane][d] != 0:
            d = step_workday(d)
        u["day"] = d
        load[lane][d] += u["fmin"]

# ---- Sofa foam bonding sub-stage (Wei Siang 2026-06-01) ----------------------
# After sofa framing+webbing (same day), foam bonding runs the NEXT working day.
#   · base + armrest -> capacity by production hour: 8h/day budget, overflow pushes
#     to next working day, DD-first  -> sets foam_day  (Wei Siang: 8h 只算 base+armrest)
#   · back cushion   -> bonded the SAME day as base (rides foam_day, follows base even
#     when base pushed), but its hours do NOT count toward the 5h cap
FOAM_CAP_MIN = 8 * 60                           # sofa base+armrest foam bonding, 8h/day (Wei Siang 2026-06-02, real avg 6.3h)
FOAM_HANDOFF = 1                                # foam bonding = sofa framing day + 1 working day (隔天)
sofa_foam = defaultdict(lambda: {"base": [], "arm": [], "cush": []})  # coSOId -> parts
for c in waiting("FOAM", cats=("SOFA",)):
    wt = c.get("wipType")
    if wt == "SOFA_BASE":      sofa_foam[c.get("coSOId")]["base"].append(c)
    elif wt == "SOFA_ARMREST": sofa_foam[c.get("coSOId")]["arm"].append(c)
    elif wt == "SOFA_CUSHION": sofa_foam[c.get("coSOId")]["cush"].append(c)

foam_units = []
for u in units:
    if u["cat"] != "SOFA":
        continue
    p = sofa_foam.get(u["so"])
    if not p or not (p["base"] + p["arm"] + p["cush"]):
        continue
    foam_units.append({"so": u["so"], "frame_day": u["day"], "cdd": u["cdd"],
                       "models": u["models"], "model_key": u["model_key"],
                       "base": p["base"], "arm": p["arm"], "cush": p["cush"],
                       "ba_min": sum(mins_of(c) for c in p["base"] + p["arm"]),
                       "bc_min": sum(mins_of(c) for c in p["cush"]),
                       "floor": add_workdays(u["day"], FOAM_HANDOFF)})

foam_load = defaultdict(int)                    # date -> base+armrest minutes (the capped pool)
foam_units.sort(key=lambda u: (_pe(u["cdd"]), u["floor"], u["model_key"], u["so"] or ""))
for u in foam_units:
    d = max(u["floor"], DAY1)
    while foam_load[d] + u["ba_min"] > FOAM_CAP_MIN and foam_load[d] != 0:
        d = step_workday(d)
    u["foam_day"] = d                          # base+armrest day (capacity); back cushion rides same day
    foam_load[d] += u["ba_min"]                # only base+armrest consumes the 5h budget

# ---- Upholstery sub-stage (包工) (Wei Siang 2026-06-02) -----------------------
# Floors (隔天 = upstream day + 1 working day):
#   · Bedframe upholstery -> webbing's next day = framing day + 1 working day
#       (bedframe webbing rides framing the same day, so upholstery = framing + 1).
#   · Sofa upholstery     -> foam bonding (base) done + 1 working day.
#   SOs whose upstream is already done (no WAITING framing/foam) start from day 1.
# Capacity by PRODUCTION HOURS, two separate lanes (Wei Siang, from real ERP history
#   5/18-5/29: bedframe ~17h/day avg, sofa ~9h/day avg):
#   · Bedframe 24 h/day, Sofa 12 h/day. Same SO whole, DD-first.
UPH_CAP_HOURS = {"BEDFRAME": 24, "SOFA": 12}
UPH_CAP_MIN = {k: v * 60 for k, v in UPH_CAP_HOURS.items()}
UPH_HANDOFF = 1                                # working days after the upstream floor (隔天)

frame_day_map = {(u["so"], u["cat"]): u["day"] for u in units}
foam_day_map = {u["so"]: u["foam_day"] for u in foam_units}

uph_by_unit = defaultdict(list)
for c in waiting("UPHOLSTERY"):
    uph_by_unit[(c.get("coSOId"), c.get("cat"))].append(c)

uph_units = []
for (so, cat), cards in uph_by_unit.items():
    if cat == "BEDFRAME":
        up = frame_day_map.get((so, cat))          # framing day; webbing rides same day
    else:
        up = foam_day_map.get(so)                  # sofa base foam-bonding day
    floor = add_workdays(up, UPH_HANDOFF) if up else DAY1
    cdds = sorted(c["cdd"] for c in cards if c.get("cdd"))
    models = sorted({(c.get("model") or "").strip() for c in cards if c.get("model")})
    umin = sum(mins_of(c) for c in cards)
    uph_units.append({"so": so, "cat": cat, "cards": cards, "umin": umin,
                      "cdd": cdds[0] if cdds else "", "models": models,
                      "floor": floor, "upstream": up,
                      "model_key": models[0] if models else ""})

uph_load = {lane: defaultdict(int) for lane in LANE_ORDER}   # minutes used per day, per lane
for lane in LANE_ORDER:
    us = [u for u in uph_units if u["cat"] == lane]
    us.sort(key=lambda u: (_pe(u["cdd"]), u["floor"], u["model_key"], u["so"] or ""))
    cap = UPH_CAP_MIN[lane]
    for u in us:
        d = max(u["floor"], DAY1)
        while uph_load[lane][d] + u["umin"] > cap and uph_load[lane][d] != 0:
            d = step_workday(d)
        u["uph_day"] = d
        uph_load[lane][d] += u["umin"]

# ---- Packing sub-stage (Wei Siang 2026-06-02) --------------------------------
# NOT capacity-limited. Packing date = its SO's upholstery date (packing 的 due date 跟
# upholstery 一起). SOs whose upholstery is already done start from day 1.
uph_day_map = {(u["so"], u["cat"]): u["uph_day"] for u in uph_units}
pack_by_unit = defaultdict(list)
for c in waiting("PACKING"):
    pack_by_unit[(c.get("coSOId"), c.get("cat"))].append(c)

pack_units = []
for (so, cat), cards in pack_by_unit.items():
    ud = uph_day_map.get((so, cat))                # same day as upholstery
    pday = ud if ud else DAY1
    cdds = sorted(c["cdd"] for c in cards if c.get("cdd"))
    models = sorted({(c.get("model") or "").strip() for c in cards if c.get("model")})
    pmin = sum(mins_of(c) for c in cards)
    pack_units.append({"so": so, "cat": cat, "cards": cards, "pmin": pmin,
                       "cdd": cdds[0] if cdds else "", "models": models,
                       "pack_day": pday, "uph_day": ud,
                       "model_key": models[0] if models else ""})

# ===================== styling ===============================================
ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
DATE_FILL = PatternFill("solid", fgColor="7030A0"); DATE_FONT = F(bold=True, color="FFFFFF", size=10)
BF_FILL = PatternFill("solid", fgColor="DDEBF7"); SOFA_FILL = PatternFill("solid", fgColor="FFF2CC")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
LANE_FILL = {"BEDFRAME": BF_FILL, "SOFA": SOFA_FILL}

wb = Workbook()

# ---- Sheet 1: Framing Calendar (by DATE) ------------------------------------
ws = wb.active; ws.title = "Framing Calendar"
headers = ["Frame Date 做框日", "Lane 队", "SO ID", "Model 款", "Item 框件",
           "件", "工时min", "Sets 套", "Webbing 同日", "HB Foam 同日",
           "Customer", "Customer DD", "木工完成 Wood done"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

units_sorted = sorted(units, key=lambda u: (u["day"], 0 if u["cat"] == "BEDFRAME" else 1,
                                             _pe(u["cdd"]), u["so"] or ""))
cur_date = None
for u in units_sorted:
    if u["day"] != cur_date:
        cur_date = u["day"]
        bfu = load["BEDFRAME"][cur_date]; sfu = load["SOFA"][cur_date]
        ws.append([cur_date.strftime("%Y-%m-%d  %a") + f"   (Day {day_num(cur_date)})"
                   + f"   —  床架 {bfu/60:.1f}h/{CAP_HOURS['BEDFRAME']}h  ·  沙发 {sfu/60:.1f}h/{CAP_HOURS['SOFA']}h"])
        rr = ws.max_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
        ws.cell(row=rr, column=1).fill = DATE_FILL; ws.cell(row=rr, column=1).font = DATE_FONT
    lane = u["cat"]
    web_txt = f"{len(u['web'])}件 / {u['web_min']}min" if u["web"] else "—"
    hbf_txt = f"{len(u['hbf'])}件 / {u['hbf_min']}min" if u["hbf"] else "—"
    wd_txt = u["wood_done"].strftime("%m-%d %a") if u["wood_done"] else "已完成/无木工"
    for i, c in enumerate(u["cards"]):
        ws.append(["", CN[lane] if i == 0 else "", u["so"] if i == 0 else "",
                   " + ".join(u["models"]) if i == 0 else "",
                   c.get("wipLabel") or c.get("model"), max(1, int(c.get("wipQty") or 1)),
                   mins_of(c),
                   u["sets"] if i == 0 else "",
                   web_txt if i == 0 else "", hbf_txt if i == 0 else "",
                   c.get("custName") if i == 0 else "", c.get("cdd"),
                   wd_txt if i == 0 else ""])
        rr = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=rr, column=col).fill = LANE_FILL[lane]
            ws.cell(row=rr, column=col).border = BORDER
            ws.cell(row=rr, column=col).font = F(size=9)
        if parse(c.get("cdd")) and parse(c.get("cdd")) < THIS_MON:
            ws.cell(row=rr, column=12).fill = OVERDUE_FILL
widths = [22, 16, 14, 14, 28, 5, 7, 6, 14, 14, 16, 12, 16]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---- Sheet 2: By Day --------------------------------------------------------
ws2 = wb.create_sheet("By Day")
h2 = ["Date 日期", "Day", "床架h", "床架cap", "沙发h", "沙发cap",
      "床架SO", "沙发SO", "Webbing件(同日)", "HB Foam件(同日)", "SOs (model)"]
ws2.append(h2)
perday = defaultdict(list)
for u in units: perday[u["day"]].append(u)
for d in sorted(perday):
    us = perday[d]
    bf = [u for u in us if u["cat"] == "BEDFRAME"]; sf = [u for u in us if u["cat"] == "SOFA"]
    txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
    ws2.append([d.strftime("%Y-%m-%d %a"), day_num(d),
                round(load["BEDFRAME"][d] / 60, 1), CAP_HOURS["BEDFRAME"],
                round(load["SOFA"][d] / 60, 1), CAP_HOURS["SOFA"],
                len(bf), len(sf),
                sum(len(u["web"]) for u in us), sum(len(u["hbf"]) for u in us), txt])
    rr = ws2.max_row
    for c in range(1, len(h2) + 1):
        ws2.cell(row=rr, column=c).border = BORDER; ws2.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h2) + 1):
    hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 7, 7, 7, 7, 7, 7, 14, 14, 64], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ---- Sheet 3: Sofa Foam Bonding (base+armrest scheduled, back cushion rides) -
ws3 = wb.create_sheet("Sofa Foam Bonding")
h3 = ["Date 日期", "Day", "Foam 工时h", "cap", "件 Base/Arm/Cush", "SOs (model)"]
ws3.append(h3)
fb_by_day = defaultdict(list)
for u in foam_units:
    fb_by_day[u["foam_day"]].append(u)
for d in sorted(fb_by_day):
    us = fb_by_day[d]
    nb = sum(len(u["base"]) for u in us); na = sum(len(u["arm"]) for u in us)
    nc = sum(len(u["cush"]) for u in us)
    so_txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
    ws3.append([d.strftime("%Y-%m-%d %a"), day_num(d),
                round(foam_load[d] / 60, 1), FOAM_CAP_MIN // 60, f"{nb}/{na}/{nc}", so_txt])
    rr = ws3.max_row
    for c in range(1, len(h3) + 1):
        ws3.cell(row=rr, column=c).border = BORDER; ws3.cell(row=rr, column=c).font = F(size=9)
        ws3.cell(row=rr, column=c).fill = SOFA_FILL
for c in range(1, len(h3) + 1):
    hc = ws3.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 12, 6, 16, 64], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ---- Sheet 4: Upholstery (包工) by day --------------------------------------
ws4 = wb.create_sheet("Upholstery")
h4 = ["Date 日期", "Day", "床架h", "床架cap", "沙发h", "沙发cap",
      "床架SO", "沙发SO", "SOs (model)"]
ws4.append(h4)
uph_perday = defaultdict(list)
for u in uph_units: uph_perday[u["uph_day"]].append(u)
for d in sorted(uph_perday):
    us = uph_perday[d]
    bf = [u for u in us if u["cat"] == "BEDFRAME"]; sf = [u for u in us if u["cat"] == "SOFA"]
    txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
    ws4.append([d.strftime("%Y-%m-%d %a"), day_num(d),
                round(uph_load["BEDFRAME"][d] / 60, 1), UPH_CAP_HOURS["BEDFRAME"],
                round(uph_load["SOFA"][d] / 60, 1), UPH_CAP_HOURS["SOFA"],
                len(bf), len(sf), txt])
    rr = ws4.max_row
    for c in range(1, len(h4) + 1):
        ws4.cell(row=rr, column=c).border = BORDER; ws4.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h4) + 1):
    hc = ws4.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 7, 7, 7, 7, 7, 7, 64], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A2"

# ---- Sheet 5: Packing by day (follows upholstery, no capacity cap) ----------
ws5 = wb.create_sheet("Packing")
h5 = ["Date 日期", "Day", "件总", "床架SO", "沙发SO", "SOs (model)"]
ws5.append(h5)
pack_perday = defaultdict(list)
for u in pack_units: pack_perday[u["pack_day"]].append(u)
for d in sorted(pack_perday):
    us = pack_perday[d]
    bf = [u for u in us if u["cat"] == "BEDFRAME"]; sf = [u for u in us if u["cat"] == "SOFA"]
    ncards = sum(len(u["cards"]) for u in us)
    txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
    ws5.append([d.strftime("%Y-%m-%d %a"), day_num(d), ncards, len(bf), len(sf), txt])
    rr = ws5.max_row
    for c in range(1, len(h5) + 1):
        ws5.cell(row=rr, column=c).border = BORDER; ws5.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h5) + 1):
    hc = ws5.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 7, 7, 7, 64], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = "A2"

# ---- Sheet 0: Summary & Notes -----------------------------------------------
ws0 = wb.create_sheet("Summary & Notes", 0)
tot_fmin = sum(u["fmin"] for u in units)
tot_web = sum(len(u["web"]) for u in units); tot_hbf = sum(len(u["hbf"]) for u in units)
nbf = sum(1 for u in units if u["cat"] == "BEDFRAME"); nsf = len(units) - nbf
lane_span = {}
for lane in LANE_ORDER:
    ds = [u["day"] for u in units if u["cat"] == lane]
    if ds: lane_span[lane] = (min(ds), max(ds))
notes = [
    ("Framing 逐日排程  Framing — Day-by-Day Schedule", True, 14),
    ("生成 2026-06-01。只排『还没做(WAITING)』的 framing 卡,不含已完成、on-hold、已取消。只读,没改系统。", False, 10),
    ("", False, 10),
    ("【整条链】裁布 → 车缝 → 木工 → Framing。Framing 排在它那张单『木工切好』之后才做。", True, 11),
    (f"  起排底线 = 该 SO 木工完成日 + {FRAME_HANDOFF} 个工作日(隔天就做)。木工已完成/没木工的,第一天就能做。", False, 10),
    ("", False, 10),
    ("【产能:按工时算,不是套数。床架、沙发分开两队人、各自工时池】", True, 11),
    (f"  · 床架 {CAP_HOURS['BEDFRAME']} 工时/天(历史好日子 5/28=22.5h、5/29=15.6h)。", False, 10),
    (f"  · 沙发 {CAP_HOURS['SOFA']} 工时/天(老板定;历史沙发量少做不到这么高,这是按实际产能定的)。", False, 10),
    (f"  · 待做总量 = {tot_fmin/60:.0f} 工时(床架 {sum(u['fmin'] for u in units if u['cat']=='BEDFRAME')/60:.0f}h / 沙发 {sum(u['fmin'] for u in units if u['cat']=='SOFA')/60:.0f}h)。", False, 10),
    ("", False, 10),
    ("【同 SO 一起】一个 SO 一定同一天做完,不拆。床架=divan+床头;沙发=底座+扶手+靠垫一起。", True, 11),
    ("", False, 10),
    ("【床架 Webbing + HB Foam Bonding 跟 framing 同一天】", True, 11),
    ("  · 哪天做这张单的 framing,它的 webbing 和 床头(HB)foam bonding 就同一天一起做。", False, 10),
    (f"  · 这两样另一批人/工位并行,不占 framing 那 {CAP_HOURS['BEDFRAME']} 小时;表里『同日』两栏标了件数+工时。", False, 10),
    (f"  · 同日带做总量:Webbing {tot_web} 件,HB Foam {tot_hbf} 件。", False, 10),
    ("", False, 10),
    ("【沙发后段:webbing 同日,foam bonding 隔天】(见『Sofa Foam Bonding』表)", True, 11),
    ("  · Webbing(底座+靠垫+扶手 三样)跟 sofa framing 同一天做,不看产能。", False, 10),
    ("  · Foam bonding 在 framing 的隔天(下一个工作日):", False, 10),
    (f"      - 底座+扶手:按产能排,每天上限 {FOAM_CAP_MIN/60:.0f} 小时,做不完往后推。", False, 10),
    (f"      - 后靠垫(back cushion):跟着底座同一天做完(底座被推后它也跟着),但不占那 {FOAM_CAP_MIN//60} 小时。", False, 10),
    ("", False, 10),
    ("【包工 Upholstery:隔天 + 按产能,床架沙发分开】(见『Upholstery』表)", True, 11),
    (f"  · 床架包工 = framing 的隔天(webbing 跟 framing 同日,所以包工 = framing+1)。每天上限 {UPH_CAP_HOURS['BEDFRAME']} 小时。", False, 10),
    (f"  · 沙发包工 = 底座 foam bonding 做好的隔天。每天上限 {UPH_CAP_HOURS['SOFA']} 小时。", False, 10),
    (f"  · 产能按近两周实际定(床架真实约 17h/天、沙发约 9h/天);做不完往后推,同 SO 一起、客期先。", False, 10),
    (f"  · 待做总量:床架 {sum(u['umin'] for u in uph_units if u['cat']=='BEDFRAME')/60:.0f}h、沙发 {sum(u['umin'] for u in uph_units if u['cat']=='SOFA')/60:.0f}h。", False, 10),
    ("", False, 10),
    ("【包装 Packing:跟着包工的日期,不看产能】(见『Packing』表)", True, 11),
    ("  · Packing 的 due date 跟 upholstery 同一天;不卡产能,包工哪天好就哪天包。", False, 10),
    ("", False, 10),
    ("【排序】先看客期(Customer DD,谁先送货谁先做),再看哪张单的木工先切好。", True, 11),
    ("", False, 10),
    ("【排出来的跨度】", True, 11),
    (f"  床架 Bedframe  {lane_span['BEDFRAME'][0].strftime('%m-%d %a')} → {lane_span['BEDFRAME'][1].strftime('%m-%d %a')}"
     f"   {nbf} 单 / {sum(u['fmin'] for u in units if u['cat']=='BEDFRAME')/60:.0f} 工时", False, 10),
    (f"  沙发 Sofa      {lane_span['SOFA'][0].strftime('%m-%d %a')} → {lane_span['SOFA'][1].strftime('%m-%d %a')}"
     f"   {nsf} 单 / {sum(u['fmin'] for u in units if u['cat']=='SOFA')/60:.0f} 工时", False, 10),
    ("", False, 10),
    ("  注:framing 跟着木工走;木工表一改,framing 底线跟着动,要重跑。只读 planning,还没写进 ERP。", False, 10),
]
for i, (t, b, s) in enumerate(notes, 1):
    ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws0.column_dimensions["A"].width = 104

out = r"C:\Users\User\hookka-erp-testing\Framing-Daily-Schedule-2026-06-02.xlsx"
try: wb.save(out)
except PermissionError:
    out = r"C:\Users\User\hookka-erp-testing\Framing-Daily-Schedule-2026-06-02-v2.xlsx"
    wb.save(out)
print("OK ->", out)
print(f"framing SO-units: {len(units)}  (BF {nbf} / SOFA {nsf})  total {tot_fmin/60:.0f}h")
for lane in LANE_ORDER:
    a, b = lane_span[lane]
    print(f"  {lane:<9} {a} -> {b}   cap {CAP_HOURS[lane]}h/day")
print(f"  same-day riders: webbing {tot_web} cards, HB-foam {tot_hbf} cards")
for lane in LANE_ORDER:
    busy = sorted(load[lane].items())
    print(f"  {lane} per-day h:", [f"{d.strftime('%m-%d')}={m/60:.1f}" for d, m in busy])

print(f"upholstery SO-units: {len(uph_units)}  (cap BF {UPH_CAP_HOURS['BEDFRAME']}h / SOFA {UPH_CAP_HOURS['SOFA']}h)")
for lane in LANE_ORDER:
    ds = [u["uph_day"] for u in uph_units if u["cat"] == lane]
    if ds:
        print(f"  {lane:<9} {min(ds)} -> {max(ds)}   {sum(u['umin'] for u in uph_units if u['cat']==lane)/60:.0f}h")
    busy = sorted(uph_load[lane].items())
    print(f"  {lane} uph per-day h:", [f"{d.strftime('%m-%d')}={m/60:.1f}" for d, m in busy])
pds = [u["pack_day"] for u in pack_units]
if pds:
    print(f"packing SO-units: {len(pack_units)}  (no cap)  {min(pds)} -> {max(pds)}")
