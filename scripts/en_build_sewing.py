# =============================================================================
# FABRIC SEWING — English report (presentation-only translation of
# build_sewing_daily_xlsx.py). Imports the ORIGINAL module (its workbook code is
# guarded by __main__, so import yields only the computed structures).
# =============================================================================
import os, importlib.util
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = r"C:\Users\User\Desktop\Production-Schedule-2026-06-02-EN"


def _load(name, fname):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, fname))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


sw = _load("sew_src", "build_sewing_daily_xlsx.py")

LANE_ORDER = sw.LANE_ORDER
results = sw.results
loads = sw.loads
groups_by_lane = sw.groups_by_lane
cards = sw.cards
cap_on = sw.cap_on
starvation = sw.starvation
workday_index = sw.workday_index
parse = sw.parse
next_workday = sw.next_workday
START = sw.START
CAP_HOURS = sw.CAP_HOURS
CAP_MIN = sw.CAP_MIN
PILLOW_MIN_EACH = sw.PILLOW_MIN_EACH
HANDOFF_WORKDAYS = sw.HANDOFF_WORKDAYS
LANE_CONFIRMED = sw.LANE_CONFIRMED
RESERVE_TIERS = sw.RESERVE_TIERS

EN = {"BEDFRAME": "Bedframe", "SOFA": "Sofa", "ACCESSORY": "Pillow"}
THIS_MON = next_workday(START)


def F(**k): return Font(name="Calibri", **k)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
DATE_FILL = PatternFill("solid", fgColor="1F6F54"); DATE_FONT = F(bold=True, color="FFFFFF", size=10)
BF_FILL = PatternFill("solid", fgColor="DDEBF7")
SOFA_FILL = PatternFill("solid", fgColor="FFF2CC")
ACC_FILL = PatternFill("solid", fgColor="FCE4D6")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
BLOCK_FILL = PatternFill("solid", fgColor="E2EFDA")
LANE_FILL = {"BEDFRAME": BF_FILL, "SOFA": SOFA_FILL, "ACCESSORY": ACC_FILL}


def tier_of(d):
    i = workday_index(d)
    for n, (upto, f) in enumerate(RESERVE_TIERS):
        if i <= upto:
            return ["packed", "medium", "loose"][n]
    return "loose"


wb = Workbook()

# ---- Sheet 1: Sew Calendar (interleaved by DATE) ----------------------
ws = wb.active; ws.title = "Sew Calendar"
headers = ["Lane", "SO ID", "Customer", "Model",
           "Size", "PO / Line", "Item", "Qty", "Mins",
           "Customer DD", "Fabric Status"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

REASON_EN = {"cut-done": "cut done", "no-cut-step": "no cut needed (ready fabric)",
             "cut-pending-unplanned": "waiting to cut (soon)", "after-cut": "waiting for fabric cut",
             "cut-pending-planned": "waiting for fabric cut"}
allc = []
for li, lane in enumerate(LANE_ORDER):
    for c in results[lane]:
        allc.append((c["day"], li, lane, c))
allc.sort(key=lambda t: (t[0], t[1], t[3]["soNo"] or "", -t[3]["mins"]))

cur_date = None; cur_so = None
for sday, li, lane, c in allc:
    if sday != cur_date:
        cur_date = sday; cur_so = None
        placed = {l: loads[l].get(sday, 0) for l in LANE_ORDER}
        tag = "  ".join(f"{EN[l]} {placed[l]/60:.1f}/{cap_on(l, sday)/60:.0f}h"
                        for l in LANE_ORDER if placed[l])
        ws.append([sday.strftime("%Y-%m-%d  %a")
                   + f"   (Day {workday_index(sday)})   {tag}"])
        rr = ws.max_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
        ws.cell(row=rr, column=1).fill = DATE_FILL; ws.cell(row=rr, column=1).font = DATE_FONT
    first = c["soNo"] != cur_so
    cur_so = c["soNo"]
    ws.append([EN[lane] if first else "", c["soNo"] if first else "",
               c["cust"] if first else "", c["model"], c["size"], c["po"],
               c["label"], c["qty"], c["mins"], c["cdd"],
               REASON_EN.get(c["reason"], c["reason"])])
    rr = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=rr, column=col).fill = LANE_FILL[lane]
        ws.cell(row=rr, column=col).border = BORDER
        ws.cell(row=rr, column=col).font = F(size=9)
    if parse(c["cdd"]) and c["day"] > parse(c["cdd"]):
        ws.cell(row=rr, column=10).fill = OVERDUE_FILL
    if c["reason"] in ("after-cut",):
        ws.cell(row=rr, column=11).fill = BLOCK_FILL
widths = [14, 14, 18, 14, 8, 14, 30, 5, 7, 13, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---- Sheet 2: By Day (per lane-day load vs cap) -----------------------
ws2 = wb.create_sheet("By Day")
h2 = ["Date", "Day", "Lane", "SOs that day",
      "Cards", "Load h", "Cap h", "Use%", "Tier",
      "Cut-blocked h"]
ws2.append(h2)
perday = defaultdict(lambda: defaultdict(list))
for lane in LANE_ORDER:
    for c in results[lane]:
        perday[lane][c["day"]].append(c)
alldays = sorted({d for lane in LANE_ORDER for d in perday[lane]})
for d in alldays:
    for lane in LANE_ORDER:
        if d not in perday[lane]:
            continue
        cs = perday[lane][d]
        load = sum(c["mins"] for c in cs)
        cap = cap_on(lane, d)
        sos = sorted({c["soNo"] for c in cs})
        blocked = sum(c["mins"] for c in results[lane]
                      if c["floor"] > d and parse(c["cdd"]) and parse(c["cdd"]) <= d)
        ws2.append([d.strftime("%Y-%m-%d %a"), workday_index(d), EN[lane],
                    ", ".join(str(s) for s in sos), len(cs),
                    round(load / 60, 1), round(cap / 60, 1),
                    f"{load/cap*100:.0f}%" if cap else "",
                    tier_of(d), round(blocked / 60, 1) if blocked else ""])
        rr = ws2.max_row
        for col in range(1, len(h2) + 1):
            ws2.cell(row=rr, column=col).fill = LANE_FILL[lane]
            ws2.cell(row=rr, column=col).border = BORDER
            ws2.cell(row=rr, column=col).font = F(size=9)
        if blocked:
            ws2.cell(row=rr, column=10).fill = BLOCK_FILL
for c in range(1, len(h2) + 1):
    hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 16, 40, 6, 8, 8, 7, 14, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ---- Sheet 0: Summary & Notes -----------------------------------------
ws0 = wb.create_sheet("Summary & Notes", 0)
spans = {}
for lane in LANE_ORDER:
    gs = groups_by_lane[lane]
    if gs:
        spans[lane] = (min(g["start"] for g in gs), max(g["end"] for g in gs),
                       len(gs), len(results[lane]))
overdue = {l: sum(1 for c in results[l] if parse(c["cdd"]) and parse(c["cdd"]) < THIS_MON)
           for l in LANE_ORDER}
causedlate = {l: sum(1 for c in results[l] if parse(c["cdd"])
                     and parse(c["cdd"]) >= THIS_MON and c["day"] > parse(c["cdd"]))
              for l in LANE_ORDER}
notes = [
    ("Fabric Sewing - Day-by-Day Schedule", True, 14),
    ("Generated 2026-06-01, built on top of the cutting schedule - fabric not yet cut can't be sewn that day. Read-only, no ERP writes.", False, 10),
    ("", False, 10),
    ("[How capacity is set] from past sewing cards with a completion date, split by lane", True, 11),
    (f"  - Sofa SOFA     {CAP_HOURS['SOFA']:.0f} h/day ({CAP_MIN['SOFA']} min)", False, 10),
    (f"  - Bedframe BEDFRAME {CAP_HOURS['BEDFRAME']:.0f} h/day ({CAP_MIN['BEDFRAME']} min)", False, 10),
    (f"  - Pillow {PILLOW_MIN_EACH} min/each (your figure; the ERP's 300 min/each is wrong, overridden).", False, 10),
    ("  Measured over the 7 working days 05-22 to 05-28 (the 29th excluded). Sofa and bedframe are two crews, separate buckets, no cross-fill.", False, 10),
    ("", False, 10),
    ("[How it's planned]", True, 11),
    ("  - By Customer DD first, earliest due sewn first; same date = shortest job first to pack the day tight.", False, 10),
    ("  - Same SO ID always sewn together on consecutive working days, never split across gaps (so the order ships complete).", False, 10),
    ("  - Each day filled to capacity, spills to the next day; a group bigger than one day runs across consecutive days.", False, 10),
    ("  - Every day packed full, no buffer (your directive). Walk-in inserts go to OT; the whole plan is re-cut about every 3 days.", False, 10),
    ("", False, 10),
    ("[Cut -> Sew link] this is the key point", True, 11),
    ("  - Fabric already cut, or no cut needed (ready fabric) -> can sew from day 1.", False, 10),
    (f"  - Fabric not yet cut -> earliest sew = that order's LAST cut in the cutting schedule + {HANDOFF_WORKDAYS} working day handoff.", False, 10),
    ("    (The green column = this card is still waiting for its fabric to be cut.)", False, 10),
    ("", False, 10),
    ("[Cut -> Sew feed (coupling)]", True, 11),
]
for lane in LANE_ORDER:
    st = starvation(lane)
    if st:
        tot = sum(b for _, b in st.values()) / 60
        ds = sorted(st)
        notes.append((f"  - {EN[lane]}: {len(st)} days with crew free but no fabric to sew, "
                      f"{tot:.1f} h of due work stuck waiting on cutting "
                      f"({ds[0].strftime('%m-%d')}->{ds[-1].strftime('%m-%d')}).", False, 10))
notes += [("    -> On those days squeeze more cutting into this lane so the sewing crew isn't idle. Want me to re-tune cutting? Just say so.", False, 10),
          ("", False, 10),
          ("[Resulting spans / late orders]", True, 11)]
for lane in LANE_ORDER:
    if lane in spans:
        a, b, ng, ni = spans[lane]
        tag = "" if LANE_CONFIRMED[lane] else "  (capacity TBC)"
        notes.append((f"  {EN[lane]:<10} {a.strftime('%m-%d %a')} -> {b.strftime('%m-%d %a')}"
                      f"   {ng} SOs / {ni} cards{tag}", False, 10))
        notes.append((f"      Already past customer date {overdue[lane]} cards (fabric/capacity limited, sew ASAP);"
                      f"  schedule-caused late {causedlate[lane]} cards.", False, 10))
notes += [("", False, 10),
          ("[How to read]", True, 11),
          ("  - Sew Calendar: by date, green band = that day; see which SOs sofa / bedframe each sew per day.", False, 10),
          ("  - By Day: one row per lane per day - hours scheduled vs capacity, % used, and how much work is stuck waiting on cutting.", False, 10),
          ("", False, 10),
          ("  Note: planning from read-only live data, not yet written to the ERP. Confirm the plan and I'll write it in.", False, 10)]
for i, (t, b, s) in enumerate(notes, 1):
    ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws0.column_dimensions["A"].width = 120

out = os.path.join(OUT_DIR, "2-FabricSewing.xlsx")
wb.save(out)
print("OK ->", out)
