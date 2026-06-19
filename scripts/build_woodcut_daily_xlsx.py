# =============================================================================
# WOOD CUTTING — day-by-day cut calendar  (read-only, no ERP writes)
# =============================================================================
# THE CHAIN (Wei Siang 2026-06-01):
#     Fabric Cutting --(done date)--> Fabric Sewing --(done date)--> Wood Cutting
#   Each stage is planned AFTER the previous stage finishes for that order.
#   So an order's wood cut can only start AFTER its fabric SEWING is done.
#   Wood cutting lags sewing by many days, so there's plenty of freedom.
#
# RULES:
#   · FLOOR per order: wood cut starts only AFTER that SO's fabric sewing is
#     complete (sew done day + WOOD_HANDOFF working days). Orders whose sewing is
#     already done / has no sewing start from day 1.
#   · SAME SO TOGETHER: a whole SO is cut on the same day — never split. A sofa SO
#     carries base + arms + back-cushion (and sectional left/right lines); ALL of
#     it is cut together. (Wei Siang: "同一个 SO 不可以分开，尤其 sofa".)
#   · SAME MODEL, NO SIZE SPLIT: all sizes of a model are cut together; same models
#     sit adjacent where the sew-done floor allows.
#   · CAPACITY in SETS/day (1 set = 1 SO-line = one finished bed / one sofa):
#     BEDFRAME 20 sets/day, SOFA 10 sets/day (Wei Siang, matches历史 completion).
#   · PRIORITY: earliest Customer DD first, then earliest ready (floor).
#
#   Calendar starts 2026-06-02 (6/1 holiday), skips Sundays + holidays.
#   READ-ONLY. Nothing written to the ERP.
# =============================================================================
import json, os, re, importlib.util
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# ---- pull the fabric-SEWING schedule (its per-SO completion day is our floor) -
_spec = importlib.util.spec_from_file_location("sew", os.path.join(HERE, "build_sewing_daily_xlsx.py"))
sew = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(sew)
sew_end = {}                                   # coSOId -> last sewing working day
for _lane, _gs in sew.groups_by_lane.items():
    for _g in _gs:
        so = _g["so"]
        if so and (so not in sew_end or _g["end"] > sew_end[so]):
            sew_end[so] = _g["end"]

START = date(2026, 6, 2)
HOLIDAYS = {date(2026, 6, 1)}
CAP_SETS = {"BEDFRAME": 20, "SOFA": 10}        # sets/day per lane (Wei Siang)
WOOD_HANDOFF = 2                                # working days after fab-sew done = 1 CLEAR buffer day (Wei Siang: 隔一天才排;没料的日子空着,人调去别部门)
LANE_ORDER = ["BEDFRAME", "SOFA"]
CN = {"BEDFRAME": "床架 Bedframe", "SOFA": "沙发 Sofa"}

def is_offday(d):  return d.weekday() == 6 or d in HOLIDAYS
def next_workday(d):
    while is_offday(d): d += timedelta(days=1)
    return d
def step_workday(d): return next_workday(d + timedelta(days=1))
def add_workdays(d, n):
    d = next_workday(d)
    for _ in range(n): d = step_workday(d)
    return d
DAY1 = next_workday(START)
def day_num(d):
    i, x = 0, DAY1
    while x < d: i += 1; x = step_workday(x)
    return i + 1
def parse(s):
    try:
        y, m, dd = str(s)[:10].split("-"); return date(int(y), int(m), int(dd))
    except Exception: return None
FAR = "9999-99-99"
def _pe(s):  return parse(s) or date(9999, 12, 31)
THIS_MON = START - timedelta(days=START.weekday())

SIZE_SUFFIX = re.compile(r"-\((SS|SK|K|Q|S)\)\s*$")
def norm_model(r):
    m = (r.get("model") or "").strip()
    if r.get("cat") == "BEDFRAME":
        m = SIZE_SUFFIX.sub("", m)
    return m or "(none)"
def is_base_card(c, cat):
    wl = (c.get("wipLabel") or "").lower()
    if cat == "BEDFRAME": return "divan" in wl
    return "-base" in wl or "stool" in wl
def so_sets(cards, cat):
    base = [c for c in cards if is_base_card(c, cat)]
    if base: return sum(max(1, int(c.get("wipQty") or 1)) for c in base)
    return len({c.get("po") for c in cards})    # fallback: count lines

# ---- load WOOD_CUT WAITING, group into SO-units (per coSOId + cat) ------------
full = json.load(open(os.path.join(HERE, "sched_dept_full.json"), encoding="utf-8"))
src = [r for r in full.get("WOOD_CUT", [])
       if r.get("cardStatus") == "WAITING"
       and r.get("orderStatus") not in ("ON_HOLD", "CANCELLED")
       and r.get("cat") in ("BEDFRAME", "SOFA")]

by_unit = defaultdict(list)
for r in src:
    by_unit[(r.get("coSOId"), r.get("cat"))].append(r)

units = []
for (so, cat), cards in by_unit.items():
    sd = sew_end.get(so)
    floor = add_workdays(sd, WOOD_HANDOFF) if sd else DAY1
    cdds = sorted(c["cdd"] for c in cards if c.get("cdd"))
    models = sorted({norm_model(c) for c in cards})
    sizes = sorted({(c.get("size") or "").strip() for c in cards if c.get("size")})
    units.append({"so": so, "cat": cat, "cards": cards, "sets": so_sets(cards, cat),
                  "cdd": cdds[0] if cdds else "", "cdd_max": cdds[-1] if cdds else "",
                  "models": models, "sizes": sizes, "floor": floor,
                  "sew_done": sd, "model_key": models[0] if models else ""})

# ---- schedule: per lane, ready(floor)-first then DD, same SO whole, same model adj
lane_units = {}
for lane in LANE_ORDER:
    us = [u for u in units if u["cat"] == lane]
    us.sort(key=lambda u: (_pe(u["cdd"]), u["floor"], u["model_key"], u["so"] or ""))
    cap = CAP_SETS[lane]; load = defaultdict(int)
    for u in us:
        d = max(u["floor"], DAY1)
        while load[d] + u["sets"] > cap and load[d] != 0:
            d = step_workday(d)
        u["day"] = d; load[d] += u["sets"]
    lane_units[lane] = us

# ===================== styling ==============================================
ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
DATE_FILL = PatternFill("solid", fgColor="1F6F54"); DATE_FONT = F(bold=True, color="FFFFFF", size=10)
BF_FILL = PatternFill("solid", fgColor="DDEBF7")
SOFA_FILL = PatternFill("solid", fgColor="FFF2CC")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
LANE_FILL = {"BEDFRAME": BF_FILL, "SOFA": SOFA_FILL}

wb = Workbook()

# ---- Sheet 1: Wood Cut Calendar (interleaved by DATE) ----------------------
ws = wb.active; ws.title = "Wood Cut Calendar"
headers = ["Cut Date 裁日", "Lane 队", "SO ID", "Model 款 (不分尺寸)", "Sizes 尺寸",
           "Item 木工件", "件", "Sets 套", "Customer", "Customer DD", "车缝完成 Sew done"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

allu = []
for li, lane in enumerate(LANE_ORDER):
    for u in lane_units[lane]:
        allu.append((u["day"], li, lane, u))
allu.sort(key=lambda t: (t[0], t[1], t[3]["model_key"], t[3]["so"] or ""))

cur_date = None
for sday, li, lane, u in allu:
    if sday != cur_date:
        cur_date = sday
        ws.append([sday.strftime("%Y-%m-%d  %a") + f"   (Day {day_num(sday)})"])
        rr = ws.max_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
        ws.cell(row=rr, column=1).fill = DATE_FILL; ws.cell(row=rr, column=1).font = DATE_FONT
    models_txt = " + ".join(u["models"])
    sizes_txt = " / ".join(u["sizes"])
    sew_txt = u["sew_done"].strftime("%m-%d %a") if u["sew_done"] else "已完成/无车缝"
    for i, c in enumerate(u["cards"]):
        ws.append(["", CN[lane] if i == 0 else "", u["so"] if i == 0 else "",
                   models_txt if i == 0 else "", sizes_txt if i == 0 else "",
                   c.get("wipLabel") or c.get("model"), max(1, int(c.get("wipQty") or 1)),
                   u["sets"] if i == 0 else "", c.get("custName") if i == 0 else "",
                   c.get("cdd"), sew_txt if i == 0 else ""])
        rr = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=rr, column=col).fill = LANE_FILL[lane]; ws.cell(row=rr, column=col).border = BORDER
            ws.cell(row=rr, column=col).font = F(size=9)
        if parse(c.get("cdd")) and parse(c.get("cdd")) < THIS_MON:
            ws.cell(row=rr, column=10).fill = OVERDUE_FILL
widths = [20, 16, 15, 18, 14, 28, 5, 6, 16, 12, 16]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---- Sheet 2: By Day -------------------------------------------------------
ws2 = wb.create_sheet("By Day")
h2 = ["Date 日期", "Day", "Lane 队", "SOs that day 当天切的单 (model)", "SO数", "Sets 套", "Cap"]
ws2.append(h2)
perday = defaultdict(lambda: defaultdict(list))
for lane in LANE_ORDER:
    for u in lane_units[lane]:
        perday[lane][u["day"]].append(u)
for d in sorted({d for lane in LANE_ORDER for d in perday[lane]}):
    for lane in LANE_ORDER:
        if d not in perday[lane]: continue
        us = perday[lane][d]
        txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
        ws2.append([d.strftime("%Y-%m-%d %a"), day_num(d), CN[lane], txt,
                    len(us), sum(u["sets"] for u in us), CAP_SETS[lane]])
        rr = ws2.max_row
        for c in range(1, len(h2)+1):
            ws2.cell(row=rr, column=c).fill = LANE_FILL[lane]; ws2.cell(row=rr, column=c).border = BORDER
            ws2.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h2)+1):
    hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 16, 70, 6, 7, 6], 1): ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ---- Sheet 0: Summary & Notes ----------------------------------------------
ws0 = wb.create_sheet("Summary & Notes", 0)
spans = {}
for lane in LANE_ORDER:
    us = lane_units[lane]
    if us:
        spans[lane] = (min(u["day"] for u in us), max(u["day"] for u in us),
                       len(us), sum(u["sets"] for u in us), sum(len(u["cards"]) for u in us))
bydist = {l: sum(1 for u in units if u["cat"] == l) for l in LANE_ORDER}
notes = [
    ("木工裁料 逐日排程  Wood Cutting — Day-by-Day Schedule (重排版)", True, 14),
    ("生成 2026-06-01。只排『还没切(WAITING)』的木工卡,不含已完成、on-hold、已取消。只读,没改系统。", False, 10),
    ("", False, 10),
    ("【整条链:每段排在上一段做完之后】", True, 11),
    ("  裁布 →(完成日)→ 车缝 →(完成日)→ 木工。木工排在它那张单『车缝做完』之后才切。", False, 10),
    ("  起排底线 = 该 SO 车缝完成那天『不可以当天切』,要隔两个工作日,第三个工作日才开始切。", False, 10),
    ("  例:车缝 6/2(二) 完成 → 6/3、6/4 空 → 6/5(五) 才能切。车缝已做完/没车缝的,第一天就能切。", False, 10),
    ("", False, 10),
    ("【范围】", True, 11),
    (f"  还要切的单 = {len(units)} 个 SO  (床架 {bydist['BEDFRAME']} / 沙发 {bydist['SOFA']})。", False, 10),
    ("", False, 10),
    ("【同 SO 一起 + 同款不分尺寸】", True, 11),
    ("  · 同一个 SO 一定同一天切完,不拆开 —— 沙发的底座+扶手+靠垫(连左右件)全部一起。", False, 10),
    ("  · 床架 1013-(Q)/(K)/(SS)/(SK) 当一个款『1013』,所有尺寸一起切;沙发同款座宽一起。", False, 10),
    ("", False, 10),
    ("【产能:套/天】(1 套 = 一个 SO-line = 一张床 / 一个沙发)", True, 11),
    (f"  · 床架 {CAP_SETS['BEDFRAME']} 套/天,沙发 {CAP_SETS['SOFA']} 套/天 (对得上历史 completion)。", False, 10),
    ("  · 床架、沙发两条线分开、同时开工(不同人)。", False, 10),
    ("", False, 10),
    ("【排序】先看客期(Customer DD,谁先送货谁先切),再看哪张单的布车好了(底线)。", True, 11),
    ("", False, 10),
    ("【怎么看】", True, 11),
    ("  · Wood Cut Calendar:按日期排,绿条 = 那一天。最右栏『车缝完成』= 这张单要等到那天车好才切。", False, 10),
    ("  · By Day:一天一队一行,看当天切哪几个 SO、几套。", False, 10),
    ("", False, 10),
    ("【排出来的跨度】", True, 11),
]
for lane in LANE_ORDER:
    if lane in spans:
        a, b, nso, nsets, ncards = spans[lane]
        notes.append((f"  {CN[lane]:<16} {a.strftime('%m-%d %a')} → {b.strftime('%m-%d %a')}"
                      f"   共 {nso} 单 / {nsets} 套 / {ncards} 件", False, 10))
notes += [("", False, 10),
          ("  注:木工跟着车缝走;车缝表一改,木工底线跟着动,要重跑。只读 planning,还没写进 ERP。", False, 10)]
for i, (t, b, s) in enumerate(notes, 1):
    ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws0.column_dimensions["A"].width = 100

out = r"C:\Users\User\hookka-erp-testing\WoodCut-Daily-Schedule-2026-06-02.xlsx"
try: wb.save(out)
except PermissionError:
    out = r"C:\Users\User\hookka-erp-testing\WoodCut-Daily-Schedule-2026-06-02-v2.xlsx"
    wb.save(out)
print("OK ->", out)
print("wood-cut SO-units:", len(units), bydist, " (SOs matched to a sew-done day:",
      sum(1 for u in units if u["sew_done"]), ")")
for lane in LANE_ORDER:
    if lane in spans:
        a, b, nso, nsets, ncards = spans[lane]
        print(f"  {lane:<10} {a} -> {b}  ({nso} SOs, {nsets} sets, {ncards} cards)")
