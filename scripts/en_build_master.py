# =============================================================================
# MASTER PRODUCTION SCHEDULE — English report (presentation-only translation of
# build_master_schedule_xlsx.py). Same import pattern + same aggregation; only
# labels are English. Numbers/dates byte-identical.
# =============================================================================
import os, io, contextlib, importlib.util
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = r"C:\Users\User\Desktop\Production-Schedule-2026-06-02-EN"


def _load(name, fname):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, fname))
    mod = importlib.util.module_from_spec(spec)
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        spec.loader.exec_module(mod)
    return mod


fr = _load("fr_master_en", "build_framing_daily_xlsx.py")
sw = _load("sw_master_en", "build_sewing_daily_xlsx.py")
wd = fr.wood
ct = sw.cut

DAY1 = fr.DAY1
step_workday = fr.step_workday
day_num = fr.day_num

# ---- per-stage per-day aggregation ------------------------------------------
cut_day = defaultdict(lambda: defaultdict(int))
for lane, gs in ct.lane_groups.items():
    for g in gs:
        for (dd, cnt) in g.get("days", []):
            cut_day[dd][lane] += cnt

sew_day = {lane: dict(sw.loads.get(lane, {})) for lane in ("SOFA", "BEDFRAME", "ACCESSORY")}

wood_day = defaultdict(lambda: defaultdict(int))
for lane, us in wd.lane_units.items():
    for u in us:
        wood_day[u["day"]][lane] += u.get("sets", 0)

web_day = defaultdict(int); hbf_day = defaultdict(int)
for u in fr.units:
    web_day[u["day"]] += len(u.get("web", []))
    hbf_day[u["day"]] += len(u.get("hbf", []))

pack_day = defaultdict(int)
for u in fr.pack_units:
    pack_day[u["pack_day"]] += len(u.get("cards", []))

# ---- master day range --------------------------------------------------------
all_days = set()
for d in cut_day: all_days.add(d)
for lane in sew_day:
    for d in sew_day[lane]: all_days.add(d)
for d in wood_day: all_days.add(d)
for lane in fr.load:
    for d in fr.load[lane]: all_days.add(d)
for d in fr.foam_load: all_days.add(d)
for lane in fr.uph_load:
    for d in fr.uph_load[lane]: all_days.add(d)
for d in pack_day: all_days.add(d)
days = sorted(d for d in all_days if d >= DAY1)

# ===================== styling ===============================================
ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
GRP_FILL = PatternFill("solid", fgColor="305496"); GRP_FONT = F(bold=True, color="FFFFFF", size=9)
SUN_FILL = PatternFill("solid", fgColor="F2F2F2")
CUT_FILL = PatternFill("solid", fgColor="FCE4D6")
SEW_FILL = PatternFill("solid", fgColor="E2EFDA")
WOOD_FILL = PatternFill("solid", fgColor="FFF2CC")
FRAME_FILL = PatternFill("solid", fgColor="DDEBF7")
FOAM_FILL = PatternFill("solid", fgColor="FBE5D6")
UPH_FILL = PatternFill("solid", fgColor="E4DFEC")
PACK_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = F(bold=True, size=14)

wb = Workbook()

# ---- Sheet 1: Chain & Links -------------------------------------------------
ws1 = wb.active; ws1.title = "Chain & Links"
def hrs(mins): return f"{mins/60:.0f}h"
fr_bf = sum(u["fmin"] for u in fr.units if u["cat"] == "BEDFRAME")
fr_sf = sum(u["fmin"] for u in fr.units if u["cat"] == "SOFA")
uph_bf = sum(u["umin"] for u in fr.uph_units if u["cat"] == "BEDFRAME")
uph_sf = sum(u["umin"] for u in fr.uph_units if u["cat"] == "SOFA")
foam_ba = sum(u["ba_min"] for u in fr.foam_units)
rows = [
    ("Master Production Schedule", True, 14),
    ("Generated 2026-06-02. Calendar: starts 6/2, skips Sundays + the 6/1 holiday. Only cards not yet done (WAITING). Read-only, no ERP changes.", False, 10),
    ("All capacity figures verified against the live ERP (2026-06-02): production hours exact; caps are the owner's targets.", False, 10),
    ("", False, 10),
    ("[How the whole chain links together]", True, 12),
    ("", False, 10),
    ("1. Fabric Cut", True, 11),
    ("     Done first. By cut (a pool of 8 cuts/day, bedframe priority, sofa fills the gaps). Output feeds sewing.", False, 10),
    ("2. Fabric Sew", True, 11),
    ("     Start = the next working day after that order's fabric is cut. Capacity: sofa 35h, bedframe 20h, pillow 4h /day.", False, 10),
    ("3. Wood Cut", True, 11),
    ("     Start = next working day after that order's sewing is done (1 clearing day). Idle days left empty, people moved elsewhere. Capacity: bedframe 20 sets, sofa 10 sets/day.", False, 10),
    ("4. Framing  (+ Webbing + headboard HB Foam same day)", True, 11),
    (f"     Start = next working day after that order's wood is cut. Capacity by hours: bedframe {fr.CAP_HOURS['BEDFRAME']}h, sofa {fr.CAP_HOURS['SOFA']}h /day (two crews).", False, 10),
    ("     * Same day, riding along: this order's Webbing + headboard (HB) Foam Bonding happen the same day as framing (other crew, doesn't use framing hours).", False, 10),
    ("5. Sofa Foam Bonding  (sofa only)", True, 11),
    (f"     Start = next working day after sofa framing. Base + armrest by capacity {fr.FOAM_CAP_MIN//60}h/day; back cushion done the same day as the base (doesn't use those {fr.FOAM_CAP_MIN//60}h).", False, 10),
    ("6. Upholstery", True, 11),
    (f"     Bedframe upholstery = next day after framing (webbing done same day). Sofa upholstery = next day after base foam bonding.", False, 10),
    (f"     Capacity by hours, two crews: bedframe {fr.UPH_CAP_HOURS['BEDFRAME']}h, sofa {fr.UPH_CAP_HOURS['SOFA']}h /day.", False, 10),
    ("7. Packing", True, 11),
    ("     Date = the same day as that order's upholstery (due date follows upholstery). Not capacity-limited.", False, 10),
    ("", False, 10),
    ("[Shared rules]", True, 12),
    ("  - One SO is always finished on the same day, never split.", False, 10),
    ("  - Sort: Customer DD first (whoever ships first is done first), then who finished the previous stage first.", False, 10),
    ("  - If an upstream stage changes, the downstream floor moves and must be re-run. Changes are planning-only, not written to the ERP.", False, 10),
    ("", False, 10),
    ("[Total still to do (WAITING)]", True, 12),
    (f"  Sewing {hrs(sum(sw.loads[l].get(d,0) for l in sw.loads for d in sw.loads[l]))}  -  "
     f"Framing bedframe {hrs(fr_bf)} / sofa {hrs(fr_sf)}", False, 10),
    (f"  Sofa Foam (base+armrest) {hrs(foam_ba)}  -  Upholstery bedframe {hrs(uph_bf)} / sofa {hrs(uph_sf)}", False, 10),
]
for i, (t, b, s) in enumerate(rows, 1):
    ws1.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws1.column_dimensions["A"].width = 135

# ---- Sheet 2: Master Timeline -----------------------------------------------
ws2 = wb.create_sheet("Master Timeline")
groups = [
    ("", 0, None),
    ("Cut (cuts)", 1, CUT_FILL),
    ("Sew (h)", 2, SEW_FILL),
    ("Wood (sets)", 2, WOOD_FILL),
    ("Framing (h)", 2, FRAME_FILL),
    ("Same-day Web/HBfoam (pcs)", 2, FRAME_FILL),
    ("Sofa Foam (h)", 1, FOAM_FILL),
    ("Upholstery (h)", 2, UPH_FILL),
    ("Pack (pcs)", 1, PACK_FILL),
]
sub = ["Date", "Day", "Wk",
       "cuts",
       "Sofa", "Bedframe",
       "Bedframe", "Sofa",
       "Bedframe", "Sofa",
       "Web", "HBf",
       "Base+Arm",
       "Bedframe", "Sofa",
       "pcs"]
ws2.append([""] * len(sub))
col = 1
for label, span, fill in groups:
    if label == "":
        ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        c = ws2.cell(row=1, column=col, value="Date"); c.font = GRP_FONT; c.fill = GRP_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        col += 3
    else:
        ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        c = ws2.cell(row=1, column=col, value=label); c.font = GRP_FONT; c.fill = GRP_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col += span
ws2.append(sub)
for j in range(1, len(sub) + 1):
    c = ws2.cell(row=2, column=j); c.font = HEAD_FONT; c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def g0(d, m, div=1, dec=1):
    v = m.get(d, 0) / div
    if v == 0: return ""
    return round(v, dec) if dec else int(round(v))

for d in days:
    cut_t = sum(cut_day[d].values()) if d in cut_day else 0
    row = [
        d.strftime("%m-%d %a"), day_num(d), d.strftime("%a"),
        cut_t or "",
        g0(d, sew_day["SOFA"], 60), g0(d, sew_day["BEDFRAME"], 60),
        (wood_day[d]["BEDFRAME"] or "") if d in wood_day else "",
        (wood_day[d]["SOFA"] or "") if d in wood_day else "",
        g0(d, fr.load["BEDFRAME"], 60), g0(d, fr.load["SOFA"], 60),
        web_day.get(d, 0) or "", hbf_day.get(d, 0) or "",
        g0(d, fr.foam_load, 60),
        g0(d, fr.uph_load["BEDFRAME"], 60), g0(d, fr.uph_load["SOFA"], 60),
        pack_day.get(d, 0) or "",
    ]
    ws2.append(row)
    rr = ws2.max_row
    fills = [None, None, None, CUT_FILL, SEW_FILL, SEW_FILL, WOOD_FILL, WOOD_FILL,
             FRAME_FILL, FRAME_FILL, FRAME_FILL, FRAME_FILL, FOAM_FILL, UPH_FILL, UPH_FILL, PACK_FILL]
    for j in range(1, len(sub) + 1):
        cell = ws2.cell(row=rr, column=j)
        cell.border = BORDER; cell.font = F(size=9)
        cell.alignment = Alignment(horizontal="center")
        if fills[j - 1]: cell.fill = fills[j - 1]
    ws2.cell(row=rr, column=1).alignment = Alignment(horizontal="left")
widths = [11, 4, 4, 5, 6, 8, 8, 6, 8, 6, 5, 5, 9, 8, 6, 5]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "D3"

# ---- Sheet 3: Spans ----------------------------------------------------------
ws3 = wb.create_sheet("Spans")
h3 = ["Stage", "Start", "End", "Orders/batches", "Total", "Daily Cap", "Floor rule"]
ws3.append(h3)

def span_of(day_list):
    ds = [d for d in day_list if d]
    return (min(ds), max(ds)) if ds else (None, None)

cut_days = [d for d in cut_day]
sew_days = [d for l in sew_day for d in sew_day[l]]
wood_days = [d for d in wood_day]
fr_bf_days = [u["day"] for u in fr.units if u["cat"] == "BEDFRAME"]
fr_sf_days = [u["day"] for u in fr.units if u["cat"] == "SOFA"]
foam_days = [u["foam_day"] for u in fr.foam_units]
uph_bf_days = [u["uph_day"] for u in fr.uph_units if u["cat"] == "BEDFRAME"]
uph_sf_days = [u["uph_day"] for u in fr.uph_units if u["cat"] == "SOFA"]
pack_days = [u["pack_day"] for u in fr.pack_units]

spans = [
    ("Fabric Cut", cut_days, sum(len(g) for g in ct.lane_groups.values()), "-", "8 cuts/day", "done first"),
    ("Fabric Sew", sew_days, len(sw.cards),
     hrs(sum(sw.loads[l].get(d, 0) for l in sw.loads for d in sw.loads[l])),
     "sofa 35h / bedframe 20h /day", "next day after fabric cut"),
    ("Wood Cut", wood_days, sum(len(u) for u in wd.lane_units.values()),
     f"{sum(u['sets'] for us in wd.lane_units.values() for u in us)} sets", "bedframe 20 / sofa 10 sets/day", "next day after sewing (+1 clearing day)"),
    ("Framing Bedframe", fr_bf_days, len(fr_bf_days), hrs(fr_bf), f"{fr.CAP_HOURS['BEDFRAME']}h/day", "next day after wood cut"),
    ("Framing Sofa", fr_sf_days, len(fr_sf_days), hrs(fr_sf), f"{fr.CAP_HOURS['SOFA']}h/day", "next day after wood cut"),
    ("Sofa Foam Base+Arm", foam_days, len(fr.foam_units), hrs(foam_ba), f"{fr.FOAM_CAP_MIN//60}h/day", "next day after sofa framing"),
    ("Upholstery Bedframe", uph_bf_days, len(uph_bf_days), hrs(uph_bf), f"{fr.UPH_CAP_HOURS['BEDFRAME']}h/day", "next day after framing"),
    ("Upholstery Sofa", uph_sf_days, len(uph_sf_days), hrs(uph_sf), f"{fr.UPH_CAP_HOURS['SOFA']}h/day", "next day after base foam"),
    ("Packing", pack_days, len(fr.pack_units), "-", "no cap", "same day as upholstery"),
]
for name, dl, n, tot, cap, rule in spans:
    s, e = span_of(dl)
    ws3.append([name,
                s.strftime("%m-%d %a") if s else "-",
                e.strftime("%m-%d %a") if e else "-",
                n, tot, cap, rule])
    rr = ws3.max_row
    for j in range(1, len(h3) + 1):
        ws3.cell(row=rr, column=j).border = BORDER; ws3.cell(row=rr, column=j).font = F(size=9)
for j in range(1, len(h3) + 1):
    c = ws3.cell(row=1, column=j); c.font = HEAD_FONT; c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([20, 12, 12, 14, 9, 28, 36], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

out = os.path.join(OUT_DIR, "0-MASTER.xlsx")
wb.save(out)
print("OK ->", out)
print(f"days {days[0]} -> {days[-1]}  ({len(days)} working days)")
