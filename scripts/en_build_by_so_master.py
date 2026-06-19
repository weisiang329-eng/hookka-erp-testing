# =============================================================================
# BY-SO MASTER — one row per SOID, the whole 8-department chain in a single row
# (read-only, English). Imports the existing scheduler modules so every stage
# date is exactly the date that stage's own report shows. Numbers byte-identical.
#
# Owner's ask: "yi tao yi tao de by SOID" - for each SO see its full chain across
# all departments together in one row, with no fragmented / missing stage.
#
# Stage date per SO = the LAST scheduled working day for that SO in that stage
# (so a multi-line SO shows the day it actually clears the stage). A BLANK stage
# cell means that SO has no WAITING card for that stage = already done upstream,
# or not applicable for that product type (e.g. Sofa Foam Bonding only applies to
# sofas; Webbing rides framing).
# =============================================================================
import os, io, contextlib, importlib.util, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = r"C:\Users\User\Desktop\Production-Schedule-2026-06-02-EN"


def _load(name, fname):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, fname))
    mod = importlib.util.module_from_spec(spec)
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        spec.loader.exec_module(mod)
    return mod


fr = _load("fr_byso", "build_framing_daily_xlsx.py")
sw = _load("sw_byso", "build_sewing_daily_xlsx.py")
wd = fr.wood
ct = sw.cut
parse = fr.parse

# ---------------------------------------------------------------------------
# per-SO LAST scheduled day for each stage
# ---------------------------------------------------------------------------
# Cutting: items carry 'po' (= SOID + line suffix); raw src maps po -> soNo.
po_so = {r.get("po"): r.get("soNo") for r in ct.src}
cut_date = {}
for lane, gs in ct.lane_groups.items():
    for g in gs:
        gday = g["days"][-1][0] if g.get("days") else g.get("end")
        for it in g["items"]:
            so = po_so.get(it.get("po"))
            if not so:
                continue
            if so not in cut_date or gday > cut_date[so]:
                cut_date[so] = gday

sew_date = {}
for lane in sw.LANE_ORDER:
    for c in sw.results[lane]:
        so = c.get("soNo")
        if not so:
            continue
        if so not in sew_date or c["day"] > sew_date[so]:
            sew_date[so] = c["day"]

wood_date = {}
for lane, us in wd.lane_units.items():
    for u in us:
        so = u["so"]
        if so not in wood_date or u["day"] > wood_date[so]:
            wood_date[so] = u["day"]

fram_date = {}; web_date = {}
for u in fr.units:
    so = u["so"]
    if so not in fram_date or u["day"] > fram_date[so]:
        fram_date[so] = u["day"]
    if u.get("web"):                                   # webbing rides framing same day
        if so not in web_date or u["day"] > web_date[so]:
            web_date[so] = u["day"]

foam_date = {}
for u in fr.foam_units:
    so = u["so"]
    if so not in foam_date or u["foam_day"] > foam_date[so]:
        foam_date[so] = u["foam_day"]

uph_date = {}
for u in fr.uph_units:
    so = u["so"]
    if so not in uph_date or u["uph_day"] > uph_date[so]:
        uph_date[so] = u["uph_day"]

pack_date = {}
for u in fr.pack_units:
    so = u["so"]
    if so not in pack_date or u["pack_day"] > pack_date[so]:
        pack_date[so] = u["pack_day"]

# ---------------------------------------------------------------------------
# per-SO meta (customer, customer delivery date, category, models, qty) from the
# full dept snapshot so the row is informative even for SOs whose early stages
# are already done.
# ---------------------------------------------------------------------------
full = json.load(open(os.path.join(HERE, "sched_dept_full.json"), encoding="utf-8"))
cust = {}; cdd = {}; cat = defaultdict(set); models = defaultdict(set); ncards = defaultdict(int)
for dept, rows in full.items():
    for r in rows:
        so = r.get("coSOId")
        if not so:
            continue
        if r.get("custName") and so not in cust:
            cust[so] = r.get("custName")
        if r.get("cdd") and (so not in cdd or r["cdd"] < cdd[so]):
            cdd[so] = r["cdd"]
        if r.get("cat"):
            cat[so].add(r["cat"])
        m = (r.get("model") or "").strip()
        if m:
            models[so].add(m)

# WAITING card count per SO across the whole chain (quantity / size of the order)
# + per-(SO,dept) WAITING count so we can tell "stage already done" (blank, fine)
# from "stage has waiting work but no scheduled date" (a real scheduling miss).
waiting_by_dept = defaultdict(lambda: defaultdict(int))   # so -> dept -> WAITING count
for dept, rows in full.items():
    for r in rows:
        if r.get("cardStatus") == "WAITING" and r.get("orderStatus") not in ("ON_HOLD", "CANCELLED"):
            so = r.get("coSOId")
            if so:
                ncards[so] += 1
                waiting_by_dept[so][dept] += 1

# all real SOIDs that appear anywhere in the chain (drop blank SOID)
all_so = (set(cut_date) | set(sew_date) | set(wood_date) | set(fram_date)
          | set(web_date) | set(foam_date) | set(uph_date) | set(pack_date))
all_so = {s for s in all_so if s}

FAR = "9999-99-99"
def sortkey(so):
    return (cdd.get(so, FAR), so)

ordered = sorted(all_so, key=sortkey)

# ---------------------------------------------------------------------------
# workbook
# ---------------------------------------------------------------------------
ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
CUT_FILL = PatternFill("solid", fgColor="FCE4D6")
SEW_FILL = PatternFill("solid", fgColor="E2EFDA")
WOOD_FILL = PatternFill("solid", fgColor="FFF2CC")
FRAME_FILL = PatternFill("solid", fgColor="DDEBF7")
WEB_FILL = PatternFill("solid", fgColor="DDEBF7")
FOAM_FILL = PatternFill("solid", fgColor="FBE5D6")
UPH_FILL = PatternFill("solid", fgColor="E4DFEC")
PACK_FILL = PatternFill("solid", fgColor="D9E1F2")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
META_FILL = PatternFill("solid", fgColor="F2F2F2")

wb = Workbook()
ws = wb.active; ws.title = "By SO (full chain)"

stage_defs = [
    ("FabCut date", cut_date, CUT_FILL),
    ("FabSew date", sew_date, SEW_FILL),
    ("WoodCut date", wood_date, WOOD_FILL),
    ("Framing date", fram_date, FRAME_FILL),
    ("Webbing date", web_date, WEB_FILL),
    ("Foam date", foam_date, FOAM_FILL),
    ("Upholstery date", uph_date, UPH_FILL),
    ("Packing date", pack_date, PACK_FILL),
]
headers = (["SO ID", "Customer", "Item Category", "Customer Delivery Date"]
           + [name for name, _, _ in stage_defs]
           + ["Models", "WAITING cards", "Missing-stage note"])
ws.append(headers)
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def dstr(d):
    return d.strftime("%Y-%m-%d %a") if d else ""


# A blank stage cell is FINE if that stage is already done for the SO (no WAITING
# card) or doesn't apply to the product type. The only thing worth flagging is a
# stage that still has WAITING cards yet got no scheduled date - a real miss.
stage_order = [d for _, d, _ in stage_defs]
# which snapshot dept feeds each stage column (Webbing/Foam share FOAM/WEBBING)
stage_dept = ["FAB_CUT", "FAB_SEW", "WOOD_CUT", "FRAMING", "WEBBING", "FOAM",
              "UPHOLSTERY", "PACKING"]
for so in ordered:
    catset = cat.get(so, set())
    catstr = " + ".join(sorted(catset)) if catset else ""
    miss = []
    for i, (name, dmap, fill) in enumerate(stage_defs):
        if dmap.get(so):
            continue                                   # has a date - fine
        if waiting_by_dept[so].get(stage_dept[i], 0) > 0:
            # Foam column tracks SOFA base/arm only; a bedframe HB-foam WAITING card
            # rides framing day, so a bedframe-only SO legitimately has no sofa-foam
            # date. Don't flag Foam for non-sofa SOs.
            if name == "Foam date" and "SOFA" not in catset:
                continue
            miss.append(name)
    note = ("CHECK: " + ", ".join(miss)) if miss else ""

    row = [so, cust.get(so, ""), catstr, cdd.get(so, "")]
    for name, dmap, fill in stage_defs:
        row.append(dstr(dmap.get(so)))
    row.append(", ".join(sorted(models.get(so, set()))))
    row.append(ncards.get(so, 0))
    row.append(note)
    ws.append(row)
    rr = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=rr, column=c)
        cell.border = BORDER; cell.font = F(size=9)
    # meta tint
    for c in range(1, 5):
        ws.cell(row=rr, column=c).fill = META_FILL
    # stage tints (cols 5..12)
    for k, (name, dmap, fill) in enumerate(stage_defs):
        ws.cell(row=rr, column=5 + k).fill = fill
        ws.cell(row=rr, column=5 + k).alignment = Alignment(horizontal="center")
    # overdue customer date
    cd = parse(cdd.get(so))
    if cd and cd < fr.THIS_MON:
        ws.cell(row=rr, column=4).fill = OVERDUE_FILL
    if note:
        ws.cell(row=rr, column=len(headers)).fill = OVERDUE_FILL

widths = [15, 18, 16, 16, 15, 15, 15, 15, 15, 15, 15, 15, 26, 7, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "B2"

# ---------------------------------------------------------------------------
# Glossary sheet
# ---------------------------------------------------------------------------
wsG = wb.create_sheet("Glossary")
gloss = [
    ("By-SO Master - Glossary", True, 14),
    ("One row per SO (sales order). Each row shows that order's whole production chain across all 8 departments,", False, 10),
    ("so you can see each SO 'set by set' (yi tao yi tao) with no missing stage. Rows are sorted by Customer Delivery Date, earliest first.", False, 10),
    ("", False, 10),
    ("[Columns]", True, 12),
    ("SO ID", True, 10), ("   The sales-order number. Every department's cards for this order roll up into this one row.", False, 10),
    ("Customer", True, 10), ("   Who the order is for.", False, 10),
    ("Item Category", True, 10), ("   BEDFRAME and/or SOFA. An order can carry both (shown as 'BEDFRAME + SOFA').", False, 10),
    ("Customer Delivery Date", True, 10), ("   The promised ship date to the customer. Highlighted orange if already past. This is the sort key.", False, 10),
    ("", False, 10),
    ("[Stage date columns - the date that stage is scheduled to FINISH for this SO]", True, 12),
    ("FabCut date", True, 10), ("   Fabric Cutting - the order's fabric is cut. Done first.", False, 10),
    ("FabSew date", True, 10), ("   Fabric Sewing - covers/upholstery fabric sewn. Starts the working day after fabric is cut.", False, 10),
    ("WoodCut date", True, 10), ("   Wood Cutting - timber frame parts cut. Starts after sewing is done.", False, 10),
    ("Framing date", True, 10), ("   Framing - the frame is built. Starts the working day after wood cut.", False, 10),
    ("Webbing date", True, 10), ("   Webbing - done the SAME day as framing (bedframe rider). Blank for pure-sofa orders.", False, 10),
    ("Foam date", True, 10), ("   Sofa Foam Bonding (base + armrest) - the working day after sofa framing. SOFA orders only; blank for bedframe.", False, 10),
    ("Upholstery date", True, 10), ("   Upholstery - bedframe = day after framing; sofa = day after foam bonding.", False, 10),
    ("Packing date", True, 10), ("   Packing - same day as upholstery. Not capacity-limited.", False, 10),
    ("", False, 10),
    ("Models", True, 10), ("   The product models in this order.", False, 10),
    ("WAITING cards", True, 10), ("   How many production cards across all departments are still not done for this order (a rough size of the remaining work).", False, 10),
    ("Missing-stage note", True, 10), ("   'CHECK: ...' flags a stage that still has unfinished (WAITING) cards but did not get a scheduled date - worth a look. Blank = nothing to flag.", False, 10),
    ("", False, 10),
    ("[Why a stage cell can be blank - and that's usually fine]", True, 12),
    ("  - That stage is already DONE for this order (no WAITING card left) - so it isn't in this fresh plan. This is normal.", False, 10),
    ("  - That stage does not apply to this product type: Webbing is a bedframe step; Sofa Foam Bonding is a sofa step.", False, 10),
    ("  A blank is only flagged ('CHECK') when the order still has unfinished cards in that department yet no date came out - a genuine miss to review.", False, 10),
    ("", False, 10),
    ("Note: read-only planning from live ERP data (2026-06-02). Numbers/dates match the per-department reports exactly. Nothing written to the ERP.", False, 10),
]
for i, (t, b, s) in enumerate(gloss, 1):
    wsG.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
wsG.column_dimensions["A"].width = 130

out = os.path.join(OUT_DIR, "5-By-SO-Master.xlsx")
wb.save(out)

# ---- report -----------------------------------------------------------------
flagged = []
for so in ordered:
    catset = cat.get(so, set())
    for i, (name, dmap, fill) in enumerate(stage_defs):
        if dmap.get(so):
            continue
        if waiting_by_dept[so].get(stage_dept[i], 0) > 0:
            if name == "Foam date" and "SOFA" not in catset:
                continue
            flagged.append((so, name))
print("OK ->", out)
print("SO rows:", len(ordered))
print("SOs with a CHECK flag (unfinished stage but no scheduled date):", len({s for s, _ in flagged}))
for s, name in flagged:
    print("   ", s, cust.get(s, ""), "->", name)
