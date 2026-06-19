# =============================================================================
# FABRIC CUTTING — English report (presentation-only translation of
# build_cutting_daily_xlsx.py). Imports the ORIGINAL module so every number,
# date, sort order and capacity is byte-identical; only labels are English.
# =============================================================================
import os, importlib.util
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = r"C:\Users\User\Desktop\Production-Schedule-2026-06-02-EN"


def _load(name, fname):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, fname))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


ct = _load("cut_src", "build_cutting_daily_xlsx.py")

# pull the already-computed structures / helpers
LANE_ORDER = ct.LANE_ORDER
lane_groups = ct.lane_groups
src = ct.src
next_workday = ct.next_workday
START = ct.START
parse = ct.parse
THIS_MON = ct.THIS_MON
_fmt_cdd = ct._fmt_cdd
LANE_CONFIRMED = ct.LANE_CONFIRMED


def deadline_label(g):
    # English reimplementation of ct.deadline_label (original appends a Chinese
    # measure word for the part-batch count). Same numbers, English label.
    lo, hi = g.get("cdd_min", ""), g.get("cdd_max", "")
    if not lo and not hi:
        lbl = "no date"
    elif lo and hi and lo != hi:
        lbl = f"{_fmt_cdd(lo)} -> {_fmt_cdd(hi)}"
    else:
        lbl = _fmt_cdd(lo or hi)
    pa, pn = g.get("part", (1, 1))
    if pn > 1:
        lbl += f"  (cut {pa}/{pn})"
    return lbl

PULL_WINDOW = ct.PULL_WINDOW
SETUP_CAP = ct.SETUP_CAP
CHUNK_LEAD = ct.CHUNK_LEAD
RESERVE_TIERS = ct.RESERVE_TIERS
cap_for = ct.cap_for
pool_cap = ct.pool_cap
bf_used = ct.bf_used
LANE_CAP = ct.LANE_CAP
from collections import defaultdict

# English lane labels
EN = {"BEDFRAME": "Bedframe", "SOFA": "Sofa", "ACCESSORY": "Pillow"}

# ===================== styling =====================
ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
DATE_FILL = PatternFill("solid", fgColor="1F6F54"); DATE_FONT = F(bold=True, color="FFFFFF", size=10)
BF_FILL = PatternFill("solid", fgColor="DDEBF7")
SOFA_FILL = PatternFill("solid", fgColor="FFF2CC")
ACC_FILL = PatternFill("solid", fgColor="FCE4D6")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
LANE_FILL = {"BEDFRAME": BF_FILL, "SOFA": SOFA_FILL, "ACCESSORY": ACC_FILL}

wb = Workbook()

# ===================== Sheet 1: Cut Calendar (interleaved by DATE) ===========
ws = wb.active; ws.title = "Cut Calendar"
headers = ["Cut Date", "Lane", "Cut #", "Model", "Size", "SO / PO",
           "Customer", "Item (one set/row)", "Fabric", "Sets", "Slots",
           "Customer DD", "Our Expected DD", "Batch Deadline (earliest cust. DD)"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

allg = []
for li, lane in enumerate(LANE_ORDER):
    for g in lane_groups[lane]:
        allg.append((g["start"], li, lane, g))
allg.sort(key=lambda t: (t[0], t[1], t[3]["cut"]))

cur_date = None
for sdate, li, lane, g in allg:
    if sdate != cur_date:
        cur_date = sdate
        ws.append([sdate.strftime("%Y-%m-%d  %a") + f"   (Day {(sdate-next_workday(START)).days+1})"])
        rr = ws.max_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
        ws.cell(row=rr, column=1).fill = DATE_FILL; ws.cell(row=rr, column=1).font = DATE_FONT
    dstr = ""
    if g["start"] != g["end"]:
        dstr = "-> " + g["end"].strftime("%m-%d %a")
    for i, r in enumerate(g["items"]):
        ws.append([dstr if i == 0 else "", EN[lane] if i == 0 else "",
                   g["cut"] if i == 0 else "", g["config"] if i == 0 else "",
                   g["size"] if i == 0 else "", r["po"], r["cust"], r["label"], r["fabric"],
                   r["sets"], g["slots"] if i == 0 else "", r["cdd"], r["edd"],
                   (deadline_label(g) if i == 0 else "")])
        rr = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = LANE_FILL[lane]; ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).font = F(size=9)
        if parse(r["cdd"]) and parse(r["cdd"]) < THIS_MON:
            ws.cell(row=rr, column=12).fill = OVERDUE_FILL
widths = [22, 16, 6, 12, 8, 15, 18, 38, 13, 6, 7, 12, 14, 28]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ===================== Sheet 2: By Day (one row per lane-day) ================
ws2 = wb.create_sheet("By Day")
h2 = ["Date", "Day", "Lane", "Models that day (model-size x slots)",
      "Slots", "Orders", "Cap"]
ws2.append(h2)
perday = defaultdict(lambda: defaultdict(list))
for lane in LANE_ORDER:
    for g in lane_groups[lane]:
        for d, sl in g["days"]: perday[lane][d].append((g, sl))
for d in sorted({d for lane in LANE_ORDER for d in perday[lane]}):
    for lane in LANE_ORDER:
        if d not in perday[lane]: continue
        entries = perday[lane][d]
        models = ", ".join(f"{g['config']} ({g['size']})" + (f"x{sl}" if sl > 1 else "")
                           + (" (cont.)" if g["start"] != g["end"] and d != g["start"] else "")
                           for g, sl in entries)
        norders = sum(len(g["items"]) for g, sl in entries if d == g["start"])
        disp_cap = cap_for(lane, d)
        if lane == "SOFA":
            disp_cap = min(LANE_CAP["SOFA"], pool_cap(d) - bf_used.get(d, 0))
        ws2.append([d.strftime("%Y-%m-%d %a"), (d-next_workday(START)).days+1, EN[lane],
                    models, sum(sl for g, sl in entries), norders, disp_cap])
        rr = ws2.max_row
        for c in range(1, len(h2)+1):
            ws2.cell(row=rr, column=c).fill = LANE_FILL[lane]; ws2.cell(row=rr, column=c).border = BORDER
            ws2.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h2)+1):
    hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 16, 64, 7, 8, 6], 1): ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ===================== Sheet 0: Summary & Notes =============================
ws0 = wb.create_sheet("Summary & Notes", 0)
spans = {}
for lane in LANE_ORDER:
    gs = lane_groups[lane]
    if gs:
        spans[lane] = (min(g["start"] for g in gs), max(g["end"] for g in gs),
                       len(gs), sum(len(g["items"]) for g in gs))
ncards = len(src); bydist = {l: sum(1 for r in src if r.get("cat") == l) for l in LANE_ORDER}
notes = [
    ("Fabric Cutting - Day-by-Day Schedule", True, 14),
    ("Generated 2026-06-01 from a fresh live pull of your logged-in ERP. Schedules ONLY cards still to be cut (WAITING):", False, 10),
    ("excludes already-cut, on-hold and cancelled. Read-only - nothing written to the ERP.", False, 10),
    ("", False, 10),
    ("[Scope] matches the 114 rows on the page", True, 11),
    (f"  Cards still to cut = {ncards}  (Bedframe {bydist['BEDFRAME']} / Sofa {bydist['SOFA']} / Pillow {bydist['ACCESSORY']}).", False, 10),
    ("  - Drops already-cut (COMPLETED) cards + on-hold + cancelled (SO-2605-052).", False, 10),
    ("  - Pillows are all scheduled this time: live data shows they are genuinely not yet cut;", False, 10),
    ("    a pillow is cut to match its SO's customer date so the order ships complete.", False, 10),
    ("", False, 10),
    ("[What counts as one 'cut']", True, 11),
    ("  - By full configuration, not model: 5531-L(LHF)+2A(RHF) and 5531-2A(LHF)+L(RHF) are two different cuts.", False, 10),
    ("  - Bedframe 1013 K/Q/SS/S (6FT/5FT/3.5FT/3FT) are also different cuts.", False, 10),
    ("  - Same config + same size are batched together (different fabric is fine, kept in one batch).", False, 10),
    ("", False, 10),
    ("[How it's planned] merge same cuts where sensible, but don't force far-future ones together", True, 11),
    (f"  - Same config + same size: orders with customer dates within {PULL_WINDOW} days are pulled together into one cut.", False, 10),
    (f"    Same model due more than {PULL_WINDOW} days out waits until near its own date - don't cut too early, don't flood storage.", False, 10),
    (f"  - One cut holds up to {SETUP_CAP['BEDFRAME']} sets; bigger batches auto-split into 2 / 3 cuts.", False, 10),
    (f"    Split by customer date: urgent goes in cut #1 (ASAP), later ones in cut #2 placed {CHUNK_LEAD} working days before their own date,", False, 10),
    ("    so cut #2 isn't jammed next to cut #1 and fabric isn't cut too early. (See the 'Batch Deadline' column's 1/2, 2/2.)", False, 10),
    ("  - Priority follows each cut's OWN earliest Customer DD - a far-due cut isn't pulled forward just because the same model has an urgent order.", False, 10),
    ("  - Same model, different size with close dates naturally fall on the same day (Wood Cut ignores size; FC/sewing/WC keep a model together).", False, 10),
    ("    Same model with far-apart dates each sit near their own date, not forced onto one day.", False, 10),
    (f"  - Bedframe & sofa share the cutters. Combined per day: first 3 working days run {RESERVE_TIERS[0][1]} cuts (orders firmest),", False, 10),
    (f"    days 4-7 run {RESERVE_TIERS[1][1]} (keep slots for walk-in orders), day 8+ looser ({RESERVE_TIERS[2][1]}).", False, 10),
    ("    Bedframe scheduled first; on bedframe-light days sofa speeds up automatically, no need to wait for all bedframe.", False, 10),
    (f"  - The {PULL_WINDOW}-day pull window is tunable: want shorter (less storage) or longer (fewer setups)? Say the word.", False, 10),
    ("", False, 10),
    ("[How to read]", True, 11),
    ("  - Cut Calendar: by date, green band = that day; see what bedframe / sofa each cut on the same day.", False, 10),
    ("  - By Day: one row per lane per day, see which model-sizes each lane cuts.", False, 10),
    ("", False, 10),
    ("[Resulting spans]", True, 11),
]
for lane in LANE_ORDER:
    if lane in spans:
        a, b, ng, ni = spans[lane]
        tag = "" if LANE_CONFIRMED[lane] else "  (capacity TBC)"
        notes.append((f"  {EN[lane]:<10} {a.strftime('%m-%d %a')} -> {b.strftime('%m-%d %a')}"
                      f"   {ng} batches / {ni} cards{tag}", False, 10))
notes += [("", False, 10),
          ("  Note: planning made from read-only live data, not yet written to the ERP. Confirm the plan and I'll write it in.", False, 10)]
for i, (t, b, s) in enumerate(notes, 1):
    ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws0.column_dimensions["A"].width = 110

out = os.path.join(OUT_DIR, "1-FabricCutting.xlsx")
wb.save(out)
print("OK ->", out, " WAITING cards:", ncards, bydist)
