# =============================================================================
# FABRIC SEWING — day-by-day sew calendar  (read-only, no ERP writes)
# =============================================================================
# WHAT THIS IS
#   A calendar date stamped on every Fab Sewing card that still has to be done,
#   built ON TOP OF the fabric-cutting calendar so sewing never runs ahead of
#   the fabric being cut.
#
#   CAPACITY (measured from completion-dated FAB_SEW cards, by category):
#       SOFA     20 h/day  = 1200 min/day  (most-stable 7-day window 05-18..25)
#       BEDFRAME 17 h/day  = 1020 min/day  (recent 7 working days incl. clearing
#                                           batches; bedframe sews in bursts)
#     Two separate crews -> two separate daily buckets (sofa work can't fill a
#     bedframe-light day and vice versa).
#
#   PRIORITY: by CUSTOMER delivery date (cdd) first; ties keep the order's own
#   sew-floor / shortest job first so the day packs tight.
#
#   CUT -> SEW LINK (the whole point):
#       * fabric already cut (cut card done) OR no cut step  -> can sew from day 1
#       * fabric still to be cut  -> sew floor = (that order's LAST cut day in the
#         cutting calendar) + 1 working day handoff ("隔一天").
#     The cut day comes straight from build_cutting_daily_xlsx's schedule, keyed
#     by order line (po). An order line that needs several cuts waits for the
#     LAST one before its sewing can start.
#
#   COUPLING SIGNAL (feed the sewers):
#     We count, per day, minutes where a sew crew has spare capacity but every
#     remaining card is still floor-blocked waiting for its fabric to be cut.
#     Big sofa numbers here = "cut more sofa earlier so sofa sewing isn't idle".
#
#   Calendar starts 2026-06-02 (6/1 holiday), skips Sundays + holidays.
#   READ-ONLY. Nothing written to the ERP.
# =============================================================================
import json, os, importlib.util
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# ---- knobs ----------------------------------------------------------------
# Locked by Wei Siang from the 05-22..05-28 completion window (29th dropped):
#   SOFA 25 h/day, BEDFRAME 20 h/day. ACCESSORY 4 h/day provisional (pillow TBC).
CAP_MIN = {"SOFA": 2100, "BEDFRAME": 1200, "ACCESSORY": 240}  # min/day, FULL (sofa 35h/day per Wei Siang 2026-06-01)
CAP_HOURS = {k: v / 60 for k, v in CAP_MIN.items()}

# Pillow standard time: ERP stores 300 min/pillow (wrong, 20x over). Wei Siang
# (2026-06-01): a pillow sews in 15 min. Override ACCESSORY card time on read.
PILLOW_MIN_EACH = 15

# PACK EVERY DAY FULL (Wei Siang 2026-06-01): sewing packs each day to full
# capacity, no reserve buffer. Walk-in / insert orders are absorbed by OT, and
# the whole plan gets re-cut roughly every 3 days. (Cutting still keeps its own
# three-tier reserve — this change is sewing-only.)
RESERVE_TIERS = [(10 ** 9, 1.0)]


def cap_on(lane, d):
    """That working day's sew budget (minutes) for a crew, after the reserve
    tier for how far out the day is."""
    i = workday_index(d)
    frac = RESERVE_TIERS[-1][1]
    for upto, f in RESERVE_TIERS:
        if i <= upto:
            frac = f
            break
    return CAP_MIN[lane] * frac


HANDOFF_WORKDAYS = 1            # sew starts this many working days after the cut (next working day, 0 clear buffer days, owner directive 2026-06-01)
LANE_ORDER = ["BEDFRAME", "SOFA", "ACCESSORY"]
CN = {"BEDFRAME": "床架 Bedframe", "SOFA": "沙发 Sofa", "ACCESSORY": "配件 Pillow"}
LANE_CONFIRMED = {"BEDFRAME": True, "SOFA": True, "ACCESSORY": True}  # pillow 15min/each (Wei Siang)
FAR = "9999-99-99"

# ---- pull the cutting calendar (gives po -> last cut day + the helpers) ----
_spec = importlib.util.spec_from_file_location(
    "cut", os.path.join(HERE, "build_cutting_daily_xlsx.py"))
cut = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(cut)
START = cut.START
next_workday = cut.next_workday
step_workday = cut.step_workday
parse = cut.parse
workday_index = cut.workday_index

# po -> last scheduled cut day (an order line may be cut over several days /
# in several config groups; sewing waits for the LAST cut).
cut_last_day = {}
for lane, groups in cut.lane_groups.items():
    for g in groups:
        gday = g["days"][-1][0] if g.get("days") else g.get("end")
        for it in g.get("items", []):
            po = it.get("po")
            if not po:
                continue
            if po not in cut_last_day or gday > cut_last_day[po]:
                cut_last_day[po] = gday


def add_workdays(d, n):
    for _ in range(n):
        d = step_workday(d)
    return d


# ---- load FAB_SEW + FAB_CUT cards from the fresh dept snapshot -------------
SNAP = os.path.join(HERE, "sched_dept_full.json")
snap = json.load(open(SNAP, encoding="utf-8"))
sew_all = snap["FAB_SEW"]

# fabric-cut state per order line (from the snapshot's own FAB_CUT cards)
cut_state = {}
for c in snap["FAB_CUT"]:
    cut_state.setdefault(c["po"], []).append(c.get("cardStatus"))


def fabric_ready_floor(po):
    """Earliest working day this line's fabric is sewn-ready.
    Returns (floor_date, reason)."""
    sts = cut_state.get(po)
    cut_done = sts is not None and all(
        s in ("COMPLETED", "TRANSFERRED") for s in sts)
    if sts is None:
        return next_workday(START), "no-cut-step"
    if cut_done:
        return next_workday(START), "cut-done"
    # cut still pending -> wait for its scheduled cut day (+ handoff)
    cd = cut_last_day.get(po)
    if cd is None:
        # cut pending but not in our cutting plan -> assume cuttable soon
        return next_workday(START), "cut-pending-unplanned"
    return add_workdays(cd, HANDOFF_WORKDAYS), "after-cut"


# ---- build the sew backlog (one entry per WAITING sew card) ----------------
cards = []
for c in sew_all:
    if c.get("cardStatus") != "WAITING":
        continue
    cat = c.get("cat") or "OTHER"
    if cat not in CAP_MIN:
        cat = "ACCESSORY"
    # prodTime is PER-UNIT (ERP B3 fix 2026-05-11: JC total = prodTime x wipQty,
    # see department-performance.ts:299-302). Multiply by quantity for the real
    # crew time — 128 of the waiting cards are qty=2.
    qty = int(c.get("wipQty") or 1)
    mins = int(c.get("prodTime") or 0) * qty
    if cat == "ACCESSORY":
        mins = PILLOW_MIN_EACH * qty   # ERP 300/pillow is wrong; use 15 each
    floor, reason = fabric_ready_floor(c["po"])
    cards.append({
        "po": c["po"], "soNo": c.get("coSOId"), "cust": c.get("custName"),
        "cat": cat, "label": c.get("wipLabel") or c.get("wipCode") or "",
        "model": c.get("model"), "size": c.get("size"),
        "mins": mins, "qty": qty,
        "cdd": c.get("cdd") or FAR, "floor": floor, "reason": reason,
    })

# ---- schedule each crew: same SO together, by cdd, fill each day to capacity --
def schedule(lane, lane_cards):
    """Group the crew's cards by SO ID so a whole customer order sews together
    (Wei Siang "同个 SO ID 一定要一起"), then place each SO-group, earliest
    customer-DD first, into the earliest working day on/after its floor that can
    hold the whole group. Daily budget shrinks on later days (reserve tiers).
    A group bigger than one day's budget spills onto the next working day(s),
    still kept contiguous. Returns the cards (each tagged 'day') + daily load."""
    groups = defaultdict(list)
    for c in lane_cards:
        groups[c["soNo"]].append(c)
    glist = []
    for so, cs in groups.items():
        glist.append({
            "so": so, "cards": cs,
            "mins": sum(c["mins"] for c in cs),
            "floor": max(c["floor"] for c in cs),
            "cdd": min(c["cdd"] for c in cs),
        })
    glist.sort(key=lambda g: (g["cdd"], g["floor"].isoformat(), -g["mins"]))

    load = defaultdict(int)            # date -> minutes placed
    for g in glist:
        d = max(g["floor"], next_workday(START))
        # find earliest day that can take the WHOLE group; if the group is
        # bigger than a single day's budget, accept the first day with any room
        # and let it spill (handled below).
        while True:
            budget = cap_on(lane, d)
            if load[d] + g["mins"] <= budget:        # fits whole on this day
                break
            if load[d] == 0 and g["mins"] > budget:  # group bigger than a day
                break                                # start here, spill forward
            d = step_workday(d)
        # place the group's cards starting at d, spilling to next days if needed
        cur, room = d, cap_on(lane, d) - load[d]
        g["start"] = d
        for c in sorted(g["cards"], key=lambda c: -c["mins"]):
            while room < c["mins"] and load[cur] > 0:
                cur = step_workday(cur)
                room = cap_on(lane, cur) - load[cur]
            c["day"] = cur
            load[cur] += c["mins"]
            room -= c["mins"]
        g["end"] = cur
    cards_out = [c for g in glist for c in g["cards"]]
    return cards_out, load, glist


lane_cards = defaultdict(list)
for c in cards:
    lane_cards[c["cat"]].append(c)

results, loads, groups_by_lane = {}, {}, {}
for lane in LANE_ORDER:
    results[lane], loads[lane], groups_by_lane[lane] = schedule(lane, lane_cards[lane])

# ---- starvation / coupling signal -----------------------------------------
# For each lane+day: spare = cap - load. Of the cards NOT yet placeable that day
# (floor > day) but due soon, how many minutes are blocked waiting on cutting?
def starvation(lane):
    cs = results[lane]
    if not cs:
        return {}
    last = max(c["day"] for c in cs)
    out = {}
    d = next_workday(START)
    while d <= last:
        placed = loads[lane].get(d, 0)
        spare = cap_on(lane, d) - placed
        # minutes of work whose fabric isn't cut yet on day d (floor>d) but is
        # due on/before this day's horizon -> would have filled the idle crew
        blocked = sum(c["mins"] for c in cs
                      if c["floor"] > d and parse(c["cdd"]) and parse(c["cdd"]) <= d)
        if spare > 0 and blocked > 0:
            out[d] = (spare, blocked)
        d = step_workday(d)
    return out


if __name__ == "__main__":
    print("CUT->SEW linkage:")
    rc = defaultdict(int)
    for c in cards:
        rc[c["reason"]] += 1
    for k, v in sorted(rc.items()):
        print(f"   {k:24s} {v}")
    print()
    for lane in LANE_ORDER:
        cs = results[lane]
        if not cs:
            continue
        days = sorted({c["day"] for c in cs})
        late = sum(1 for c in cs if parse(c["cdd"]) and c["day"] > parse(c["cdd"]))
        print(f"{lane}: {len(cs)} cards  {days[0]} -> {days[-1]}  "
              f"({len(days)} working days)  cap {CAP_HOURS[lane]:.0f}h/day  "
              f"late_vs_cdd={late}")
        st = starvation(lane)
        if st:
            tot = sum(b for _, b in st.values())
            print(f"   STARVATION: {len(st)} days idle waiting on cutting, "
                  f"{tot} blocked-min ({tot/60:.1f}h) of due work")

    # ===================== build the workbook ==============================
    def F(**k):
        return Font(name="Calibri", **k)

    thin = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
    DATE_FILL = PatternFill("solid", fgColor="1F6F54"); DATE_FONT = F(bold=True, color="FFFFFF", size=10)
    BF_FILL = PatternFill("solid", fgColor="DDEBF7")
    SOFA_FILL = PatternFill("solid", fgColor="FFF2CC")
    ACC_FILL = PatternFill("solid", fgColor="FCE4D6")
    OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
    BLOCK_FILL = PatternFill("solid", fgColor="E2EFDA")
    LANE_FILL = {"BEDFRAME": BF_FILL, "SOFA": SOFA_FILL, "ACCESSORY": ACC_FILL}
    THIS_MON = next_workday(START)

    def tier_of(d):
        i = workday_index(d)
        for n, (upto, f) in enumerate(RESERVE_TIERS):
            if i <= upto:
                return ["packed 排满", "medium 中等", "loose 松散"][n]
        return "loose 松散"

    wb = Workbook()

    # ---- Sheet 1: Sew Calendar (interleaved by DATE) ----------------------
    ws = wb.active; ws.title = "Sew Calendar"
    headers = ["Lane 队", "SO ID", "Customer", "Model 款",
               "Size", "PO / Line", "Item 项目", "Qty", "Mins 分钟",
               "Customer DD 客期", "Fabric 布料状态"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # flatten every card, tagged with its sew day + lane order
    REASON_CN = {"cut-done": "已裁好", "no-cut-step": "无需裁(现成布)",
                 "cut-pending-unplanned": "待裁(很快)", "after-cut": "等裁料",
                 "cut-pending-planned": "等裁料"}
    allc = []
    for li, lane in enumerate(LANE_ORDER):
        for c in results[lane]:
            allc.append((c["day"], li, lane, c))
    allc.sort(key=lambda t: (t[0], t[1], t[3]["soNo"] or "", -t[3]["mins"]))

    cur_date = None; cur_so = None
    for sday, li, lane, c in allc:
        if sday != cur_date:
            cur_date = sday; cur_so = None
            placed = {l: loads[l].get(sday, 0) for l in LANE_ORDER}
            tag = "  ".join(f"{CN[l].split()[0]} {placed[l]/60:.1f}/{cap_on(l, sday)/60:.0f}h"
                            for l in LANE_ORDER if placed[l])
            ws.append([sday.strftime("%Y-%m-%d  %a")
                       + f"   (Day {workday_index(sday)})   {tag}"])
            rr = ws.max_row
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
            ws.cell(row=rr, column=1).fill = DATE_FILL; ws.cell(row=rr, column=1).font = DATE_FONT
        first = c["soNo"] != cur_so
        cur_so = c["soNo"]
        ws.append([CN[lane] if first else "", c["soNo"] if first else "",
                   c["cust"] if first else "", c["model"], c["size"], c["po"],
                   c["label"], c["qty"], c["mins"], c["cdd"],
                   REASON_CN.get(c["reason"], c["reason"])])
        rr = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=rr, column=col).fill = LANE_FILL[lane]
            ws.cell(row=rr, column=col).border = BORDER
            ws.cell(row=rr, column=col).font = F(size=9)
        if parse(c["cdd"]) and c["day"] > parse(c["cdd"]):
            ws.cell(row=rr, column=10).fill = OVERDUE_FILL
        if c["reason"] in ("after-cut",):
            ws.cell(row=rr, column=11).fill = BLOCK_FILL
    widths = [14, 14, 18, 14, 8, 14, 30, 5, 7, 13, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- Sheet 2: By Day (per lane-day load vs cap) -----------------------
    ws2 = wb.create_sheet("By Day")
    h2 = ["Date 日期", "Day", "Lane 队", "SOs that day 当天车的单",
          "Cards 张", "Load h 排了", "Cap h 产能", "Use% 用量", "Tier 档",
          "Cut-blocked h 等裁料"]
    ws2.append(h2)
    perday = defaultdict(lambda: defaultdict(list))
    for lane in LANE_ORDER:
        for c in results[lane]:
            perday[lane][c["day"]].append(c)
    alldays = sorted({d for lane in LANE_ORDER for d in perday[lane]})
    for d in alldays:
        for lane in LANE_ORDER:
            if d not in perday[lane]:
                continue
            cs = perday[lane][d]
            load = sum(c["mins"] for c in cs)
            cap = cap_on(lane, d)
            sos = sorted({c["soNo"] for c in cs})
            blocked = sum(c["mins"] for c in results[lane]
                          if c["floor"] > d and parse(c["cdd"]) and parse(c["cdd"]) <= d)
            ws2.append([d.strftime("%Y-%m-%d %a"), workday_index(d), CN[lane],
                        ", ".join(str(s) for s in sos), len(cs),
                        round(load / 60, 1), round(cap / 60, 1),
                        f"{load/cap*100:.0f}%" if cap else "",
                        tier_of(d), round(blocked / 60, 1) if blocked else ""])
            rr = ws2.max_row
            for col in range(1, len(h2) + 1):
                ws2.cell(row=rr, column=col).fill = LANE_FILL[lane]
                ws2.cell(row=rr, column=col).border = BORDER
                ws2.cell(row=rr, column=col).font = F(size=9)
            if blocked:
                ws2.cell(row=rr, column=10).fill = BLOCK_FILL
    for c in range(1, len(h2) + 1):
        hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate([18, 5, 16, 40, 6, 8, 8, 7, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ---- Sheet 0: Summary & Notes -----------------------------------------
    ws0 = wb.create_sheet("Summary & Notes", 0)
    spans = {}
    for lane in LANE_ORDER:
        gs = groups_by_lane[lane]
        if gs:
            spans[lane] = (min(g["start"] for g in gs), max(g["end"] for g in gs),
                           len(gs), len(results[lane]))
    rc = defaultdict(int)
    for c in cards:
        rc[c["reason"]] += 1
    overdue = {l: sum(1 for c in results[l] if parse(c["cdd"]) and parse(c["cdd"]) < THIS_MON)
               for l in LANE_ORDER}
    causedlate = {l: sum(1 for c in results[l] if parse(c["cdd"])
                         and parse(c["cdd"]) >= THIS_MON and c["day"] > parse(c["cdd"]))
                  for l in LANE_ORDER}
    notes = [
        ("车缝 逐日排程  Fabric Sewing — Day-by-Day Schedule", True, 14),
        ("生成 2026-06-01,接着裁料表来排 —— 布没裁好,车缝排不到那天。只读,没改系统。", False, 10),
        ("", False, 10),
        ("【产能怎么算】看过去有完成日的车缝卡,按队分开", True, 11),
        (f"  · 沙发 SOFA   {CAP_HOURS['SOFA']:.0f} 小时/天 ({CAP_MIN['SOFA']} 分钟)", False, 10),
        (f"  · 床架 BEDFRAME {CAP_HOURS['BEDFRAME']:.0f} 小时/天 ({CAP_MIN['BEDFRAME']} 分钟)", False, 10),
        (f"  · 配件 Pillow {PILLOW_MIN_EACH} 分钟/个 (你定的;ERP 那个 300 分钟/个是错的,已覆盖)。", False, 10),
        ("  取 05-22→05-28 那 7 个工作天的完成量算的(29号没算)。沙发床架两组人,各自一桶,不互相补。", False, 10),
        ("", False, 10),
        ("【排法】", True, 11),
        ("  · 先按 Customer DD(客期)排,客期早的先车;同客期就短工先车,把一天塞紧。", False, 10),
        ("  · 同一个 SO ID 一定排一起、车在连着的工作天,不拆开跳天(出货才配得齐)。", False, 10),
        (f"  · 一天排到产能上限,塞满就溢到隔天;一组比一天还大就连着车几天。", False, 10),
        (f"  · 每天排满,不留 buffer(你定的)。临时插单走 OT,大约每 3 天重排一次。", False, 10),
        ("", False, 10),
        ("【裁料→车缝 接得上】这是重点", True, 11),
        ("  · 布已经裁好 或 不用裁(现成布)→ 第一天就能车。", False, 10),
        (f"  · 布还没裁 → 车缝最早 = 那张单在裁料表的『最后一刀』+ 隔 {HANDOFF_WORKDAYS} 个工作天。", False, 10),
        ("    (绿色那栏 = 这张还在等布裁好才能车。)", False, 10),
        ("", False, 10),
        ("【裁料→车缝 接料情况(coupling)】", True, 11),
    ]
    for lane in LANE_ORDER:
        st = starvation(lane)
        if st:
            tot = sum(b for _, b in st.values()) / 60
            ds = sorted(st)
            notes.append((f"  · {CN[lane].split()[0]}:{len(st)} 天有人手却没布车,"
                          f"共 {tot:.1f} 小时该车的活卡在等裁料"
                          f"({ds[0].strftime('%m-%d')}→{ds[-1].strftime('%m-%d')})。", False, 10))
    notes += [("    → 这几天该把裁料多挤一点这一队,别让车缝组空等。要不要回头调裁料,你讲一声。", False, 10),
              ("", False, 10),
              ("【排出来的跨度 / 迟单】", True, 11)]
    for lane in LANE_ORDER:
        if lane in spans:
            a, b, ng, ni = spans[lane]
            tag = "" if LANE_CONFIRMED[lane] else "  (产能待确认)"
            notes.append((f"  {CN[lane]:<16} {a.strftime('%m-%d %a')} → {b.strftime('%m-%d %a')}"
                          f"   {ng} 个 SO / {ni} 张{tag}", False, 10))
            notes.append((f"      已经过期客期 {overdue[lane]} 张(布/产能限制,尽快车);"
                          f"  排程造成迟交 {causedlate[lane]} 张。", False, 10))
    notes += [("", False, 10),
              ("【怎么看】", True, 11),
              ("  · Sew Calendar:按日期排,绿色横条 = 那一天;每天沙发/床架各车哪些 SO 一眼看到。", False, 10),
              ("  · By Day:一天一队一行,看排了几小时 vs 产能、用了几成、还有多少活卡在等裁料。", False, 10),
              ("", False, 10),
              ("  注:只读线上资料做的 planning,还没写进 ERP。你确认排法 OK,我再写进系统。", False, 10)]
    for i, (t, b, s) in enumerate(notes, 1):
        ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
    ws0.column_dimensions["A"].width = 98

    for out in (r"C:\Users\User\hookka-erp-testing\Sewing-Daily-Schedule-2026-06-02.xlsx",
                r"C:\Users\User\hookka-erp-testing\Sewing-Daily-Schedule-2026-06-02-v2.xlsx",
                r"C:\Users\User\Documents\Sewing-Daily-Schedule-2026-06-02-NEW.xlsx"):
        try:
            wb.save(out); break
        except PermissionError:
            continue
    print("OK ->", out)
