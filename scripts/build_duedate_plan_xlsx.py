# =============================================================================
# Per-department PROCESS DUE DATE planning table  (read-only — no ERP writes)
# =============================================================================
# WHAT THIS IS
#   The owner asked for a planning book that looks EXACTLY like our ERP, but
#   with the planned process due date filled in:
#     "每一个部门只要有的 column 都开出来,然后你把它分成 8 个部门,分 8 个 sheet"
#   So: one sheet per department (8 sheets), and on each sheet we show every
#   column that the live ERP department production page shows — Customer PO ID,
#   Customer Ref, Customer, State, Company SO ID, Model, WIP, Size, Quantity,
#   Colour, Gap, Divan, Leg, Total, Special Order, Notes, the previous-process
#   completed-date column(s), that department's DD, PIC 1, PIC 2, Completed
#   Date, Overdue, Status (+ Raw Material Ready for Cutting, Rack for Packing).
#
#   To the LEFT of each row we add ONE extra column the ERP does not have:
#     "Plan Due" = the day our forward finite-capacity cascade scheduler
#     (dept_flow_scheduler.py) predicts that process can realistically be done,
#     given our real last-30-day daily output and the upstream hand-off.
#   The list on each sheet = the work that department still has to do (WAITING
#   cards), sorted earliest Plan Due first.
#
#   2026-06-01 is a no-work day (holiday) — the plan starts 2026-06-02.
#   READ-ONLY. Nothing is written back to the ERP.
#
# DATA
#   sched_data.json      — per order×dept scheduling aggregates (drives the algo)
#   sched_dept_full.json — per department, every job card with the full ERP
#                          column set (drives the display). Pulled read-only
#                          from the live ERP. Schema: { DEPT: [ {row}, ... ] }
# =============================================================================
import os, json
from datetime import date, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import dept_flow_scheduler as S

HERE = os.path.dirname(os.path.abspath(__file__))
FULL = os.path.join(HERE, "sched_dept_full.json")

ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
LATE_FILL = PatternFill("solid", fgColor="F8CBAD")
OVERDUE_FILL = PatternFill("solid", fgColor="E59866")
RISK_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="D5F5E3")
ONHOLD_FILL = PatternFill("solid", fgColor="FCF3CF")
DUE_FILL = PatternFill("solid", fgColor="EBF5FB")
BAND_A = PatternFill("solid", fgColor="DDEBF7")
BAND_B = PatternFill("solid", fgColor="FFFFFF")
OVERDUE_TXT = F(size=10, bold=True, color="C0392B")

# (sheet-tab name, dept code, label, that department's own DD header in the ERP)
DEPT_SHEETS = [
    ("1-Cut",   "FAB_CUT",    "裁布 Fabric Cut",   "Fab Cutting DD"),
    ("2-Sew",   "FAB_SEW",    "车缝 Fabric Sew",   "Fab Sewing DD"),
    ("3-Wood",  "WOOD_CUT",   "锯木 Wood Cut",     "Wood Cutting DD"),
    ("4-Foam",  "FOAM",       "粘棉 Foam",         "Foaming DD"),
    ("5-Frame", "FRAMING",    "组框 Framing",      "Framing DD"),
    ("6-Web",   "WEBBING",    "打底 Webbing",      "Webbing DD"),
    ("7-Uphol", "UPHOLSTERY", "包面 Upholstery",   "Upholstery DD"),
    ("8-Pack",  "PACKING",    "包装 Packing",      "Packing DD"),
]
DEPT_SHORT = {"FAB_CUT": "Cut", "FAB_SEW": "Sew", "WOOD_CUT": "Wood",
              "FOAM": "Foam", "FRAMING": "Frame", "WEBBING": "Web",
              "UPHOLSTERY": "Uphol", "PACKING": "Pack"}

# Departments that show a per-WIP Quantity column in the ERP
SHOW_QTY = {"FRAMING", "UPHOLSTERY", "PACKING"}

# Previous-process completed-date columns, exactly as the ERP labels them
PREV_CD = {
    "FRAMING":    [("WOOD_CUT",  "Wood Cutting CD")],
    "WEBBING":    [("FRAMING",   "Framing CD")],
    "UPHOLSTERY": [("FAB_SEW",   "Fab Sewing CD"),
                   ("FOAM",      "Foaming CD"),
                   ("FRAMING",   "Framing CD")],
    "PACKING":    [("UPHOLSTERY", "Upholstery CD")],
}


def datestr(v):
    """Trim ISO date/timestamp to YYYY-MM-DD; pass through anything else."""
    if not v:
        return ""
    s = str(v)
    return s[:10] if len(s) >= 10 and s[4] == "-" else s


def status_of(o, ready):
    cdd = S.parse(o.get("cdd")); hdd = S.parse(o.get("hdd"))
    if o.get("st") == "ON_HOLD":
        return "ON-HOLD", ONHOLD_FILL
    if cdd and cdd < S.START:
        return "OVERDUE", OVERDUE_FILL
    if ready and cdd and ready > cdd:
        return "LATE", LATE_FILL
    if ready and hdd and ready > hdd:
        return "AT-RISK", RISK_FILL
    return "ON-TIME", OK_FILL


def style_header(ws, n, row=1):
    for c in range(1, n + 1):
        hc = ws.cell(row=row, column=c)
        hc.font = HEAD_FONT; hc.fill = HEAD_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=5).coordinate
    ws.auto_filter.ref = f"A{row}:{get_column_letter(n)}{ws.max_row}"


def build_columns(dept, dd_header):
    """Return the ordered (header, key) column list for a department, mirroring
    the live ERP production page column-for-column, plus our Plan Due in front."""
    cols = [
        ("Plan Due\n计划到期", "_plan"),     # our prediction (only extra column)
        ("Sent", "_sent"),
        ("#", "_idx"),
        ("SO / PO", "po"),
        ("Customer PO ID", "custPOId"),
        ("Customer Ref", "custRef"),
        ("Customer", "custName"),
        ("State", "custState"),
        ("Company SO ID", "coSOId"),
        ("Model", "model"),
        ("WIP", "wipLabel"),
        ("Size", "size"),
    ]
    if dept in SHOW_QTY:
        cols.append(("Quantity", "wipQty"))
    cols += [
        ("Colour", "colour"),
        ("Gap", "gap"),
        ("Divan", "divan"),
        ("Leg", "leg"),
        ("Total", "total"),
        ("Special Order", "special"),
        ("Notes", "notes"),
    ]
    for code, label in PREV_CD.get(dept, []):
        cols.append((label, ("_prev", code)))
    cols.append((dd_header, "dd"))
    cols += [
        ("PIC 1", "pic1"),
        ("PIC 2", "pic2"),
        ("Completed Date", "completedDate"),
        ("Overdue", "_overdue"),
    ]
    if dept == "FAB_CUT":
        cols.append(("Raw Material Ready", "_blank"))
    if dept == "PACKING":
        cols.append(("Rack", "rack"))
    cols.append(("Status", "cardStatus"))
    return cols


def cell_value(r, key, plan_due):
    if isinstance(key, tuple):                       # ("_prev", deptcode)
        return datestr((r.get("prevCD") or {}).get(key[1]))
    if key == "_plan":
        return plan_due.strftime("%Y-%m-%d  %a") if plan_due else ""
    if key == "_sent":
        return datestr(r.get("distributedAt"))
    if key == "_idx" or key == "_blank":
        return ""
    if key == "_overdue":
        d = S.parse(datestr(r.get("dd")))
        return "OVERDUE" if (r.get("cardStatus") != "COMPLETED" and d and d < S.START) else ""
    if key in ("dd", "completedDate"):
        return datestr(r.get(key))
    return r.get(key, "")


def main():
    orders = S.load_orders()
    full = json.load(open(FULL, encoding="utf-8"))
    finish, start_at, load, load_models = S.schedule(orders)
    for o in orders:
        o["_ready"] = S.predicted_ready(o, finish)
        o["_status"], o["_fill"] = status_of(o, o["_ready"])
    by_po = {o["po"]: o for o in orders}

    wb = Workbook()
    ws0 = wb.active; ws0.title = "Summary & Notes"

    # ---------------- Overview: process due dates across all 8 depts ----------
    ws = wb.create_sheet("Process Due Dates")
    H = ["SO / PO", "Customer", "Model", "Type"] + \
        [DEPT_SHORT[d] for _, d, _, _ in DEPT_SHEETS] + \
        ["Ready", "Customer DD", "Hookka DD", "Status"]
    ws.append(H)
    rank = {"OVERDUE": 0, "LATE": 1, "AT-RISK": 2, "ON-HOLD": 3, "ON-TIME": 4}
    for o in sorted(orders, key=lambda x: (rank[x["_status"]], x["_ready"] or date(2099, 1, 1))):
        po = o["po"]
        row = [po, o.get("cust"), o.get("pc"), o.get("cat")]
        for _, d, _, _ in DEPT_SHEETS:
            cell = o["d"].get(d)
            if cell is None:
                row.append("-")
            elif cell[S.I_DONE] == 1:
                row.append("done")
            else:
                fi = finish[po].get(d)
                row.append(fi.strftime("%m-%d") if fi else "")
        row += [o["_ready"].strftime("%m-%d") if o["_ready"] else "",
                o.get("cdd"), o.get("hdd"), o["_status"]]
        ws.append(row); rr = ws.max_row
        for j in range(len(DEPT_SHEETS)):
            c = ws.cell(row=rr, column=5 + j)
            if c.value not in ("-", "", "done"):
                c.fill = DUE_FILL
            c.alignment = Alignment(horizontal="center")
        ws.cell(row=rr, column=len(H)).fill = o["_fill"]
    for i, w in enumerate([16, 18, 13, 9] + [8] * 8 + [9, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, len(H))

    # ---------------- 8 Department sheets (full ERP columns) ------------------
    dept_stats = {}
    for tab, dept, label, dd_header in DEPT_SHEETS:
        wsx = wb.create_sheet(tab)
        cols = build_columns(dept, dd_header)
        headers = [h for h, _ in cols]

        # title row 1
        wsx.cell(row=1, column=1,
                 value=f"{label}   —   工序 Plan Due 排程 (该部门要做的活,按计划到期先后;栏位跟 ERP 一样)")
        wsx.cell(row=1, column=1).font = F(bold=True, size=12, color="1F4E78")
        wsx.append(headers)   # row 2

        # worklist = WAITING cards for this dept
        rows = []
        for r in full.get(dept, []):
            if r.get("cardStatus") == "COMPLETED":
                continue
            po = r["po"]
            if po not in by_po:
                continue
            plan = finish.get(po, {}).get(dept)
            rows.append((plan or date(2099, 1, 1), r))
        rows.sort(key=lambda t: (t[0], (t[1].get("model") or ""), t[1]["po"]))
        dept_stats[dept] = (len(rows), sum(int(r.get("wipQty") or 1) for _, r in rows))

        # write, banding by plan-due day
        last_due = None; band = BAND_A; idx = 0
        for plan, r in rows:
            idx += 1
            if plan != last_due:
                band = BAND_B if band is BAND_A else BAND_A
                last_due = plan
            line = []
            for h, key in cols:
                v = "" if key == "_idx" else cell_value(r, key, plan if plan.year < 2099 else None)
                line.append(idx if key == "_idx" else v)
            wsx.append(line); rr = wsx.max_row
            o = by_po[r["po"]]
            for cc in range(1, len(cols) + 1):
                cell = wsx.cell(row=rr, column=cc)
                cell.fill = band; cell.border = BORDER; cell.font = F(size=9)
            wsx.cell(row=rr, column=1).fill = o["_fill"]      # Plan Due tinted by order status
            wsx.cell(row=rr, column=1).font = F(size=9, bold=True)
            # Overdue cell red
            for ci, (h, key) in enumerate(cols, 1):
                if key == "_overdue" and wsx.cell(row=rr, column=ci).value == "OVERDUE":
                    wsx.cell(row=rr, column=ci).font = OVERDUE_TXT

        # widths
        wmap = {"Plan Due\n计划到期": 16, "Sent": 11, "#": 5, "SO / PO": 15,
                "Customer PO ID": 14, "Customer Ref": 12, "Customer": 16, "State": 7,
                "Company SO ID": 14, "Model": 12, "WIP": 30, "Size": 8, "Quantity": 8,
                "Colour": 11, "Gap": 6, "Divan": 6, "Leg": 6, "Total": 6,
                "Special Order": 16, "Notes": 18, "PIC 1": 10, "PIC 2": 10,
                "Completed Date": 13, "Overdue": 9, "Raw Material Ready": 14,
                "Rack": 8, "Status": 10}
        for i, (h, _) in enumerate(cols, 1):
            wsx.column_dimensions[get_column_letter(i)].width = wmap.get(h, 13)
        style_header(wsx, len(cols), row=2)

    # ---------------- Summary & Notes ----------------------------------------
    cnt = defaultdict(int)
    for o in orders:
        cnt[o["_status"]] += 1
    notes = [
        ("各部门工序 Plan Due 排程  (Process Plan-Due, 栏位 = ERP 原样)", True, 14),
        ("生成 2026-06-01   6/1 放假不排工,计划从 6/2 开始。只读,没有改动系统任何东西。", False, 10),
        ("", False, 10),
        ("【这本表怎么看】", True, 11),
        ("  · Process Due Dates:一张单一行,横看 8 个部门各自要哪天做好(done=已做完,-=不需要)。", False, 10),
        ("  · 后面 8 页 = 8 个部门各一页(1-Cut … 8-Pack)。每页的栏位跟你 ERP 的部门生产页一模一样,", False, 10),
        ("      只在最左边多加一栏『Plan Due 计划到期』= 我们排出来这件活该哪天做好。", False, 10),
        ("  · 每页只列该部门『还没做完(WAITING)』的活,按 Plan Due 先后排,最上面的先做。", False, 10),
        ("      同一个计划到期日的用底色分组。最左 Plan Due 那格的颜色 = 整张单的状态(见下)。", False, 10),
        ("", False, 10),
        ("【Plan Due 怎么算出来的】", True, 11),
        ("  按近 30 天每天真实做得出来的量,一个部门接一个部门算:上一道工序哪天做完,下一道", False, 10),
        ("  才那天开工;每天只能做到产能上限;周日 + 6/1 放假不排。算出来那天 = 该工序最迟做好日。", False, 10),
        ("", False, 10),
        ("【状态 Status(最左 Plan Due 格的底色)】", True, 11),
        ("  OVERDUE 已逾期 / LATE 赶不上 Customer DD / AT-RISK 破内部缓冲 Hookka DD"
         " / ON-TIME 准时 / ON-HOLD 暂停。", False, 10),
        ("", False, 10),
        ("【结果一览】", True, 11),
        (f"  共 {len(orders)} 张活跃单。  OVERDUE={cnt['OVERDUE']}  LATE={cnt['LATE']}"
         f"  AT-RISK={cnt['AT-RISK']}  ON-TIME={cnt['ON-TIME']}  ON-HOLD={cnt['ON-HOLD']}", False, 10),
        ("", False, 10),
        ("【各部门:还要做的活 / 产能(近30天实际平均)】", True, 11),
    ]
    for _, d, label, _ in DEPT_SHEETS:
        n_cards, n_pieces = dept_stats.get(d, (0, 0))
        notes.append((f"  {label:<16} 待做 {n_cards:>4} 件 (WAITING) "
                      f"  产能 {S.CAP_MIN_PER_DAY[d]:>5} 分钟/天 ≈ {S.CAP_PIECES_PER_DAY[d]} 件/天",
                      False, 10))
    notes += [
        ("", False, 10),
        ("  瓶颈:FOAM 粘棉 818 分钟/天最低,整条线卡在这里。要把 Plan Due 提前,先在这里加人。", False, 10),
        ("", False, 10),
        ("  注:这只是 planning 表,还没写进 ERP。你看过、确认排法 OK,我再帮你把 due date 写进系统。", False, 10),
    ]
    for i, (t, b, s) in enumerate(notes, 1):
        ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
    ws0.column_dimensions["A"].width = 104

    out = r"C:\Users\User\hookka-erp-testing\Process-DueDate-Plan-2026-06-01.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        out = r"C:\Users\User\hookka-erp-testing\Process-DueDate-Plan-2026-06-01-v2.xlsx"
        wb.save(out)
    print("OK ->", out)
    print("status:", dict(cnt))
    print("dept WAITING:", {d: dept_stats[d] for _, d, _, _ in DEPT_SHEETS})


if __name__ == "__main__":
    main()
