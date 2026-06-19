# =============================================================================
# MASTER PRODUCTION SCHEDULE — whole chain in one workbook (read-only)
# =============================================================================
# Pulls the already-built per-stage schedules from the other scripts and lays
# the WHOLE factory out in one place:
#   Sheet 1  全链总览 Chain & Links   — how the 7 stages connect (floors/handoffs)
#   Sheet 2  每日总览 Master Timeline — every working day, every stage's load
#   Sheet 3  各阶段跨度 Spans          — start→end + totals + caps per stage
#
# THE CHAIN (Wei Siang, 2026-06):
#   裁布 Fab Cut → 车缝 Fab Sew → 木工 Wood Cut → Framing(+Webbing+HB-foam 同日)
#     → 沙发 Foam Bonding → 包工 Upholstery → 包装 Packing
# Nothing is written to the ERP. All numbers verified against live ERP 2026-06-02.
# =============================================================================
import os, io, contextlib, importlib.util
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

def _load(name, fname):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, fname))
    mod = importlib.util.module_from_spec(spec)
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        spec.loader.exec_module(mod)
    return mod

fr = _load("fr_master", "build_framing_daily_xlsx.py")   # framing+foam+upholstery+packing (imports wood)
sw = _load("sw_master", "build_sewing_daily_xlsx.py")     # sewing (imports cutting)
wd = fr.wood
ct = sw.cut

DAY1 = fr.DAY1
step_workday = fr.step_workday
day_num = fr.day_num

# ---- per-stage per-day aggregation ------------------------------------------
# Cutting: cuts/day per lane (cutting's native unit; pool cap 8 cuts/day)
cut_day = defaultdict(lambda: defaultdict(int))
for lane, gs in ct.lane_groups.items():
    for g in gs:
        for (dd, cnt) in g.get("days", []):
            cut_day[dd][lane] += cnt

# Sewing: minutes/day per lane
sew_day = {lane: dict(sw.loads.get(lane, {})) for lane in ("SOFA", "BEDFRAME", "ACCESSORY")}

# Wood: sets/day per lane
wood_day = defaultdict(lambda: defaultdict(int))
for lane, us in wd.lane_units.items():
    for u in us:
        wood_day[u["day"]][lane] += u.get("sets", 0)

# Framing: minutes/day per lane + same-day webbing / HB-foam riders
web_day = defaultdict(int); hbf_day = defaultdict(int)
for u in fr.units:
    web_day[u["day"]] += len(u.get("web", []))
    hbf_day[u["day"]] += len(u.get("hbf", []))

# Foam (sofa base+armrest): minutes/day  (fr.foam_load)
# Upholstery: minutes/day per lane (fr.uph_load) ; Packing: cards/day
pack_day = defaultdict(int)
for u in fr.pack_units:
    pack_day[u["pack_day"]] += len(u.get("cards", []))

# ---- master day range --------------------------------------------------------
all_days = set()
for d in cut_day: all_days.add(d)
for lane in sew_day:
    for d in sew_day[lane]: all_days.add(d)
for d in wood_day: all_days.add(d)
for lane in fr.load:
    for d in fr.load[lane]: all_days.add(d)
for d in fr.foam_load: all_days.add(d)
for lane in fr.uph_load:
    for d in fr.uph_load[lane]: all_days.add(d)
for d in pack_day: all_days.add(d)
days = sorted(d for d in all_days if d >= DAY1)

# ===================== styling ===============================================
ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
GRP_FILL = PatternFill("solid", fgColor="305496"); GRP_FONT = F(bold=True, color="FFFFFF", size=9)
SUN_FILL = PatternFill("solid", fgColor="F2F2F2")
CUT_FILL = PatternFill("solid", fgColor="FCE4D6")
SEW_FILL = PatternFill("solid", fgColor="E2EFDA")
WOOD_FILL = PatternFill("solid", fgColor="FFF2CC")
FRAME_FILL = PatternFill("solid", fgColor="DDEBF7")
FOAM_FILL = PatternFill("solid", fgColor="FBE5D6")
UPH_FILL = PatternFill("solid", fgColor="E4DFEC")
PACK_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = F(bold=True, size=14)

wb = Workbook()

# ---- Sheet 1: Chain & Links -------------------------------------------------
ws1 = wb.active; ws1.title = "全链总览 Chain & Links"
def hrs(mins): return f"{mins/60:.0f}h"
fr_bf = sum(u["fmin"] for u in fr.units if u["cat"] == "BEDFRAME")
fr_sf = sum(u["fmin"] for u in fr.units if u["cat"] == "SOFA")
uph_bf = sum(u["umin"] for u in fr.uph_units if u["cat"] == "BEDFRAME")
uph_sf = sum(u["umin"] for u in fr.uph_units if u["cat"] == "SOFA")
foam_ba = sum(u["ba_min"] for u in fr.foam_units)
rows = [
    ("生产全链排程总表  Master Production Schedule", True, 14),
    ("生成 2026-06-02。日历:6/2 起,跳星期日 + 6/1 假期。只排『还没做(WAITING)』的卡。只读,没改 ERP。", False, 10),
    ("所有产能数字已对实时 ERP 核对(2026-06-02):生产工时一分不差,cap 是老板定的目标。", False, 10),
    ("", False, 10),
    ("【整条链怎么环环相扣】", True, 12),
    ("", False, 10),
    ("1. 裁布 Fab Cut", True, 11),
    ("     最先做。按 cut(每天 8 cut 的池子,床架优先、沙发填空)。出料喂给车缝。", False, 10),
    ("2. 车缝 Fab Sew", True, 11),
    ("     起排 = 该单裁布好的『隔天』(下一个工作日)。产能:沙发 35h、床架 20h、配件 4h /天。", False, 10),
    ("3. 木工 Wood Cut", True, 11),
    ("     起排 = 该单车缝好的『隔天』(1 个清空日)。没料的日子留空,人调去别部门。产能:床架 20 套、沙发 10 套/天。", False, 10),
    ("4. Framing 做框  (+ Webbing + 床头HB Foam 同日)", True, 11),
    (f"     起排 = 该单木工切好的『隔天』。产能按工时:床架 {fr.CAP_HOURS['BEDFRAME']}h、沙发 {fr.CAP_HOURS['SOFA']}h /天(分两队)。", False, 10),
    ("     ★ 同一天连带做:这张单的 Webbing + 床头(HB)Foam Bonding 跟 framing 同日(另一批人,不占 framing 工时)。", False, 10),
    ("5. 沙发 Foam Bonding  (只沙发)", True, 11),
    (f"     起排 = 沙发 framing 的『隔天』。底座+扶手按产能 {fr.FOAM_CAP_MIN//60}h/天;后靠垫跟着底座同日做完(不占那 {fr.FOAM_CAP_MIN//60}h)。", False, 10),
    ("6. 包工 Upholstery", True, 11),
    (f"     床架包工 = framing 的『隔天』(webbing 同日好了)。沙发包工 = 底座 foam bonding 的『隔天』。", False, 10),
    (f"     产能按工时,分两队:床架 {fr.UPH_CAP_HOURS['BEDFRAME']}h、沙发 {fr.UPH_CAP_HOURS['SOFA']}h /天。", False, 10),
    ("7. 包装 Packing", True, 11),
    ("     日期 = 该单包工的同一天(due date 跟包工走)。不卡产能。", False, 10),
    ("", False, 10),
    ("【共同规则】", True, 12),
    ("  · 同一个 SO 一定同一天做完,不拆开。", False, 10),
    ("  · 排序:先看客期(Customer DD,谁先送货谁先做),再看上一站谁先做好。", False, 10),
    ("  · 上游一改,下游底线跟着动,要重跑。改动只在 planning,没写进 ERP。", False, 10),
    ("", False, 10),
    ("【待做总量(WAITING)】", True, 12),
    (f"  车缝 {hrs(sum(sw.loads[l].get(d,0) for l in sw.loads for d in sw.loads[l]))}  ·  "
     f"Framing 床架 {hrs(fr_bf)} / 沙发 {hrs(fr_sf)}", False, 10),
    (f"  沙发 Foam(底座+扶手) {hrs(foam_ba)}  ·  包工 床架 {hrs(uph_bf)} / 沙发 {hrs(uph_sf)}", False, 10),
]
for i, (t, b, s) in enumerate(rows, 1):
    ws1.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws1.column_dimensions["A"].width = 118

# ---- Sheet 2: Master Timeline -----------------------------------------------
ws2 = wb.create_sheet("每日总览 Master Timeline")
# two header rows: group band + sub columns
groups = [
    ("", 0, None),                                   # date block (3 cols)
    ("裁布 Cut(cut)", 1, CUT_FILL),
    ("车缝 Sew(h)", 2, SEW_FILL),
    ("木工 Wood(套)", 2, WOOD_FILL),
    ("Framing(h)", 2, FRAME_FILL),
    ("同日 Web/HBfoam(件)", 2, FRAME_FILL),
    ("沙发Foam(h)", 1, FOAM_FILL),
    ("包工 Upholstery(h)", 2, UPH_FILL),
    ("包装 Pack(件)", 1, PACK_FILL),
]
sub = ["日期 Date", "Day", "Wk",
       "cut",
       "沙发", "床架",
       "床架", "沙发",
       "床架", "沙发",
       "Web", "HBf",
       "底+扶",
       "床架", "沙发",
       "件"]
# group band row
ws2.append([""] * len(sub))
col = 1
band_cells = []
for label, span, fill in groups:
    if label == "":
        ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        c = ws2.cell(row=1, column=col, value="日期"); c.font = GRP_FONT; c.fill = GRP_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        col += 3
    else:
        ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        c = ws2.cell(row=1, column=col, value=label); c.font = GRP_FONT; c.fill = GRP_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col += span
# sub header row
ws2.append(sub)
for j in range(1, len(sub) + 1):
    c = ws2.cell(row=2, column=j); c.font = HEAD_FONT; c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def g0(d, m, div=1, dec=1):
    v = m.get(d, 0) / div
    if v == 0: return ""
    return round(v, dec) if dec else int(round(v))

for d in days:
    cut_t = sum(cut_day[d].values()) if d in cut_day else 0
    row = [
        d.strftime("%m-%d %a"), day_num(d), d.strftime("%a"),
        cut_t or "",
        g0(d, sew_day["SOFA"], 60), g0(d, sew_day["BEDFRAME"], 60),
        (wood_day[d]["BEDFRAME"] or "") if d in wood_day else "",
        (wood_day[d]["SOFA"] or "") if d in wood_day else "",
        g0(d, fr.load["BEDFRAME"], 60), g0(d, fr.load["SOFA"], 60),
        web_day.get(d, 0) or "", hbf_day.get(d, 0) or "",
        g0(d, fr.foam_load, 60),
        g0(d, fr.uph_load["BEDFRAME"], 60), g0(d, fr.uph_load["SOFA"], 60),
        pack_day.get(d, 0) or "",
    ]
    ws2.append(row)
    rr = ws2.max_row
    fills = [None, None, None, CUT_FILL, SEW_FILL, SEW_FILL, WOOD_FILL, WOOD_FILL,
             FRAME_FILL, FRAME_FILL, FRAME_FILL, FRAME_FILL, FOAM_FILL, UPH_FILL, UPH_FILL, PACK_FILL]
    for j in range(1, len(sub) + 1):
        cell = ws2.cell(row=rr, column=j)
        cell.border = BORDER; cell.font = F(size=9)
        cell.alignment = Alignment(horizontal="center")
        if fills[j - 1]: cell.fill = fills[j - 1]
    ws2.cell(row=rr, column=1).alignment = Alignment(horizontal="left")
widths = [11, 4, 4, 5, 6, 6, 6, 6, 6, 6, 5, 5, 6, 6, 6, 5]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "D3"

# ---- Sheet 3: Spans ----------------------------------------------------------
ws3 = wb.create_sheet("各阶段跨度 Spans")
h3 = ["阶段 Stage", "起 Start", "止 End", "单/批", "总量", "每天上限 Cap", "起排规则 Floor rule"]
ws3.append(h3)

def span_of(day_list):
    ds = [d for d in day_list if d]
    return (min(ds), max(ds)) if ds else (None, None)

cut_days = [d for d in cut_day]
sew_days = [d for l in sew_day for d in sew_day[l]]
wood_days = [d for d in wood_day]
fr_bf_days = [u["day"] for u in fr.units if u["cat"] == "BEDFRAME"]
fr_sf_days = [u["day"] for u in fr.units if u["cat"] == "SOFA"]
foam_days = [u["foam_day"] for u in fr.foam_units]
uph_bf_days = [u["uph_day"] for u in fr.uph_units if u["cat"] == "BEDFRAME"]
uph_sf_days = [u["uph_day"] for u in fr.uph_units if u["cat"] == "SOFA"]
pack_days = [u["pack_day"] for u in fr.pack_units]

spans = [
    ("裁布 Fab Cut", cut_days, sum(len(g) for g in ct.lane_groups.values()), "—", "8 cut/天", "最先做"),
    ("车缝 Fab Sew", sew_days, len(sw.cards),
     hrs(sum(sw.loads[l].get(d, 0) for l in sw.loads for d in sw.loads[l])),
     "沙发35h/床架20h/天", "裁布好的隔天"),
    ("木工 Wood Cut", wood_days, sum(len(u) for u in wd.lane_units.values()),
     f"{sum(u['sets'] for us in wd.lane_units.values() for u in us)} 套", "床架20/沙发10 套/天", "车缝好的隔天(留1清空日)"),
    ("Framing 床架", fr_bf_days, len(fr_bf_days), hrs(fr_bf), f"{fr.CAP_HOURS['BEDFRAME']}h/天", "木工切好的隔天"),
    ("Framing 沙发", fr_sf_days, len(fr_sf_days), hrs(fr_sf), f"{fr.CAP_HOURS['SOFA']}h/天", "木工切好的隔天"),
    ("沙发Foam 底座+扶手", foam_days, len(fr.foam_units), hrs(foam_ba), f"{fr.FOAM_CAP_MIN//60}h/天", "沙发framing的隔天"),
    ("包工 床架 Upholstery", uph_bf_days, len(uph_bf_days), hrs(uph_bf), f"{fr.UPH_CAP_HOURS['BEDFRAME']}h/天", "framing的隔天"),
    ("包工 沙发 Upholstery", uph_sf_days, len(uph_sf_days), hrs(uph_sf), f"{fr.UPH_CAP_HOURS['SOFA']}h/天", "底座foam的隔天"),
    ("包装 Packing", pack_days, len(fr.pack_units), "—", "不限", "跟包工同一天"),
]
for name, dl, n, tot, cap, rule in spans:
    s, e = span_of(dl)
    ws3.append([name,
                s.strftime("%m-%d %a") if s else "—",
                e.strftime("%m-%d %a") if e else "—",
                n, tot, cap, rule])
    rr = ws3.max_row
    for j in range(1, len(h3) + 1):
        ws3.cell(row=rr, column=j).border = BORDER; ws3.cell(row=rr, column=j).font = F(size=9)
for j in range(1, len(h3) + 1):
    c = ws3.cell(row=1, column=j); c.font = HEAD_FONT; c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([20, 12, 12, 7, 8, 18, 24], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ---- save -------------------------------------------------------------------
saved = None
for out in (r"C:\Users\User\hookka-erp-testing\Master-Production-Schedule-2026-06-02.xlsx",
            r"C:\Users\User\hookka-erp-testing\Master-Production-Schedule-2026-06-02-v2.xlsx",
            r"C:\Users\User\Documents\Master-Production-Schedule-2026-06-02-NEW.xlsx"):
    try:
        wb.save(out); saved = out; break
    except PermissionError:
        continue
print("OK ->", saved)
print(f"days {days[0]} -> {days[-1]}  ({len(days)} working days)")
for name, dl, n, tot, cap, rule in spans:
    s, e = span_of(dl)
    print(f"  {name:<22} {s} -> {e}   {n} units  {tot}")
