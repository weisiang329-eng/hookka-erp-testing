import json, os
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "planning_rows.jsonl")
rows = [json.loads(l) for l in open(DATA, encoding="utf-8") if l.strip()]
# row keys: s(soNo) c(customer) p(productCode) l(merged label) f(fabric)
#           t(category) w(system wipQty) d(customerDD) o(ourExpectedDD) x(status)
GEN_DATE = "2026-06-01"
FIRST_WD = "2026-06-02"

ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = F(bold=True, color="FFFFFF", size=11)
ONHOLD_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow
BUG_FILL = PatternFill("solid", fgColor="F8CBAD")      # orange = system shows wrong qty
ALT_FILL = PatternFill("solid", fgColor="F2F6FB")      # alternating model band

def model_of(p):
    p = (p or "").strip()
    return p.split("-")[0] if p else "(none)"

def cat_label(t):
    return {"SOFA": "Sofa", "BEDFRAME": "Bedframe", "ACCESSORY": "Accessory"}.get(t, t)

wb = Workbook()

# ===================== Sheet: Production Planning =====================
ws = wb.active
ws.title = "Production Planning"
headers = ["Model", "SO / CO No", "Customer", "Item (one set per row)",
           "Category", "Fabric", "Qty", "Customer DD", "Our Expected DD", "Status"]
ws.append(headers)

# sort: group by model, then by Our Expected DD, then SO No
srt = sorted(rows, key=lambda r: (model_of(r["p"]), r.get("o") or "9999", r.get("s") or ""))

band_for_model = {}
band_toggle = False
prev_model = None
for r in srt:
    m = model_of(r["p"])
    if m != prev_model:
        band_toggle = not band_toggle
        prev_model = m
    band_for_model_flag = band_toggle
    ws.append([m, r.get("s") or "", r.get("c") or "", r.get("l") or r.get("p") or "",
               cat_label(r.get("t")), r.get("f") or "", 1,
               r.get("d") or "", r.get("o") or "", r.get("x") or ""])
    rr = ws.max_row
    if band_for_model_flag:
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = ALT_FILL
    if r.get("x") == "ON_HOLD":
        ws.cell(row=rr, column=10).fill = ONHOLD_FILL

widths = [12, 16, 20, 46, 11, 14, 6, 13, 15, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c)
    hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        if cell.font is None or cell.font.name != ARIAL:
            cell.font = F(size=10)
        cell.border = BORDER
        if cell.column in (7,):  # Qty centered
            cell.alignment = Alignment(horizontal="center")
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# ===================== Sheet: Qty Bug Evidence =====================
ws2 = wb.create_sheet("Qty Bug (system mis-count)")
bug = [r for r in srt if (r.get("w") or 0) > 1]
h2 = ["Model", "SO / CO No", "Customer", "Item (one set per row)",
      "System shows (WRONG)", "Should be", "What the system counted"]
ws2.append(h2)
def why(r):
    if r.get("t") == "SOFA":
        return "Counted the merged compartments / pieces, not the set"
    if r.get("t") == "BEDFRAME":
        return "Counted fabric panels (e.g. Queen/King = 2 panels)"
    return "Counted internal pieces, not the order"
for r in bug:
    ws2.append([model_of(r["p"]), r.get("s") or "", r.get("c") or "",
                r.get("l") or r.get("p") or "", r.get("w"), 1, why(r)])
    rr = ws2.max_row
    ws2.cell(row=rr, column=5).fill = BUG_FILL
w2 = [12, 16, 20, 46, 20, 10, 46]
for i, w in enumerate(w2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for c in range(1, len(h2) + 1):
    hc = ws2.cell(row=1, column=c)
    hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center")
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(h2)):
    for cell in row:
        cell.font = F(size=10); cell.border = BORDER
        if cell.column in (5, 6):
            cell.alignment = Alignment(horizontal="center")
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{ws2.max_row}"

# ===================== Sheet: Summary & Notes (index 0) =====================
ws3 = wb.create_sheet("Summary & Notes", 0)
total = len(rows)
n_bug = len(bug)
by_cat = Counter(r.get("t") for r in rows)
n_models = len(set(model_of(r["p"]) for r in rows))
no_cust_dd = sum(1 for r in rows if not r.get("d"))
wdist = Counter(r.get("w") for r in rows)

notes = [
    ("生产计划表  Production Planning", True, 14),
    (f"生成日期 Generated: {GEN_DATE}    第一个开工天 First working day: {FIRST_WD}", False, 10),
    ("", False, 10),

    ("【这张表的规矩 The rules you set】", True, 11),
    ("  1. 数量一律算 1 —— 一行 = 一套 = 一件。订单本来每一行就是数量 1,这里全部照 1 走。", False, 10),
    ("  2. 一套不拆开 —— 沙发的角位 / 1A / 2A / 凳 / L 位… 已经併成一行(就是你切料时併好的那一排)。", False, 10),
    ("  3. 同款排在一起 —— 按款号(Model)分组排序,方便一次过看同款有几单。", False, 10),
    ("  4. 两个日期分两栏 —— Customer DD = 顾客要的交货日;Our Expected DD = 我们内部预计做好的日。", False, 10),
    ("     (这两个不一样,所以分开放,不会再混成一栏。)", False, 10),
    ("", False, 10),

    ("【这次发现的 Bug — 你问的「为什么 Quantity 会有 2、3、4」】", True, 11),
    ("  系统旧的「Qty」栏,拿错了数字:", False, 10),
    ("  · 它显示的不是『订单数量』,而是『切料要切几片 / 几个格』。", False, 10),
    ("  · 沙发:你把整套(角位+1A+2A+凳…)併成一排切,系统却把『併进来的格数』当成数量 ——", False, 10),
    ("        所以 5540-1A(LHF)+1NA+CSL+L(RHF) 这种一套,会显示成 4;有六个格的显示成 6。", False, 10),
    ("  · 床架:Queen / King 一张床要两片布,系统又把『布的片数 2』当成数量。", False, 10),
    ("  · 真相:这些其实全部都是『一套』『一张』—— 数量就是 1。你查的那张 D1 就是这样:只有一个数量,却被叫去切两片。", False, 10),
    (f"  · 全部 {total} 行里,有 {n_bug} 行被系统显示成多过 1。这 {n_bug} 行列在「Qty Bug」那一页,你可以逐张对。", False, 10),
    ("  · 这张计划表已经把它改回正确的 1。系统里那一栏要不要也跟着改,我先不动,等你点头(因为那个数字还连着贴纸张数、用布、工时,要一起想清楚)。", False, 10),
    ("", False, 10),

    ("【数字一览 Headline】", True, 11),
    (f"  在排的总行数 Active rows: {total}  (PENDING + ON_HOLD,做好的和取消的不算)", False, 10),
    (f"  款号种类 Models: {n_models}", False, 10),
    (f"  分类 By category: 床架 Bedframe {by_cat.get('BEDFRAME',0)} / 沙发 Sofa {by_cat.get('SOFA',0)} / 配件 Accessory {by_cat.get('ACCESSORY',0)}", False, 10),
    (f"  系统旧 Qty 分布 (本应全是 1): {dict(sorted(wdist.items()))}", False, 10),
    (f"  没填 Customer DD 的: {no_cust_dd} 行 (那张寄卖单 CO 没有顾客交期)", False, 10),
    ("", False, 10),

    ("【分页说明 Sheets】", True, 11),
    ("  · Production Planning:正式计划表 —— 数量全 1、一套一行、同款一起、两个日期分栏。", False, 10),
    ("  · Qty Bug (system mis-count):被系统数错的那几行,左边是系统现在显示的错数字,右边是正确的 1。", False, 10),
    ("", False, 10),
    ("  注:这张表只读了线上资料来做,没有改动系统任何东西。", False, 10),
]
for i, (txt, bold, size) in enumerate(notes, start=1):
    c = ws3.cell(row=i, column=1, value=txt)
    c.font = F(bold=bold, size=size)
ws3.column_dimensions["A"].width = 100

out = r"C:\Users\User\hookka-erp-testing\Production-Planning-2026-06-02.xlsx"
try:
    wb.save(out)
except PermissionError:
    out = r"C:\Users\User\hookka-erp-testing\Production-Planning-2026-06-02-v2.xlsx"
    wb.save(out)
print("OK total=", total, "bug_rows=", n_bug, "models=", n_models, "->", out)
