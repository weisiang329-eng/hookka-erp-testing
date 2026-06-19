import json, os
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "planning_rows.jsonl")
rows = [json.loads(l) for l in open(DATA, encoding="utf-8") if l.strip()]
# keys: s(soNo) c(customer) p(productCode) l(merged label) f(fabric)
#       t(category) w(system wipQty) d(customerDD) o(ourExpectedDD) x(status)
TODAY = date(2026, 6, 1)
THIS_MON = TODAY - timedelta(days=TODAY.weekday())   # Monday of current week
FAR = "9999-99-99"

# ---- balancing knob: batch same-model orders only inside the SAME window ----
# WINDOW = "week" groups orders due in the same Mon-Sun week. Tighten to a few
# days or loosen to 2 weeks by changing WINDOW_DAYS.
WINDOW_DAYS = 7

ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = F(bold=True, color="FFFFFF", size=11)
BAND_A = PatternFill("solid", fgColor="DDEBF7")
BAND_B = PatternFill("solid", fgColor="FFFFFF")
OVERDUE_HDR = PatternFill("solid", fgColor="F8CBAD")
ONHOLD_FILL = PatternFill("solid", fgColor="FFF2CC")

def model_of(p):
    p = (p or "").strip()
    return p.split("-")[0] if p else "(none)"

def cat_label(t):
    return {"SOFA": "Sofa", "BEDFRAME": "Bedframe", "ACCESSORY": "Accessory"}.get(t, t)

def parse(s):
    try:
        y, m, d = (s or "").split("-")
        return date(int(y), int(m), int(d))
    except Exception:
        return None

def window_key(o):
    """Bucket by deadline window. Past/this-week -> bucket 0 (OVERDUE/this week).
    Future weeks -> 1,2,3..."""
    dt = parse(o)
    if dt is None:
        return 9999
    if dt < THIS_MON:
        return 0
    wk_mon = dt - timedelta(days=dt.weekday())
    return ((wk_mon - THIS_MON).days // WINDOW_DAYS) + 1

def window_label(wk):
    if wk == 0:
        return "0) OVERDUE / this week — cut first"
    if wk == 9999:
        return "9) no date"
    start = THIS_MON + timedelta(days=(wk - 1) * WINDOW_DAYS)
    end = start + timedelta(days=WINDOW_DAYS - 1)
    return f"{wk}) {start.strftime('%d %b')} – {end.strftime('%d %b')}"

# ---- group orders: (window, model) ----
buckets = defaultdict(lambda: defaultdict(list))   # window -> model -> [rows]
for r in rows:
    buckets[window_key(r["o"])][model_of(r["p"])].append(r)

# build ordered list of (window, model) cut-groups
cut_groups = []   # (window, model, items)
for wk in sorted(buckets.keys()):
    models = buckets[wk]
    # within a window, models ordered by their earliest Our Expected DD
    def grp_urgency(m):
        ods = [it["o"] for it in models[m] if it.get("o")]
        return min(ods) if ods else FAR
    for m in sorted(models.keys(), key=lambda m: (grp_urgency(m), m)):
        items = sorted(models[m], key=lambda r: (r.get("o") or FAR, r.get("d") or FAR, r.get("s") or ""))
        cut_groups.append((wk, m, items))

wb = Workbook()

# ===================== Sheet 1: Cut Order =====================
ws = wb.active
ws.title = "Cut Order"
headers = ["Window (deadline)", "Cut #", "Model", "SO / CO No", "Customer",
           "Item (one set per row)", "Fabric", "Qty", "Customer DD",
           "Our Expected DD", "Status"]
ws.append(headers)

cut_no = 0
for (wk, m, items) in cut_groups:
    cut_no += 1
    band = BAND_A if (cut_no % 2 == 1) else BAND_B
    for r in items:
        ws.append([window_label(wk), cut_no, m, r.get("s") or "", r.get("c") or "",
                   r.get("l") or r.get("p") or "", r.get("f") or "", 1,
                   r.get("d") or "", r.get("o") or "", r.get("x") or ""])
        rr = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = band
        if wk == 0:
            ws.cell(row=rr, column=1).fill = OVERDUE_HDR
        if r.get("x") == "ON_HOLD":
            ws.cell(row=rr, column=11).fill = ONHOLD_FILL

widths = [26, 6, 12, 16, 20, 44, 14, 6, 13, 15, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c)
    hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        if cell.font is None or cell.font.name != ARIAL:
            cell.font = F(size=10)
        cell.border = BORDER
        if cell.column in (2, 8):
            cell.alignment = Alignment(horizontal="center")
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# ===================== Sheet 2: Batch Summary =====================
ws2 = wb.create_sheet("Batch Summary")
h2 = ["Cut #", "Window (deadline)", "Model", "Category", "# Orders",
      "Earliest Customer DD", "Earliest Our Expected DD", "Latest Our Expected DD"]
ws2.append(h2)
cut_no = 0
for (wk, m, items) in cut_groups:
    cut_no += 1
    cds = [it["d"] for it in items if it.get("d")]
    ods = [it["o"] for it in items if it.get("o")]
    ws2.append([cut_no, window_label(wk), m, cat_label(items[0].get("t")), len(items),
                min(cds) if cds else "", min(ods) if ods else "", max(ods) if ods else ""])
    rr = ws2.max_row
    if wk == 0:
        ws2.cell(row=rr, column=2).fill = OVERDUE_HDR
w2 = [6, 26, 12, 11, 9, 18, 22, 22]
for i, w in enumerate(w2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for c in range(1, len(h2) + 1):
    hc = ws2.cell(row=1, column=c)
    hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center")
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(h2)):
    for cell in row:
        cell.font = F(size=10); cell.border = BORDER
        if cell.column in (1, 5):
            cell.alignment = Alignment(horizontal="center")
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{ws2.max_row}"

# ===================== Sheet 3: Summary & Notes (index 0) =====================
ws3 = wb.create_sheet("Summary & Notes", 0)
total = len(rows)
n_groups = len(cut_groups)
n_windows = len(set(wk for wk, _, _ in cut_groups))
# split-model example: a model that appears in >1 window
appear = defaultdict(set)
for wk, m, _ in cut_groups:
    appear[m].add(wk)
split_models = sorted([m for m, wks in appear.items() if len(wks) > 1])
notes = [
    ("裁料排程  Cutting Schedule (鱼与熊掌版)", True, 14),
    (f"生成日期 Generated: 2026-06-01    批次窗口 Window: 每 {WINDOW_DAYS} 天（一周）", False, 10),
    ("", False, 10),

    ("【这次的排法 —— 又快又不让别款 overdue】", True, 11),
    ("  你讲得对:不能把一个款『全部』单一次过剪完,不然 1007 剪完 36 张,别的款全 overdue 了。", False, 10),
    ("  所以这版是这样平衡的:", False, 10),
    ("  1. 先按『交期那一周』分段 —— 这周该交的先剪,下周的下周剪,不会把下个月的拉上来抢产能。", False, 10),
    ("  2. 在『同一周』里面,才把同款的单併在一起剪(省一次开料设定),剪得快。", False, 10),
    ("  3. 所以一个款(例如 1007)的单会按它各自的交期,分散到各周;不会全部塞在最前面。", False, 10),
    ("  4. 最前面是「OVERDUE / 这周」那一段 —— 已经到期 / 这周到期的,全部先剪(橙色)。", False, 10),
    ("  5. 数量一律 1。", False, 10),
    ("", False, 10),

    ("【日期 —— 用回你系统的两个真实日期】", True, 11),
    ("  · Customer DD = 顾客交货日（跟系统一样）。Our Expected DD = 我们内部预计做好日（跟系统一样）。", False, 10),
    ("  · 分周、排先后,都是看 Our Expected DD（我们自己的目标日）。", False, 10),
    ("  · 已经把上一版我自己算的『Planned Ready』拿掉了,只剩你系统里的真实日期。", False, 10),
    ("", False, 10),

    ("【数字一览】", True, 11),
    (f"  在排总行数 Active rows: {total}", False, 10),
    (f"  裁料窗口段数 Windows: {n_windows}    裁料批次数 Cut groups: {n_groups}", False, 10),
    (f"  会分到多周剪的款(没有全塞在前面)Split across weeks: {', '.join(split_models) if split_models else '—'}", False, 10),
    ("", False, 10),

    ("【可调 —— 一句话就能改】", True, 11),
    ("  · 现在窗口是『一周』。想剪得更集中(同款拉更多张一起)→ 改成两周;想更贴紧交期 → 改成几天。你说一声我就调。", False, 10),
    ("  · 如果你要一张『每天剪哪几个款』的逐日表(用 床架 6 款/天、沙发 2 款/天 的产能算),我也能出 —— 那张会有我算的日期,你之前不太信那种,所以先做这版只用真实交期的。", False, 10),
    ("", False, 10),

    ("【分页】", True, 11),
    ("  · Cut Order：逐张单,按裁料先后;同款同色一条带。Window 栏 = 该周交期段。", False, 10),
    ("  · Batch Summary：每个裁料批次一行 —— 第几个剪、哪一周、哪个款、几张单。先看这页定先后。", False, 10),
    ("", False, 10),
    ("  注:只读了线上资料来做,没有改动系统任何东西。", False, 10),
]
for i, (txt, bold, size) in enumerate(notes, start=1):
    c = ws3.cell(row=i, column=1, value=txt)
    c.font = F(bold=bold, size=size)
ws3.column_dimensions["A"].width = 102

out = r"C:\Users\User\hookka-erp-testing\Cutting-Schedule-2026-06-02.xlsx"
try:
    wb.save(out)
except PermissionError:
    out = r"C:\Users\User\hookka-erp-testing\Cutting-Schedule-2026-06-02-v2.xlsx"
    wb.save(out)
print("OK total=", total, "groups=", n_groups, "windows=", n_windows,
      "split_models=", split_models, "->", out)
