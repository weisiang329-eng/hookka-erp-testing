# =============================================================================
# Report builder for the multi-department forward schedule.
# Pure presentation — all scheduling logic lives in dept_flow_scheduler.py
# (the Agent core). This file just runs it and renders an .xlsx the owner reads.
# =============================================================================
import os
from datetime import date
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import dept_flow_scheduler as S

ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = F(bold=True, color="FFFFFF", size=11)
LATE_FILL = PatternFill("solid", fgColor="F8CBAD")      # red-ish
OVERDUE_FILL = PatternFill("solid", fgColor="E59866")   # darker
RISK_FILL = PatternFill("solid", fgColor="FFF2CC")      # amber
OK_FILL = PatternFill("solid", fgColor="D5F5E3")        # green
ONHOLD_FILL = PatternFill("solid", fgColor="FCF3CF")

DEPT_SHORT = {"FAB_CUT": "Cut", "FAB_SEW": "Sew", "WOOD_CUT": "Wood",
              "FOAM": "Foam", "FRAMING": "Frame", "WEBBING": "Web",
              "UPHOLSTERY": "Uphol", "PACKING": "Pack"}


def status_of(o, ready):
    cdd = S.parse(o.get("cdd")); hdd = S.parse(o.get("hdd"))
    if o.get("st") == "ON_HOLD":
        return "ON-HOLD", ONHOLD_FILL
    if cdd and cdd < S.START:
        return "OVERDUE", OVERDUE_FILL
    if ready and cdd and ready > cdd:
        return "LATE", LATE_FILL
    if ready and hdd and ready > hdd:
        return "AT-RISK", RISK_FILL
    return "ON-TIME", OK_FILL


def current_dept(o):
    for d in S.DEPT_ORDER:
        cell = o["d"].get(d)
        if cell and cell[S.I_DONE] == 0:
            return d
    return "DONE"


def style_header(ws, n):
    for c in range(1, n + 1):
        hc = ws.cell(row=1, column=c)
        hc.font = HEAD_FONT; hc.fill = HEAD_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n)}{ws.max_row}"


def main():
    orders = S.load_orders()
    finish, start_at, load, load_models = S.schedule(orders)
    for o in orders:
        o["_ready"] = S.predicted_ready(o, finish)
        o["_status"], o["_fill"] = status_of(o, o["_ready"])

    wb = Workbook()

    # ---------- Sheet 1: Ready to Deliver (headline) ----------
    ws = wb.active; ws.title = "Ready to Deliver"
    H = ["SO / PO", "Customer", "Model", "Type", "Now at",
         "Predicted Ready", "Customer DD", "Hookka DD", "Status"]
    ws.append(H)
    rank = {"OVERDUE": 0, "LATE": 1, "AT-RISK": 2, "ON-HOLD": 3, "ON-TIME": 4}
    for o in sorted(orders, key=lambda x: (rank[x["_status"]],
                    x["_ready"] or date(2099, 1, 1))):
        ws.append([o["po"], o["cust"], o["pc"], o.get("cat"), current_dept(o),
                   o["_ready"].isoformat() if o["_ready"] else "",
                   o.get("cdd"), o.get("hdd"), o["_status"]])
        rr = ws.max_row
        ws.cell(row=rr, column=9).fill = o["_fill"]
    for i, w in enumerate([16, 18, 14, 10, 11, 16, 13, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, len(H))

    # ---------- Sheet 2: Department Schedule (full cascade) ----------
    ws2 = wb.create_sheet("Dept Schedule")
    H2 = ["SO / PO", "Model", "Type"] + [DEPT_SHORT[d] for d in S.DEPT_ORDER] + ["Ready", "Cust DD", "Status"]
    ws2.append(H2)
    for o in sorted(orders, key=lambda x: (x["_ready"] or date(2099, 1, 1), x["pc"])):
        row = [o["po"], o["pc"], o.get("cat")]
        for d in S.DEPT_ORDER:
            cell = o["d"].get(d)
            if cell is None:
                row.append("-")
            elif cell[S.I_DONE] == 1:
                row.append("done")
            else:
                fi = finish[o["po"]].get(d)
                row.append(fi.strftime("%m-%d") if fi else "")
        row += [o["_ready"].strftime("%m-%d") if o["_ready"] else "",
                o.get("cdd"), o["_status"]]
        ws2.append(row)
        ws2.cell(row=ws2.max_row, column=len(H2)).fill = o["_fill"]
    for i, w in enumerate([16, 13, 9] + [7] * 8 + [8, 12, 9], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    style_header(ws2, len(H2))

    # ---------- Sheet 3: Daily Plan per department ----------
    ws3 = wb.create_sheet("Daily Plan")
    H3 = ["Department", "Day", "Load (min)", "Cap (min)", "Util %", "Models (pieces)"]
    ws3.append(H3)
    for d in S.DEPT_ORDER:
        cap = S.CAP_MIN_PER_DAY[d]
        for iso in sorted(load[d].keys()):
            mins = load[d][iso]
            models = load_models[d][iso]
            mtxt = ", ".join(f"{m}×{round(p)}" for m, p in
                             sorted(models.items(), key=lambda kv: -kv[1]) if round(p) > 0)
            ws3.append([d, iso, round(mins), cap,
                        f"{round(100*mins/cap)}%", mtxt])
        ws3.append(["", "", "", "", "", ""])  # spacer between depts
    for i, w in enumerate([12, 12, 11, 10, 8, 70], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    style_header(ws3, len(H3))

    # ---------- Sheet 4: Summary & Notes ----------
    ws4 = wb.create_sheet("Summary & Notes", 0)
    cnt = defaultdict(int)
    for o in orders:
        cnt[o["_status"]] += 1
    notes = [
        ("生产排程 — 全部门接力 (Forward Capacity Schedule)", True, 14),
        ("生成 2026-06-01    产能窗口 2026-05-03~06-01 (近30天真实产出)", False, 10),
        ("", False, 10),
        ("【这张表怎么排出来的 — 就是给 Agent 用的逻辑】", True, 11),
        ("  1. 工序链:裁布→车缝、锯木→组框→打底、粘棉  三条线汇合→包面→包装。", False, 10),
        ("  2. 一张单只有当它需要的上游工序全做完,才进下一个部门(已完成前序的,马上能进)。", False, 10),
        ("  3. 每个部门每天只能做这么多(产能=近30天平均每天做的分钟数,见下表)。", False, 10),
        ("  4. 每个部门里:最急的交期先做;同一周到期的同款拼一起做(省开料/换模)。", False, 10),
        ("  5. 一张单在上一部门做完那天 = 它在下一部门能开工那天(接力就接这里)。", False, 10),
        ("  6. 排到最后包装做好那天 = 预计可交货日,跟 Customer DD 一比就知道迟不迟。", False, 10),
        ("", False, 10),
        ("【结果一览】", True, 11),
        (f"  共 {len(orders)} 张活跃单。", False, 10),
        (f"  已逾期 OVERDUE(交期已过)= {cnt['OVERDUE']}", False, 10),
        (f"  会迟交 LATE(按现在产能赶不上 Customer DD)= {cnt['LATE']}", False, 10),
        (f"  有风险 AT-RISK(赶得上客户,但破了内部缓冲 Hookka DD)= {cnt['AT-RISK']}", False, 10),
        (f"  准时 ON-TIME = {cnt['ON-TIME']}    暂停 ON-HOLD = {cnt['ON-HOLD']}", False, 10),
        ("", False, 10),
        ("【各部门产能 / 天 (近30天实际平均)】", True, 11),
    ]
    for d in S.DEPT_ORDER:
        notes.append((f"  {d:<11} {S.CAP_MIN_PER_DAY[d]:>5} 分钟/天   ≈ {S.CAP_PIECES_PER_DAY[d]} 件/天", False, 10))
    notes += [
        ("", False, 10),
        ("【瓶颈提示】", True, 11),
        ("  FOAM 粘棉产能最低(818 分钟/天),最容易卡住整条线 — 要加快交期,先看这里加人。", False, 10),
        ("", False, 10),
        ("【分页】", True, 11),
        ("  · Ready to Deliver:每张单预计可交货日 + 迟/准,按最急排前。先看这页。", False, 10),
        ("  · Dept Schedule:每张单在 8 个部门各自的预计完成日(整条接力)。", False, 10),
        ("  · Daily Plan:每个部门每天做哪几个款、做多满 — 给现场用。", False, 10),
        ("", False, 10),
        ("  注:只读线上资料,没改动系统任何东西。", False, 10),
    ]
    for i, (t, b, s) in enumerate(notes, 1):
        ws4.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
    ws4.column_dimensions["A"].width = 100

    out = r"C:\Users\User\hookka-erp-testing\Dept-Schedule-2026-06-01.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        out = r"C:\Users\User\hookka-erp-testing\Dept-Schedule-2026-06-01-v2.xlsx"
        wb.save(out)
    print("OK ->", out)
    print("status:", dict(cnt))


if __name__ == "__main__":
    main()
