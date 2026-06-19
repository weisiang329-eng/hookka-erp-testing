# =============================================================================
# HOOKKA ERP - Full Production Planning Simulation
# =============================================================================
# Implements docs/PRODUCTION-PLANNING-LOGIC.md (STEP 0-9) end to end, OFFLINE,
# from read-only snapshots. Produces:
#   1. A per-SO ready-date / late / stalled table (STEP 8 rollup).
#   2. A per-station day-by-day planning view for the 7 downstream stations,
#      styled to match the existing Cutting-Schedule book.
#
# It does NOT touch the live ERP or any source under src/. Pure planning.
#
# KEY SPEC POINTS HONOURED
#   - Team split by itemCategory (SOFA / BEDFRAME / OTHER), NOT card wipType.
#   - Capacity bucketed per (department x team), key "<DEPT>|<TEAM>".
#   - FAB_CUT capacity = models/day (BEDFRAME 6, SOFA 2); downstream 7 stations
#     = minutes/day measured from COMPLETED/TRANSFERRED cards in the recent
#     working-day window (2026-05-21 -> 2026-05-30, the FAB_CUT-validated win).
#   - FAB_CUT stage work in MODEL-SLOTS: sum over distinct models of
#     ceil(modelSets / setupCap), setupCap SOFA 3 / BEDFRAME 5.
#   - Each order's stage ORDER derived per-order from the cards' staged due date
#     (the snapshot has no `sequence` field; per-dept `dd` is monotonic down the
#     real production chain and is the faithful per-order proxy), with the
#     canonical dept index as a deterministic tie-breaker.
#   - Integer day-index simulation, LOCAL-date working calendar (skip Sundays +
#     public holidays; Saturdays ARE working days), one stage advance per order
#     per working day, next-day handoff availDay = d+1.
#   - SO rollup: ready = max(line finishes), dd = min(line dd), late = ready-dd,
#     stalled if any line never finishes (cap 0).
# =============================================================================

import json, os, math
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEPT_FULL = os.path.join(HERE, "sched_dept_full.json")

# Run date = snapshot generatedAt. First working day is "tomorrow" (STEP 7).
TODAY = date(2026, 6, 1)

# Public holidays: kv_config['public_holidays']. The snapshot does not embed the
# list; from scripts/seed-public-holiday-2026-05-01.mjs the seeded holiday is
# 2026-05-01 (Labour Day). 2026-06-01 (run date) is a confirmed no-production day
# per the existing scheduler. Add more here as the Employee module grows them.
HOLIDAYS = {date(2026, 5, 1), date(2026, 6, 1)}

# Canonical reference order (STEP 2) - used only as a deterministic tie-breaker
# for stage ordering; the primary order key is each card group's staged due date.
DEPT_REF = ["FAB_CUT", "FAB_SEW", "WOOD_CUT", "FRAMING", "WEBBING",
            "FOAM", "UPHOLSTERY", "PACKING"]
DEPT_IDX = {d: i for i, d in enumerate(DEPT_REF)}
DOWNSTREAM = ["FAB_SEW", "WOOD_CUT", "FRAMING", "WEBBING", "FOAM",
              "UPHOLSTERY", "PACKING"]

# STEP 4: capacity measurement window (FAB_CUT-validated recent working days).
WIN_FROM = date(2026, 5, 21)
WIN_TO = date(2026, 5, 30)

# FAB_CUT confirmed capacity (models/day) and model-slot setup caps (STEP 4).
FAB_CUT_CAP = {"SOFA": 2, "BEDFRAME": 6}        # models/day
SETUP_CAP = {"SOFA": 3, "BEDFRAME": 5}          # sets per cut slot
MODELS_CUT = {"FAB_CUT"}                         # measured in models/day


# ---------------------------------------------------------------------------
# Calendar helpers - LOCAL date, never UTC/toISOString (STEP 7).
# ---------------------------------------------------------------------------
def fmt(d):
    return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"


def is_working(d):
    # Skip Sundays (weekday()==6) and public holidays. Saturdays ARE working.
    return d.weekday() != 6 and d not in HOLIDAYS


def parse(s):
    if not s:
        return None
    s = str(s)[:10]
    try:
        y, m, d = s.split("-")
        return date(int(y), int(m), int(d))
    except Exception:
        return None


def team_of(cat):
    c = (cat or "").upper()
    if "SOFA" in c:
        return "SOFA"
    if "BED" in c:
        return "BEDFRAME"
    return "OTHER"


def model_of(pc):
    pc = (pc or "").strip()
    return pc.split("-")[0] if pc else "(none)"


def build_calendar(start, n=900):
    """Working days starting tomorrow (the day AFTER `start`). index d -> date."""
    cal = []
    d = start + timedelta(days=1)
    while len(cal) < n:
        if is_working(d):
            cal.append(d)
        d += timedelta(days=1)
    return cal


# ---------------------------------------------------------------------------
# STEP 4 - capacity per (dept x team)
# ---------------------------------------------------------------------------
def compute_capacity(dept_full):
    """Returns cap[key] in the dept's measurement unit, plus a human report.
    FAB_CUT  -> models/day (confirmed, hard-coded SOFA 2 / BEDFRAME 6).
    others   -> minutes/day = sum(prodTime*max(1,wipQty)) / opdays, where opdays
                = distinct working days in the window that dept|team logged a
                completion (skip not-yet-logged days; never count 0 as a day).
    """
    cap = {}
    report = []

    # FAB_CUT confirmed numbers (STEP 4).
    for tm, c in FAB_CUT_CAP.items():
        cap["FAB_CUT|" + tm] = c
        report.append(("FAB_CUT|" + tm, "models/day", float(c), None,
                       "confirmed (STEP 4)"))

    minSum = defaultdict(float)
    daysWithComp = defaultdict(set)
    for dept in DOWNSTREAM:
        for r in dept_full.get(dept, []):
            if r.get("cardStatus") not in ("COMPLETED", "TRANSFERRED"):
                continue
            cd = parse(r.get("completedDate"))
            if not cd or cd > TODAY:          # drop dirty future-dated cards
                continue
            if not (WIN_FROM <= cd <= WIN_TO):
                continue
            if not is_working(cd):
                continue
            key = dept + "|" + team_of(r.get("cat"))
            mins = (r.get("prodTime") or 0) * max(1, r.get("wipQty") or 1)
            minSum[key] += mins
            daysWithComp[key].add(cd)

    for dept in DOWNSTREAM:
        for tm in ("SOFA", "BEDFRAME", "OTHER"):
            key = dept + "|" + tm
            op = len(daysWithComp.get(key, ()))
            if op > 0:
                cap[key] = minSum[key] / op
                report.append((key, "min/day", cap[key], op, f"{minSum[key]:.0f} min / {op} opdays"))
            else:
                cap[key] = 0.0   # no completions -> no capacity -> stalled
    return cap, report


# ---------------------------------------------------------------------------
# STEP 5 - build each LINE's remaining stages from its WAITING cards
# ---------------------------------------------------------------------------
def build_lines(dept_full):
    """A 'line' = one (po, lineNo). Aggregate its WAITING cards per dept into
    stages. Returns list of line dicts."""
    # Gather waiting cards per line per dept, plus line meta.
    lines = {}   # (po,lineNo) -> {meta, depts:{dept:[cards]}}
    for dept, rows in dept_full.items():
        for r in rows:
            if r.get("cardStatus") != "WAITING":
                continue
            key = (r["po"], r["lineNo"])
            ln = lines.get(key)
            if ln is None:
                ln = lines[key] = {
                    "po": r["po"], "lineNo": r["lineNo"], "soNo": r["coSOId"],
                    "cust": r.get("custName"), "model": r.get("model"),
                    "pc": r.get("model"), "cat": r.get("cat"),
                    "cdd": r.get("cdd"), "hdd": r.get("hdd"),
                    "depts": defaultdict(list),
                }
            ln["depts"][dept].append(r)

    out = []
    for key, ln in lines.items():
        team = team_of(ln["cat"])
        stages = []
        for dept, cards in ln["depts"].items():
            stkey = dept + "|" + team
            # Stage-order proxy: the snapshot has no `sequence` field. The per-
            # dept `dd` is a coarse/clamped staged date that ties across depts,
            # but `edd` (the finer staged due date) is strictly monotonic down
            # the real production chain, so it faithfully recovers per-order
            # stage order. Canonical dept index breaks any remaining ties.
            edds = [parse(c.get("edd")) for c in cards if c.get("edd")]
            order_dd = min(edds) if edds else date(9999, 12, 31)
            if dept in MODELS_CUT:
                # FAB_CUT work measured in model-slots: sum over distinct models
                # of ceil(modelSets / setupCap).
                by_model = defaultdict(float)
                for c in cards:
                    by_model[model_of(c.get("model"))] += max(1, c.get("wipQty") or 1)
                sc = SETUP_CAP.get(team, 3)
                work = sum(math.ceil(s / sc) for s in by_model.values())
            else:
                # downstream measured in minutes
                work = sum((c.get("prodTime") or 0) * max(1, c.get("wipQty") or 1)
                           for c in cards)
            stages.append({"dept": dept, "key": stkey, "work": float(work),
                           "order_dd": order_dd})
        # STEP 5: sort stages by the group's staged due date, dept index tiebreak
        stages.sort(key=lambda s: (s["order_dd"], DEPT_IDX.get(s["dept"], 99)))
        # priority dd (STEP 5): customer DD ?? hookka DD ?? far future
        dd = parse(ln["cdd"]) or parse(ln["hdd"]) or date(9999, 12, 31)
        out.append({
            "po": ln["po"], "lineNo": ln["lineNo"], "soNo": ln["soNo"],
            "cust": ln["cust"], "pc": ln["pc"], "cat": ln["cat"],
            "team": team, "dd": dd,
            "cdd": parse(ln["cdd"]), "hdd": parse(ln["hdd"]),
            "stages": stages,
        })
    return out


# ---------------------------------------------------------------------------
# STEP 6 - day-by-day finite-capacity simulation (integer day index)
# ---------------------------------------------------------------------------
def simulate(lines, cap, cal):
    """Mutates per-line ptr/availDay/finishDay. Records per-station daily load.
    load[key][d] = list of (line, work_taken) worked that day.
    Returns load."""
    # Reset sim state.
    for ln in lines:
        ln["ptr"] = 0
        ln["availDay"] = 0
        ln["finishDay"] = None
        ln["stalled"] = False
        # remaining work per stage (copy so we can rerun)
        for st in ln["stages"]:
            st["rem"] = st["work"]

    # priority: ascending dd (STEP 6). overdue still scheduled by dd.
    priority = sorted(lines, key=lambda l: (l["dd"], l["soNo"], l["lineNo"]))

    load = defaultdict(lambda: defaultdict(list))   # key -> day -> [(line,take)]
    unfinished = set(id(l) for l in priority if l["stages"])
    by_id = {id(l): l for l in priority}

    horizon = len(cal)
    for d in range(horizon):
        if not unfinished:
            break
        budget = {k: float(v) for k, v in cap.items()}
        for ln in priority:
            if id(ln) not in unfinished:
                continue
            if ln["availDay"] > d:
                continue
            if ln["ptr"] >= len(ln["stages"]):
                continue
            st = ln["stages"][ln["ptr"]]
            c = cap.get(st["key"], 0.0)
            if c <= 0:
                # no capacity -> this line can never advance -> stalled
                ln["stalled"] = True
                unfinished.discard(id(ln))
                continue
            b = budget.get(st["key"], 0.0)
            if b <= 0:
                continue
            take = min(st["rem"], b)
            st["rem"] -= take
            budget[st["key"]] = b - take
            load[st["key"]][d].append((ln, take))
            if st["rem"] <= 1e-9:
                ln["ptr"] += 1
                if ln["ptr"] >= len(ln["stages"]):
                    ln["finishDay"] = d
                    unfinished.discard(id(ln))
                else:
                    ln["availDay"] = d + 1   # next-day handoff
    # any line still unfinished after horizon with no stall flag = ran out of
    # horizon (shouldn't happen with 900 days) -> treat as stalled defensively
    for ln in lines:
        if ln["stages"] and ln["finishDay"] is None and not ln["stalled"]:
            ln["stalled"] = True
    return load


# ---------------------------------------------------------------------------
# STEP 8 - SO-level rollup
# ---------------------------------------------------------------------------
def rollup(lines, cal):
    by_so = defaultdict(list)
    for ln in lines:
        by_so[ln["soNo"]].append(ln)
    out = []
    for so, lns in by_so.items():
        team = lns[0]["team"]
        dds = [l["dd"] for l in lns]
        dd = min(dds)
        stalled = any(l["stalled"] for l in lns)
        finishes = [l["finishDay"] for l in lns if l["finishDay"] is not None]
        if stalled or not finishes:
            ready = None
            late = None
        else:
            fd = max(finishes)
            ready = cal[fd]
            late = (ready - dd).days
        out.append({
            "soNo": so, "team": team, "cust": lns[0]["cust"],
            "dd": dd, "ready": ready, "late": late, "stalled": stalled,
            "nlines": len(lns),
        })
    out.sort(key=lambda r: (r["ready"] or date(9999, 12, 31), r["soNo"]))
    return out


# ---------------------------------------------------------------------------
# XLSX rendering (style matched to Cutting-Schedule book)
# ---------------------------------------------------------------------------
ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = F(bold=True, color="FFFFFF", size=11)
BAND_A = PatternFill("solid", fgColor="DDEBF7")
BAND_B = PatternFill("solid", fgColor="FFFFFF")
LATE_FILL = PatternFill("solid", fgColor="FCE4D6")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
ONTIME_FILL = PatternFill("solid", fgColor="E2EFDA")
STALL_FILL = PatternFill("solid", fgColor="D9D9D9")


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        hc = ws.cell(row=1, column=c)
        hc.font = HEAD_FONT
        hc.fill = HEAD_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"


def render(cap, report, so_rows, lines, load, cal):
    wb = Workbook()

    # ---------- Sheet: Summary & Notes ----------
    ws0 = wb.active
    ws0.title = "Summary & Notes"
    n = len(so_rows)
    scheduled = [r for r in so_rows if r["ready"] is not None]
    stalled = [r for r in so_rows if r["stalled"]]
    late = [r for r in scheduled if r["late"] is not None and r["late"] > 0]
    sofa = [r for r in scheduled if r["team"] == "SOFA"]
    bed = [r for r in scheduled if r["team"] == "BEDFRAME"]
    all_ready = [r["ready"] for r in scheduled]
    rng = (min(all_ready), max(all_ready)) if all_ready else (None, None)
    notes = [
        ("Production Planning - Capacity-Grounded Ready Dates  生产排程预测", True, 14),
        (f"Generated 生成: {fmt(TODAY)}    First working day 首个开工天: {fmt(cal[0])}", False, 10),
        (f"Capacity window 产能窗口: {fmt(WIN_FROM)} -> {fmt(WIN_TO)}  (recent working days)", False, 10),
        ("", False, 10),
        ("[Result 结果]", True, 11),
        (f"  Active SOs 活动单: {n}    Scheduled 已排到: {len(scheduled)}    Stalled 排不动: {len(stalled)}", False, 10),
        (f"  Late vs Customer DD 比客人交期迟: {len(late)}", False, 10),
        (f"  Ready-date range 做好日期范围: {fmt(rng[0]) if rng[0] else '-'} -> {fmt(rng[1]) if rng[1] else '-'}", False, 10),
        (f"  Sofa scheduled 沙发: {len(sofa)}    Bedframe scheduled 床架: {len(bed)}", False, 10),
        ("", False, 10),
        ("[How it works 逻辑] (per docs/PRODUCTION-PLANNING-LOGIC.md)", True, 11),
        ("  - Team split by item category (SOFA / BEDFRAME / OTHER). 按品类分队。", False, 10),
        ("  - Capacity per (station x team). FAB_CUT = models/day (Sofa 2, Bed 6);", False, 10),
        ("    downstream = minutes/day from recent completions. 下游用近期完成工时/天。", False, 10),
        ("  - Day-by-day finite-capacity queue, priority by customer DD; one", False, 10),
        ("    station advance per order per day, next-day handoff. 一天推一站,隔天接手。", False, 10),
        ("  - Accessories (OTHER) have ~0 downstream capacity -> stalled, handle", False, 10),
        ("    separately. 配件几乎没下游产能,标 stalled 单独处理。", False, 10),
        ("", False, 10),
        ("[Capacity derived 产能数字]", True, 11),
    ]
    for key, unit, val, op, note in report:
        v = f"{val:.0f}" if unit == "models/day" else f"{val:.0f}"
        opx = f" ({op} opdays)" if op else ""
        notes.append((f"    {key:24} = {v:>6} {unit}{opx}", False, 10))
    notes += [
        ("", False, 10),
        ("[Sheets 分页]", True, 11),
        ("  - By Order: one row per SO - ready date, days late, stalled flag.", False, 10),
        ("  - Station Day-by-Day: which orders each downstream station works each day.", False, 10),
        ("", False, 10),
        ("  Note: planning only, read from snapshots; nothing written to the ERP.", False, 10),
        ("  注:只是排程预测,从快照读取,没有改动系统。", False, 10),
    ]
    for i, (txt, bold, size) in enumerate(notes, start=1):
        ws0.cell(row=i, column=1, value=txt).font = F(bold=bold, size=size)
    ws0.column_dimensions["A"].width = 104

    # ---------- Sheet: By Order (SO rollup) ----------
    ws = wb.create_sheet("By Order")
    headers = ["SO No", "Team", "Customer", "Customer DD", "Predicted Ready",
               "Days Late", "Status"]
    ws.append(headers)
    for i, r in enumerate(so_rows):
        if r["stalled"]:
            status = "Stalled (no capacity)"
        elif r["ready"] is None:
            status = "Not scheduled"
        elif r["late"] is not None and r["late"] > 0:
            status = "Late"
        else:
            status = "On time"
        ws.append([
            r["soNo"], r["team"], r["cust"] or "",
            fmt(r["dd"]) if r["dd"].year < 9999 else "",
            fmt(r["ready"]) if r["ready"] else "(not scheduled)",
            r["late"] if (r["late"] is not None) else "",
            status,
        ])
        rr = ws.max_row
        band = BAND_A if (i % 2 == 0) else BAND_B
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = band
        sc = ws.cell(row=rr, column=7)
        if status == "Stalled (no capacity)":
            sc.fill = STALL_FILL
        elif status == "Late":
            sc.fill = LATE_FILL
        elif status == "On time":
            sc.fill = ONTIME_FILL
    widths = [16, 10, 22, 14, 16, 11, 22]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    style_header(ws, len(headers))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            if cell.font is None or cell.font.name != ARIAL:
                cell.font = F(size=10)
            cell.border = BORDER
            if cell.column in (2, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center")

    # ---------- Sheet: Station Day-by-Day ----------
    ws2 = wb.create_sheet("Station Day-by-Day")
    h2 = ["Date", "Station", "Team", "SO No", "Customer", "Model",
          "Work loaded", "Unit"]
    ws2.append(h2)
    band_i = 0
    # iterate stations in chain order, then by day
    for dept in DOWNSTREAM:
        for tm in ("SOFA", "BEDFRAME"):
            key = dept + "|" + tm
            day_map = load.get(key)
            if not day_map:
                continue
            for d in sorted(day_map.keys()):
                band_i += 1
                band = BAND_A if (band_i % 2 == 1) else BAND_B
                for (ln, take) in day_map[d]:
                    ws2.append([
                        fmt(cal[d]), dept, tm, ln["soNo"], ln["cust"] or "",
                        ln["pc"] or "", round(take, 1), "min",
                    ])
                    for c in range(1, len(h2) + 1):
                        ws2.cell(row=ws2.max_row, column=c).fill = band
    widths2 = [12, 13, 10, 16, 22, 16, 12, 6]
    for j, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    style_header(ws2, len(h2))
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(h2)):
        for cell in row:
            if cell.font is None or cell.font.name != ARIAL:
                cell.font = F(size=10)
            cell.border = BORDER
            if cell.column in (2, 3, 7, 8):
                cell.alignment = Alignment(horizontal="center")

    # ---------- Sheet: Station Daily Totals ----------
    ws3 = wb.create_sheet("Station Daily Totals")
    h3 = ["Date", "Station", "Team", "Orders", "Total min loaded", "Capacity/day", "Util %"]
    ws3.append(h3)
    for dept in DOWNSTREAM:
        for tm in ("SOFA", "BEDFRAME"):
            key = dept + "|" + tm
            day_map = load.get(key)
            if not day_map:
                continue
            capd = cap.get(key, 0.0)
            for d in sorted(day_map.keys()):
                tot = sum(t for _, t in day_map[d])
                util = (tot / capd * 100) if capd else 0
                ws3.append([fmt(cal[d]), dept, tm, len(day_map[d]),
                            round(tot, 0), round(capd, 0), round(util, 0)])
    widths3 = [12, 13, 10, 8, 16, 13, 8]
    for j, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(j)].width = w
    style_header(ws3, len(h3))
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=len(h3)):
        for cell in row:
            if cell.font is None or cell.font.name != ARIAL:
                cell.font = F(size=10)
            cell.border = BORDER
            if cell.column in (2, 3, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")

    out = os.path.join(ROOT, "Production-Plan-2026-06-02.xlsx")
    try:
        wb.save(out)
    except PermissionError:
        out = os.path.join(ROOT, "Production-Plan-2026-06-02-v2.xlsx")
        wb.save(out)
    return out


# ---------------------------------------------------------------------------
def main():
    dept_full = json.load(open(DEPT_FULL, encoding="utf-8"))
    cap, report = compute_capacity(dept_full)
    cal = build_calendar(TODAY)
    lines = build_lines(dept_full)
    load = simulate(lines, cap, cal)
    so_rows = rollup(lines, cal)
    out = render(cap, report, so_rows, lines, load, cal)

    # console report
    scheduled = [r for r in so_rows if r["ready"] is not None]
    stalled = [r for r in so_rows if r["stalled"]]
    late = [r for r in scheduled if r["late"] and r["late"] > 0]
    print("=== CAPACITY (per dept x team) ===")
    for key, unit, val, op, note in report:
        print(f"  {key:24} {val:8.0f} {unit:11} {note}")
    print()
    print(f"=== SIMULATION ===")
    print(f"  lines={len(lines)}  SOs={len(so_rows)}")
    print(f"  scheduled={len(scheduled)}  stalled={len(stalled)}  late_vs_customerDD={len(late)}")
    if scheduled:
        rd = [r['ready'] for r in scheduled]
        print(f"  ready-date range: {fmt(min(rd))} -> {fmt(max(rd))}")
        for tm in ("SOFA", "BEDFRAME"):
            t = [r['ready'] for r in scheduled if r['team'] == tm]
            if t:
                print(f"    {tm:9} cleared by {fmt(max(t))}  ({len(t)} SOs)")
    # station date ranges
    print("  station date ranges (downstream):")
    for dept in DOWNSTREAM:
        days = set()
        for tm in ("SOFA", "BEDFRAME", "OTHER"):
            days |= set(load.get(dept + "|" + tm, {}).keys())
        if days:
            print(f"    {dept:11} {fmt(cal[min(days)])} -> {fmt(cal[max(days)])} ({len(days)} working days)")
    print()
    print("->", out)


if __name__ == "__main__":
    main()
