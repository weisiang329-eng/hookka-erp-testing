# =============================================================================
# FRAMING + Sofa Foam Bonding + Upholstery + Packing — English report
# (presentation-only translation of build_framing_daily_xlsx.py). Imports the
# ORIGINAL module so all dates/hours/sets are byte-identical; only labels are
# English.
# =============================================================================
import os, importlib.util
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = r"C:\Users\User\Desktop\Production-Schedule-2026-06-02-EN"


def _load(name, fname):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, fname))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


fr = _load("fram_src", "build_framing_daily_xlsx.py")

LANE_ORDER = fr.LANE_ORDER
units = fr.units
foam_units = fr.foam_units
uph_units = fr.uph_units
pack_units = fr.pack_units
load = fr.load
foam_load = fr.foam_load
uph_load = fr.uph_load
CAP_HOURS = fr.CAP_HOURS
CAP_MIN = fr.CAP_MIN
FOAM_CAP_MIN = fr.FOAM_CAP_MIN
UPH_CAP_HOURS = fr.UPH_CAP_HOURS
FRAME_HANDOFF = fr.FRAME_HANDOFF
day_num = fr.day_num
parse = fr.parse
_pe = fr._pe
THIS_MON = fr.THIS_MON
mins_of = fr.mins_of

EN = {"BEDFRAME": "Bedframe", "SOFA": "Sofa"}

ARIAL = "Arial"
def F(**kw): return Font(name=ARIAL, **kw)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78"); HEAD_FONT = F(bold=True, color="FFFFFF", size=10)
DATE_FILL = PatternFill("solid", fgColor="7030A0"); DATE_FONT = F(bold=True, color="FFFFFF", size=10)
BF_FILL = PatternFill("solid", fgColor="DDEBF7"); SOFA_FILL = PatternFill("solid", fgColor="FFF2CC")
OVERDUE_FILL = PatternFill("solid", fgColor="F8CBAD")
LANE_FILL = {"BEDFRAME": BF_FILL, "SOFA": SOFA_FILL}

wb = Workbook()

# ---- Sheet 1: Framing Calendar (by DATE) ------------------------------------
ws = wb.active; ws.title = "Framing Calendar"
headers = ["Frame Date", "Lane", "SO ID", "Model", "Item (frame part)",
           "Qty", "Mins", "Sets", "Webbing same-day", "HB Foam same-day",
           "Customer", "Customer DD", "Wood done"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    hc = ws.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

units_sorted = sorted(units, key=lambda u: (u["day"], 0 if u["cat"] == "BEDFRAME" else 1,
                                             _pe(u["cdd"]), u["so"] or ""))
cur_date = None
for u in units_sorted:
    if u["day"] != cur_date:
        cur_date = u["day"]
        bfu = load["BEDFRAME"][cur_date]; sfu = load["SOFA"][cur_date]
        ws.append([cur_date.strftime("%Y-%m-%d  %a") + f"   (Day {day_num(cur_date)})"
                   + f"   -  Bedframe {bfu/60:.1f}h/{CAP_HOURS['BEDFRAME']}h  -  Sofa {sfu/60:.1f}h/{CAP_HOURS['SOFA']}h"])
        rr = ws.max_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
        ws.cell(row=rr, column=1).fill = DATE_FILL; ws.cell(row=rr, column=1).font = DATE_FONT
    lane = u["cat"]
    web_txt = f"{len(u['web'])} pcs / {u['web_min']}min" if u["web"] else "-"
    hbf_txt = f"{len(u['hbf'])} pcs / {u['hbf_min']}min" if u["hbf"] else "-"
    wd_txt = u["wood_done"].strftime("%m-%d %a") if u["wood_done"] else "done / no wood"
    for i, c in enumerate(u["cards"]):
        ws.append(["", EN[lane] if i == 0 else "", u["so"] if i == 0 else "",
                   " + ".join(u["models"]) if i == 0 else "",
                   c.get("wipLabel") or c.get("model"), max(1, int(c.get("wipQty") or 1)),
                   mins_of(c),
                   u["sets"] if i == 0 else "",
                   web_txt if i == 0 else "", hbf_txt if i == 0 else "",
                   c.get("custName") if i == 0 else "", c.get("cdd"),
                   wd_txt if i == 0 else ""])
        rr = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=rr, column=col).fill = LANE_FILL[lane]
            ws.cell(row=rr, column=col).border = BORDER
            ws.cell(row=rr, column=col).font = F(size=9)
        if parse(c.get("cdd")) and parse(c.get("cdd")) < THIS_MON:
            ws.cell(row=rr, column=12).fill = OVERDUE_FILL
widths = [22, 16, 14, 14, 28, 5, 7, 6, 16, 16, 16, 12, 16]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---- Sheet 2: By Day --------------------------------------------------------
ws2 = wb.create_sheet("By Day")
h2 = ["Date", "Day", "Bedframe h", "Bedframe cap", "Sofa h", "Sofa cap",
      "Bedframe SO", "Sofa SO", "Webbing pcs (same-day)", "HB Foam pcs (same-day)", "SOs (model)"]
ws2.append(h2)
perday = defaultdict(list)
for u in units: perday[u["day"]].append(u)
for d in sorted(perday):
    us = perday[d]
    bf = [u for u in us if u["cat"] == "BEDFRAME"]; sf = [u for u in us if u["cat"] == "SOFA"]
    txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
    ws2.append([d.strftime("%Y-%m-%d %a"), day_num(d),
                round(load["BEDFRAME"][d] / 60, 1), CAP_HOURS["BEDFRAME"],
                round(load["SOFA"][d] / 60, 1), CAP_HOURS["SOFA"],
                len(bf), len(sf),
                sum(len(u["web"]) for u in us), sum(len(u["hbf"]) for u in us), txt])
    rr = ws2.max_row
    for c in range(1, len(h2) + 1):
        ws2.cell(row=rr, column=c).border = BORDER; ws2.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h2) + 1):
    hc = ws2.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 9, 9, 9, 9, 9, 9, 18, 18, 64], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ---- Sheet 3: Sofa Foam Bonding ---------------------------------------------
ws3 = wb.create_sheet("Sofa Foam Bonding")
h3 = ["Date", "Day", "Foam hours h", "cap", "Pcs Base/Arm/Cush", "SOs (model)"]
ws3.append(h3)
fb_by_day = defaultdict(list)
for u in foam_units:
    fb_by_day[u["foam_day"]].append(u)
for d in sorted(fb_by_day):
    us = fb_by_day[d]
    nb = sum(len(u["base"]) for u in us); na = sum(len(u["arm"]) for u in us)
    nc = sum(len(u["cush"]) for u in us)
    so_txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
    ws3.append([d.strftime("%Y-%m-%d %a"), day_num(d),
                round(foam_load[d] / 60, 1), FOAM_CAP_MIN // 60, f"{nb}/{na}/{nc}", so_txt])
    rr = ws3.max_row
    for c in range(1, len(h3) + 1):
        ws3.cell(row=rr, column=c).border = BORDER; ws3.cell(row=rr, column=c).font = F(size=9)
        ws3.cell(row=rr, column=c).fill = SOFA_FILL
for c in range(1, len(h3) + 1):
    hc = ws3.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 12, 6, 18, 64], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ---- Sheet 4: Upholstery by day ---------------------------------------------
ws4 = wb.create_sheet("Upholstery")
h4 = ["Date", "Day", "Bedframe h", "Bedframe cap", "Sofa h", "Sofa cap",
      "Bedframe SO", "Sofa SO", "SOs (model)"]
ws4.append(h4)
uph_perday = defaultdict(list)
for u in uph_units: uph_perday[u["uph_day"]].append(u)
for d in sorted(uph_perday):
    us = uph_perday[d]
    bf = [u for u in us if u["cat"] == "BEDFRAME"]; sf = [u for u in us if u["cat"] == "SOFA"]
    txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
    ws4.append([d.strftime("%Y-%m-%d %a"), day_num(d),
                round(uph_load["BEDFRAME"][d] / 60, 1), UPH_CAP_HOURS["BEDFRAME"],
                round(uph_load["SOFA"][d] / 60, 1), UPH_CAP_HOURS["SOFA"],
                len(bf), len(sf), txt])
    rr = ws4.max_row
    for c in range(1, len(h4) + 1):
        ws4.cell(row=rr, column=c).border = BORDER; ws4.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h4) + 1):
    hc = ws4.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 9, 9, 9, 9, 9, 9, 64], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A2"

# ---- Sheet 5: Packing by day ------------------------------------------------
ws5 = wb.create_sheet("Packing")
h5 = ["Date", "Day", "Total pcs", "Bedframe SO", "Sofa SO", "SOs (model)"]
ws5.append(h5)
pack_perday = defaultdict(list)
for u in pack_units: pack_perday[u["pack_day"]].append(u)
for d in sorted(pack_perday):
    us = pack_perday[d]
    bf = [u for u in us if u["cat"] == "BEDFRAME"]; sf = [u for u in us if u["cat"] == "SOFA"]
    ncards = sum(len(u["cards"]) for u in us)
    txt = ", ".join(f"{u['so']}({'+'.join(u['models'])})" for u in us)
    ws5.append([d.strftime("%Y-%m-%d %a"), day_num(d), ncards, len(bf), len(sf), txt])
    rr = ws5.max_row
    for c in range(1, len(h5) + 1):
        ws5.cell(row=rr, column=c).border = BORDER; ws5.cell(row=rr, column=c).font = F(size=9)
for c in range(1, len(h5) + 1):
    hc = ws5.cell(row=1, column=c); hc.font = HEAD_FONT; hc.fill = HEAD_FILL
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate([18, 5, 9, 9, 9, 64], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = "A2"

# ---- Sheet 0: Summary & Notes -----------------------------------------------
ws0 = wb.create_sheet("Summary & Notes", 0)
tot_fmin = sum(u["fmin"] for u in units)
tot_web = sum(len(u["web"]) for u in units); tot_hbf = sum(len(u["hbf"]) for u in units)
nbf = sum(1 for u in units if u["cat"] == "BEDFRAME"); nsf = len(units) - nbf
lane_span = {}
for lane in LANE_ORDER:
    ds = [u["day"] for u in units if u["cat"] == lane]
    if ds: lane_span[lane] = (min(ds), max(ds))
notes = [
    ("Framing (+ Webbing, HB Foam, Sofa Foam Bonding, Upholstery, Packing) - Day-by-Day Schedule", True, 14),
    ("Generated 2026-06-01. Schedules ONLY cards still to do (WAITING); excludes done, on-hold, cancelled. Read-only, no ERP writes.", False, 10),
    ("", False, 10),
    ("[The chain] Fabric Cut -> Fabric Sew -> Wood Cut -> Framing. Framing runs after that order's wood cut is done.", True, 11),
    (f"  Start floor = that SO's wood-cut done day + {FRAME_HANDOFF} working day (next day). Wood already done / no wood -> can frame from day 1.", False, 10),
    ("", False, 10),
    ("[Capacity: by production hours, not set count. Bedframe and sofa are two separate crews, each its own hour pool]", True, 11),
    (f"  - Bedframe {CAP_HOURS['BEDFRAME']} hours/day (historical good days 5/28=22.5h, 5/29=15.6h).", False, 10),
    (f"  - Sofa {CAP_HOURS['SOFA']} hours/day (owner-set; sofa volume is low so history can't hit this, set from real capacity).", False, 10),
    (f"  - Total to do = {tot_fmin/60:.0f} hours (Bedframe {sum(u['fmin'] for u in units if u['cat']=='BEDFRAME')/60:.0f}h / Sofa {sum(u['fmin'] for u in units if u['cat']=='SOFA')/60:.0f}h).", False, 10),
    ("", False, 10),
    ("[Same SO together] one SO is always framed on the same day, never split. Bedframe = divan + headboard; sofa = base + arms + cushions together.", True, 11),
    ("", False, 10),
    ("[Bedframe Webbing + HB Foam Bonding on the same day as framing]", True, 11),
    ("  - Whichever day this SO is framed, its webbing and headboard (HB) foam bonding are done the same day.", False, 10),
    (f"  - These two run in parallel (other crew / station), they do NOT eat the {CAP_HOURS['BEDFRAME']} framing hours; the 'same-day' columns show pcs + minutes.", False, 10),
    (f"  - Same-day riders total: Webbing {tot_web} pcs, HB Foam {tot_hbf} pcs.", False, 10),
    ("", False, 10),
    ("[Sofa back stages: webbing same-day, foam bonding next day] (see 'Sofa Foam Bonding' sheet)", True, 11),
    ("  - Webbing (base + cushion + arm, three parts) done the same day as sofa framing, not capacity-limited.", False, 10),
    ("  - Foam bonding on the next working day after framing:", False, 10),
    (f"      - Base + armrest: scheduled by capacity, max {FOAM_CAP_MIN/60:.0f} hours/day, overflow pushed forward.", False, 10),
    (f"      - Back cushion: done the same day as the base (follows the base even if pushed), but does NOT count toward the {FOAM_CAP_MIN//60} hours.", False, 10),
    ("", False, 10),
    ("[Upholstery: next day + by capacity, bedframe and sofa separate] (see 'Upholstery' sheet)", True, 11),
    (f"  - Bedframe upholstery = framing's next day (webbing rides framing same day, so upholstery = framing + 1). Max {UPH_CAP_HOURS['BEDFRAME']} hours/day.", False, 10),
    (f"  - Sofa upholstery = next day after the base foam bonding is done. Max {UPH_CAP_HOURS['SOFA']} hours/day.", False, 10),
    (f"  - Capacity from the last two weeks' actuals (bedframe ~17h/day, sofa ~9h/day); overflow pushed forward, same SO together, due date first.", False, 10),
    (f"  - Total to do: Bedframe {sum(u['umin'] for u in uph_units if u['cat']=='BEDFRAME')/60:.0f}h, Sofa {sum(u['umin'] for u in uph_units if u['cat']=='SOFA')/60:.0f}h.", False, 10),
    ("", False, 10),
    ("[Packing: follows the upholstery date, not capacity-limited] (see 'Packing' sheet)", True, 11),
    ("  - Packing's due date is the same day as upholstery; no capacity cap, packs whichever day upholstery is done.", False, 10),
    ("", False, 10),
    ("[Sort] Customer DD first (whoever ships first is done first), then which order's wood is cut first.", True, 11),
    ("", False, 10),
    ("[Resulting spans]", True, 11),
    (f"  Bedframe  {lane_span['BEDFRAME'][0].strftime('%m-%d %a')} -> {lane_span['BEDFRAME'][1].strftime('%m-%d %a')}"
     f"   {nbf} SOs / {sum(u['fmin'] for u in units if u['cat']=='BEDFRAME')/60:.0f} hours", False, 10),
    (f"  Sofa      {lane_span['SOFA'][0].strftime('%m-%d %a')} -> {lane_span['SOFA'][1].strftime('%m-%d %a')}"
     f"   {nsf} SOs / {sum(u['fmin'] for u in units if u['cat']=='SOFA')/60:.0f} hours", False, 10),
    ("", False, 10),
    ("  Note: framing follows wood cut; if wood changes the framing floor moves and this is re-run. Read-only planning, not written to the ERP.", False, 10),
]
for i, (t, b, s) in enumerate(notes, 1):
    ws0.cell(row=i, column=1, value=t).font = F(bold=b, size=s)
ws0.column_dimensions["A"].width = 124

out = os.path.join(OUT_DIR, "4-Framing-Foam-Upholstery-Packing.xlsx")
wb.save(out)
print("OK ->", out)
print(f"framing SO-units: {len(units)}  (BF {nbf} / SOFA {nsf})  total {tot_fmin/60:.0f}h")
