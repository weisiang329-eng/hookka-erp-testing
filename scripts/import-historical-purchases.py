#!/usr/bin/env python3
# ---------------------------------------------------------------------------
# import-historical-purchases.py — one-shot driver for the 1206-line
# historical procurement backfill (19 suppliers).
#
# Runs four phases against the ERP API:
#   A. Raw materials      — POST /api/raw-materials/bulk-import
#                            (items.xlsx + 6 hand-curated UNMATCHED- items)
#   B. Suppliers          — POST /api/import/suppliers-from-history
#   C. Bindings           — POST /api/import/supplier-bindings-from-history
#   D. PO/GRN/PI triples  — POST /api/import/historical-purchases-backfill
#
# Source files (hard-coded paths):
#   C:/Users/User/Downloads/items.xlsx              (454 raw_materials)
#   C:/Users/User/Downloads/supplier item.xlsx      (1206 PI lines)
#   C:/Users/User/Downloads/supplier_sku_mapping.xlsx
#       (per-supplier sheets w/ description -> internal_code mapping)
#
# Auth: logs in via POST /api/auth/login and reuses the bearer token. Same
# pattern as scripts/import-prod-completions-2026-04-30.mjs. Override creds
# with HOOKKA_EMAIL / HOOKKA_PASSWORD env vars; override target with
# HOOKKA_API_BASE (default points at the prod Pages deployment).
#
# Usage:
#   python scripts/import-historical-purchases.py --dry        # all-phases dry
#   python scripts/import-historical-purchases.py --live       # all-phases live
#   python scripts/import-historical-purchases.py --live --phase A,B,C
#
# Bails on any phase reporting errors > 0 (exit code 1).
# Per project_migration_in_progress: this script is one-shot and gets
# deleted post-migration. Don't generalize.
# ---------------------------------------------------------------------------

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import requests

try:
    from rapidfuzz import fuzz, process as rf_process

    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DEFAULT_API_BASE = "https://hookka-erp-testing.pages.dev"
DEFAULT_EMAIL = process.env.HOOKKA_EMAIL ?? ""
DEFAULT_PASSWORD = process.env.HOOKKA_PASSWORD ?? ""

ITEMS_XLSX = Path("C:/Users/User/Downloads/items.xlsx")
SUPPLIER_ITEM_XLSX = Path("C:/Users/User/Downloads/supplier item.xlsx")
SUPPLIER_SKU_MAP_XLSX = Path("C:/Users/User/Downloads/supplier_sku_mapping.xlsx")

OUT_DIR = Path("tmp/import-historical-purchases")

# 6 special raw_materials for items the supplier_sku_mapping file flagged
# as 'unmatched' but represent real stock (not fees/rebates). Codes are
# slugified so they don't collide with anything in items.xlsx.
SPECIAL_UNMATCHED_ITEMS: list[dict[str, Any]] = [
    {
        "itemCode": "UNMATCHED-SPRAY-GUN",
        "description": "SPRAY GUN827 2.0MM (WHITE) WITH AIR CUP SET",
        "itemGroup": "EQUIPMEN",
        "baseUOM": "UNIT",
        # description in supplier_sku_mapping (used for matching)
        "_descMatch": "SPRAY GUN827 2.0MM (WHITE) WITH AIR CUP SET MADE IN TAIWAN",
    },
    {
        "itemCode": "UNMATCHED-STAR-FABRIC-CLOUD",
        "description": "STAR FABRIC, COLOR: CLOUD (#13)",
        "itemGroup": "S.M-FABR",
        "baseUOM": "M",
        "_descMatch": "STAR FABRIC, COLOR: CLOUD(COLOR : 13#)",
    },
    {
        "itemCode": "UNMATCHED-FINGER-JOINT-1X2X12",
        "description": "FINGER JOINT 1' X 2' X 12'",
        "itemGroup": "PLYWOOD",
        "baseUOM": "PCS",
        "_descMatch": "FINGER JOINT 1' X 2' X 12' (2BDLS X 10PCS X 12FT)",
    },
    {
        "itemCode": "UNMATCHED-NB-JACOB-K",
        "description": "NB-JACOB 2.0 (F) BEDFRAME (K)",
        "itemGroup": "B.OTHERS",
        "baseUOM": "SET",
        "_descMatch": "NB-JACOB 2.0 (F) B/FRAME (K) (183X190CM)",
    },
    {
        "itemCode": "UNMATCHED-PICCO-FG66151-11",
        "description": "PICCO FG66151-11 FABRIC",
        "itemGroup": "S.M-FABR",
        "baseUOM": "M",
        "_descMatch": "PICCO FG66151-11 (FABRIC) 2 ROLLS: 48.60M, 42.50M",
    },
    {
        "itemCode": "UNMATCHED-NLY25DG-12MM",
        "description": "1875X1050X12MM NLY25D/G D/GREY SPONGE",
        "itemGroup": "S.FILLER",
        "baseUOM": "PCS",
        "_descMatch": "1875MM X 1050MM X 12MM NLY25D/G D/GREY SPONGE (1PKT X 40 PCS)",
    },
]

# Descriptions that should NEVER produce a raw_materials row (fees/rebates).
# These appear as PI-only line items in the backfill.
FEE_DESCRIPTIONS = {
    "REBATE 3% FOR NOV INVOICE",
    "TRANSPORTATION FEE - HOOKKA TO JURU",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def normalise_desc(s: Any) -> str:
    return " ".join(str(s or "").split()).upper()


def to_iso_date(v: Any) -> str:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    if not s:
        return ""
    # Accept YYYY-MM-DD as-is
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass
    return s  # caller will validate


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def dump_json(p: Path, obj: Any) -> None:
    ensure_dir(p.parent)
    with p.open("w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False, default=str)


def chunked(seq: list[Any], n: int) -> list[list[Any]]:
    return [seq[i : i + n] for i in range(0, len(seq), n)]


# ---------------------------------------------------------------------------
# API client
# ---------------------------------------------------------------------------
class Client:
    def __init__(self, base: str, email: str, password: str) -> None:
        self.base = base.rstrip("/")
        self.email = email
        self.password = password
        self.token: str | None = None

    def login(self) -> None:
        url = f"{self.base}/api/auth/login"
        r = requests.post(
            url,
            json={"email": self.email, "password": self.password},
            timeout=30,
        )
        r.raise_for_status()
        j = r.json()
        token = (j.get("data") or {}).get("token")
        if not token:
            raise RuntimeError(f"login failed: {j}")
        self.token = token

    def post(self, path: str, body: dict[str, Any]) -> dict[str, Any]:
        if not self.token:
            raise RuntimeError("not logged in")
        url = f"{self.base}{path}"
        r = requests.post(
            url,
            json=body,
            headers={
                "content-type": "application/json",
                "authorization": f"Bearer {self.token}",
            },
            timeout=120,
        )
        if r.status_code >= 400:
            sys.stderr.write(f"\nAPI error {r.status_code} on {path}\n{r.text[:1000]}\n")
        r.raise_for_status()
        return r.json()


# ---------------------------------------------------------------------------
# Phase A — raw_materials
# ---------------------------------------------------------------------------
def load_items_xlsx() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(ITEMS_XLSX, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name: str) -> int:
        return headers.index(name) + 1

    ic = col("Item Code")
    bu = col("Base UOM")
    uc = col("UOM Count")
    de = col("Description")
    ig = col("Item Group")
    it = col("Item Type")
    sc = col("Stock Control")
    ia = col("Is Active")

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        item_code = ws.cell(r, ic).value
        if not item_code:
            continue
        rows.append(
            {
                "itemCode": str(item_code).strip(),
                "baseUOM": str(ws.cell(r, bu).value or "PCS").strip() or "PCS",
                "uomCount": float(ws.cell(r, uc).value or 1),
                "description": str(ws.cell(r, de).value or item_code).strip(),
                "itemGroup": (
                    str(ws.cell(r, ig).value).strip()
                    if ws.cell(r, ig).value
                    else "OTHERS"
                ),
                "itemType": (
                    str(ws.cell(r, it).value).strip()
                    if ws.cell(r, it).value
                    else None
                ),
                "stockControl": ws.cell(r, sc).value == "Checked",
                "isActive": ws.cell(r, ia).value == "Checked",
            }
        )
    return rows


def phase_a(client: Client, dry: bool) -> tuple[int, int]:
    print("\n=== Phase A — raw_materials ===")
    items = load_items_xlsx()
    print(f"  loaded {len(items)} rows from items.xlsx")

    # Add the 6 hand-curated UNMATCHED- items (drop the _descMatch helper).
    for spec in SPECIAL_UNMATCHED_ITEMS:
        items.append(
            {
                "itemCode": spec["itemCode"],
                "description": spec["description"],
                "baseUOM": spec["baseUOM"],
                "itemGroup": spec["itemGroup"],
                "uomCount": 1,
                "itemType": None,
                "stockControl": True,
                "isActive": True,
            }
        )
    print(f"  + 6 UNMATCHED- items = {len(items)} total")

    if dry:
        dump_json(OUT_DIR / "phase-a-payload.json", items)
        print(f"  DRY: payload written to {OUT_DIR / 'phase-a-payload.json'}")
        return (0, 0)

    chunks = chunked(items, 100)
    created = updated = 0
    for i, ch in enumerate(chunks, 1):
        res = client.post("/api/raw-materials/bulk-import", {"rows": ch})
        d = res.get("data") or {}
        created += int(d.get("created", 0))
        updated += int(d.get("updated", 0))
        print(
            f"  chunk {i}/{len(chunks)} ({len(ch)} rows) "
            f"created={d.get('created', 0)} updated={d.get('updated', 0)}"
        )
    print(f"Phase A done: created={created} updated={updated}")
    return (created, updated)


# ---------------------------------------------------------------------------
# Phase B — suppliers
# ---------------------------------------------------------------------------
def load_suppliers_from_history() -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(SUPPLIER_ITEM_XLSX, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    sc_idx = headers.index("Creditor Code") + 1
    sn_idx = headers.index("Creditor Name") + 1

    seen: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, sc_idx).value
        name = ws.cell(r, sn_idx).value
        if code and name and code not in seen:
            seen[str(code).strip()] = str(name).strip()
    return [{"code": k, "name": v} for k, v in sorted(seen.items())]


def phase_b(client: Client, dry: bool) -> tuple[int, int, list[Any]]:
    print("\n=== Phase B — suppliers ===")
    sups = load_suppliers_from_history()
    print(f"  {len(sups)} suppliers to upsert:")
    for s in sups:
        print(f"    {s['code']}  {s['name']}")

    if dry:
        dump_json(OUT_DIR / "phase-b-payload.json", sups)
        print(f"  DRY: payload written to {OUT_DIR / 'phase-b-payload.json'}")
        return (0, 0, [])

    res = client.post("/api/import/suppliers-from-history", {"suppliers": sups})
    inserted = int(res.get("inserted", 0))
    skipped = int(res.get("skipped", 0))
    errors = res.get("errors") or []
    print(f"Phase B done: inserted={inserted} skipped={skipped} errors={len(errors)}")
    return (inserted, skipped, errors)


# ---------------------------------------------------------------------------
# Phase C — supplier_material_bindings
# ---------------------------------------------------------------------------
def load_supplier_sku_map() -> dict[tuple[str, str], dict[str, Any]]:
    """Returns (supplier_code, normalised_description) -> mapping row dict."""
    wb = openpyxl.load_workbook(SUPPLIER_SKU_MAP_XLSX, data_only=True)
    out: dict[tuple[str, str], dict[str, Any]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue
        ws = wb[sheet_name]
        # Header is row 3 in these per-supplier sheets.
        if ws.max_row < 4:
            continue
        headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
        try:
            sc_i = headers.index("supplier_code") + 1
            in_i = headers.index("internal_code") + 1
            ex_i = headers.index("external_code") + 1
            de_i = headers.index("description") + 1
            sp_i = headers.index("total_spend") + 1
            ap_i = headers.index("avg_unit_price") + 1
            mm_i = headers.index("match_method") + 1
            lp_i = headers.index("last_purchase_date") + 1
        except ValueError:
            continue

        for r in range(4, ws.max_row + 1):
            scode = ws.cell(r, sc_i).value
            if not scode:
                continue
            desc = ws.cell(r, de_i).value
            internal = ws.cell(r, in_i).value
            external = ws.cell(r, ex_i).value
            method = ws.cell(r, mm_i).value
            spend = ws.cell(r, sp_i).value
            avg_price = ws.cell(r, ap_i).value
            last_purchase = ws.cell(r, lp_i).value

            row = {
                "supplier_code": str(scode).strip(),
                "internal_code": str(internal).strip() if internal else None,
                "external_code": str(external).strip() if external else None,
                "description": str(desc).strip() if desc else "",
                "match_method": method,
                "total_spend": float(spend or 0),
                "avg_unit_price": float(avg_price or 0),
                "last_purchase": last_purchase,
            }
            key = (str(scode).strip(), normalise_desc(desc))
            out[key] = row
    return out


def derive_bindings(
    sku_map: dict[tuple[str, str], dict[str, Any]],
    desc_to_unmatched_code: dict[str, str],
) -> list[dict[str, Any]]:
    """
    Walks the supplier_sku_mapping per-supplier sheets. For each row whose
    match_method is 'direct_code' or 'fuzzy_match', emits a binding using
    avg_unit_price (RM) → senate. For 'unmatched' rows we use the
    UNMATCHED-* itemCode mapped via _descMatch.

    isMainSupplier=1 for the supplier with the highest total_spend on each
    materialCode (one main per materialCode). Bindings without a known
    materialCode (unmatched + not in our 6-special list) are dropped.
    """
    candidates: list[dict[str, Any]] = []
    for (scode, _ndesc), row in sku_map.items():
        method = row["match_method"]
        material_code: str | None = None
        if method in ("direct_code", "fuzzy_match"):
            material_code = row["internal_code"]
        elif method == "unmatched":
            # description-keyed lookup → UNMATCHED-* code or skip (fee/rebate)
            material_code = desc_to_unmatched_code.get(
                normalise_desc(row["description"])
            )
            if not material_code:
                continue  # fee/rebate/genuinely-unknown — drop
        else:
            continue
        if not material_code:
            continue

        candidates.append(
            {
                "materialCode": material_code,
                "supplierCode": scode,
                "supplierSKU": row["external_code"]
                or row["internal_code"]
                or material_code,
                "unitPriceSen": int(round(float(row["avg_unit_price"] or 0) * 100)),
                "isMainSupplier": 0,  # filled in below
                "_total_spend": row["total_spend"],
            }
        )

    # Pick the main supplier per materialCode (highest total_spend). Ties go
    # to first-seen.
    by_material: dict[str, dict[str, Any]] = {}
    for c in candidates:
        mc = c["materialCode"]
        if mc not in by_material or c["_total_spend"] > by_material[mc]["_total_spend"]:
            by_material[mc] = c
    for c in candidates:
        if by_material[c["materialCode"]] is c:
            c["isMainSupplier"] = 1

    # Drop the helper key.
    for c in candidates:
        c.pop("_total_spend", None)
    return candidates


def phase_c(client: Client, dry: bool) -> tuple[int, int, int, list[Any]]:
    print("\n=== Phase C — supplier_material_bindings ===")
    sku_map = load_supplier_sku_map()
    desc_to_unmatched = {
        normalise_desc(spec["_descMatch"]): spec["itemCode"]
        for spec in SPECIAL_UNMATCHED_ITEMS
    }
    bindings = derive_bindings(sku_map, desc_to_unmatched)
    main_count = sum(1 for b in bindings if b["isMainSupplier"] == 1)
    print(f"  {len(bindings)} bindings ({main_count} flagged as main supplier)")

    if dry:
        dump_json(OUT_DIR / "phase-c-payload.json", bindings)
        print(f"  DRY: payload written to {OUT_DIR / 'phase-c-payload.json'}")
        return (0, 0, 0, [])

    chunks = chunked(bindings, 200)
    inserted = updated = skipped = 0
    errors: list[Any] = []
    for i, ch in enumerate(chunks, 1):
        res = client.post(
            "/api/import/supplier-bindings-from-history", {"bindings": ch}
        )
        inserted += int(res.get("inserted", 0))
        updated += int(res.get("updated", 0))
        skipped += int(res.get("skipped", 0))
        errors.extend(res.get("errors") or [])
        print(
            f"  chunk {i}/{len(chunks)} ({len(ch)} rows) "
            f"inserted={res.get('inserted', 0)} updated={res.get('updated', 0)} "
            f"skipped={res.get('skipped', 0)} errors={len(res.get('errors') or [])}"
        )
    print(
        f"Phase C done: inserted={inserted} updated={updated} skipped={skipped} "
        f"errors={len(errors)}"
    )
    return (inserted, updated, skipped, errors)


# ---------------------------------------------------------------------------
# Phase D — PO/GRN/PI triples
# ---------------------------------------------------------------------------
def load_supplier_items() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(SUPPLIER_ITEM_XLSX, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name: str) -> int:
        return headers.index(name) + 1

    dn_i = col("Doc No")
    dd_i = col("Doc Date")
    sc_i = col("Creditor Code")
    sn_i = col("Creditor Name")
    ic_i = col("Item Code")
    de_i = col("Detail Description")
    qt_i = col("Qty")
    up_i = col("Unit Price")

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        doc_no = ws.cell(r, dn_i).value
        if not doc_no:
            continue
        rows.append(
            {
                "doc_no": str(doc_no).strip(),
                "doc_date": to_iso_date(ws.cell(r, dd_i).value),
                "supplier_code": str(ws.cell(r, sc_i).value or "").strip(),
                "supplier_name": str(ws.cell(r, sn_i).value or "").strip(),
                "item_code": (
                    str(ws.cell(r, ic_i).value).strip()
                    if ws.cell(r, ic_i).value
                    else None
                ),
                "description": str(ws.cell(r, de_i).value or "").strip(),
                "qty": float(ws.cell(r, qt_i).value or 0),
                "unit_price": float(ws.cell(r, up_i).value or 0),
            }
        )
    return rows


def build_self_desc_map(
    supplier_rows: list[dict[str, Any]],
) -> dict[tuple[str, str], str]:
    """Build (supplier_code, normalised_desc) -> item_code from source rows
    that already have Item Code set. Lets us resolve null-Item-Code rows
    whose description appears verbatim elsewhere for the same supplier."""
    out: dict[tuple[str, str], str] = {}
    for r in supplier_rows:
        if r["item_code"] and r["description"]:
            key = (r["supplier_code"], normalise_desc(r["description"]))
            out[key] = r["item_code"]
    return out


def resolve_material_code(
    line: dict[str, Any],
    self_desc_map: dict[tuple[str, str], str],
    sku_map: dict[tuple[str, str], dict[str, Any]],
    desc_to_unmatched: dict[str, str],
    fee_norm: set[str],
    fuzzy_index: dict[str, list[tuple[str, str]]],
    fuzzy_threshold: int = 88,
) -> tuple[str | None, str]:
    """Resolution priority for a supplier-item line:
    1. Source row has an Item Code  → use it
    2. Description matches a fee/rebate                → null (PI-only)
    3. Description matches one of the 6 UNMATCHED-* spec
    4. (supplier_code, desc) was seen before in the source with an Item Code
    5. (supplier_code, desc) is in supplier_sku_mapping with direct_code/fuzzy
    6. token_set fuzzy match (>=threshold) against the supplier's known descs
    Otherwise → null.

    Returns (material_code, resolution_path) where resolution_path is one of:
    'item_code' | 'fee' | 'unmatched_spec' | 'self_desc' | 'sku_map' |
    'fuzzy' | 'unresolved'.
    """
    if line["item_code"]:
        return (line["item_code"], "item_code")
    ndesc = normalise_desc(line["description"])
    if not ndesc:
        return (None, "unresolved")
    if ndesc in fee_norm:
        return (None, "fee")
    if ndesc in desc_to_unmatched:
        return (desc_to_unmatched[ndesc], "unmatched_spec")
    key = (line["supplier_code"], ndesc)
    if key in self_desc_map:
        return (self_desc_map[key], "self_desc")
    m = sku_map.get(key)
    if m and m.get("internal_code") and m.get("match_method") in (
        "direct_code",
        "fuzzy_match",
    ):
        return (m["internal_code"], "sku_map")

    if HAVE_RAPIDFUZZ:
        candidates = fuzzy_index.get(line["supplier_code"], [])
        if candidates:
            choices = [c[0] for c in candidates]
            best = rf_process.extractOne(
                ndesc, choices, scorer=fuzz.token_set_ratio
            )
            if best and best[1] >= fuzzy_threshold:
                code = candidates[best[2]][1]
                return (code, "fuzzy")

    return (None, "unresolved")


def build_fuzzy_index(
    self_desc_map: dict[tuple[str, str], str],
    sku_map: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, list[tuple[str, str]]]:
    """Per-supplier list of (normalised_desc, item_code) for fuzzy fallback.
    Combines source-derived descriptions and supplier_sku_mapping descriptions
    so we cast a wide net. We dedupe on (supplier_code, ndesc) keeping the
    self_desc_map entry preferentially."""
    by_supplier: dict[str, dict[str, str]] = defaultdict(dict)
    for (scode, ndesc), m in sku_map.items():
        if m.get("internal_code") and m.get("match_method") in (
            "direct_code",
            "fuzzy_match",
        ):
            by_supplier[scode][ndesc] = m["internal_code"]
    for (scode, ndesc), code in self_desc_map.items():
        by_supplier[scode][ndesc] = code  # source wins on conflict
    return {sc: list(d.items()) for sc, d in by_supplier.items()}


def build_documents(
    supplier_rows: list[dict[str, Any]],
    sku_map: dict[tuple[str, str], dict[str, Any]],
    desc_to_unmatched: dict[str, str],
) -> list[dict[str, Any]]:
    """Group source rows by Doc No and resolve materialCode for each line."""
    self_desc_map = build_self_desc_map(supplier_rows)
    print(f"  built self-desc map with {len(self_desc_map)} entries")
    fuzzy_index = build_fuzzy_index(self_desc_map, sku_map)
    if HAVE_RAPIDFUZZ:
        print(
            f"  fuzzy fallback enabled (rapidfuzz, token_set_ratio >= 88) — "
            f"{sum(len(v) for v in fuzzy_index.values())} candidate descs"
        )
    else:
        print("  fuzzy fallback DISABLED (rapidfuzz not installed)")

    fee_norm = {normalise_desc(d) for d in FEE_DESCRIPTIONS}
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in supplier_rows:
        groups[r["doc_no"]].append(r)

    resolution_counts: dict[str, int] = defaultdict(int)
    docs: list[dict[str, Any]] = []
    null_count = total_lines = 0
    unresolved_descs: dict[str, int] = defaultdict(int)
    for doc_no, lines in groups.items():
        first = lines[0]
        items = []
        for ln in lines:
            total_lines += 1
            material_code, path = resolve_material_code(
                ln, self_desc_map, sku_map, desc_to_unmatched, fee_norm, fuzzy_index
            )
            resolution_counts[path] += 1
            if material_code is None:
                null_count += 1
                unresolved_descs[ln["description"]] += 1

            items.append(
                {
                    "materialCode": material_code,
                    "supplierSKU": material_code or ln["item_code"] or None,
                    "description": ln["description"] or material_code or "",
                    "qty": ln["qty"],
                    "unitPriceSen": int(round(ln["unit_price"] * 100)),
                }
            )
        docs.append(
            {
                "docNo": doc_no,
                "docDate": first["doc_date"],
                "supplierCode": first["supplier_code"],
                "items": items,
            }
        )

    # Sort by docDate, docNo for deterministic processing.
    docs.sort(key=lambda d: (d["docDate"], d["docNo"]))
    print(
        f"  {len(docs)} documents, {total_lines} lines total, "
        f"{null_count} lines without a materialCode (fee/rebate/unresolved)"
    )
    print(
        "  resolution paths: "
        + ", ".join(f"{k}={v}" for k, v in sorted(resolution_counts.items()))
    )
    if unresolved_descs:
        dump_json(
            OUT_DIR / "phase-d-unresolved-descs.json",
            sorted(unresolved_descs.items(), key=lambda kv: -kv[1]),
        )
        top = sorted(unresolved_descs.items(), key=lambda kv: -kv[1])[:8]
        print("  top unresolved descriptions:")
        for d, n in top:
            print(f"    {n}x  {d!r}")
    return docs


def phase_d(client: Client, dry: bool) -> tuple[dict[str, int], list[Any]]:
    print("\n=== Phase D — PO/GRN/PI triples ===")
    supplier_rows = load_supplier_items()
    sku_map = load_supplier_sku_map()
    desc_to_unmatched = {
        normalise_desc(spec["_descMatch"]): spec["itemCode"]
        for spec in SPECIAL_UNMATCHED_ITEMS
    }
    docs = build_documents(supplier_rows, sku_map, desc_to_unmatched)

    if dry:
        dump_json(OUT_DIR / "phase-d-payload.json", docs)
        print(f"  DRY: payload written to {OUT_DIR / 'phase-d-payload.json'}")
        return (
            {
                "documentsProcessed": 0,
                "posCreated": 0,
                "grnsCreated": 0,
                "pisCreated": 0,
                "lineItemsTotal": 0,
                "skipped": 0,
            },
            [],
        )

    chunks = chunked(docs, 20)
    totals = {
        "documentsProcessed": 0,
        "posCreated": 0,
        "grnsCreated": 0,
        "pisCreated": 0,
        "lineItemsTotal": 0,
        "skipped": 0,
    }
    errors: list[Any] = []
    for i, ch in enumerate(chunks, 1):
        t0 = time.time()
        res = client.post(
            "/api/import/historical-purchases-backfill", {"documents": ch}
        )
        for k in totals:
            totals[k] += int(res.get(k, 0))
        errors.extend(res.get("errors") or [])
        dt = time.time() - t0
        print(
            f"  chunk {i}/{len(chunks)} ({len(ch)} docs) {dt:.1f}s "
            f"docs={res.get('documentsProcessed', 0)} "
            f"POs={res.get('posCreated', 0)} GRNs={res.get('grnsCreated', 0)} "
            f"PIs={res.get('pisCreated', 0)} skipped={res.get('skipped', 0)} "
            f"errors={len(res.get('errors') or [])}"
        )
    print(
        "Phase D done:",
        " ".join(f"{k}={v}" for k, v in totals.items()),
        f"errors={len(errors)}",
    )
    return (totals, errors)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(description="Hookka historical-purchase backfill driver")
    parser.add_argument("--dry", action="store_true", help="Build payloads, write to tmp/, do NOT call API")
    parser.add_argument("--live", action="store_true", help="Run live against the API")
    parser.add_argument(
        "--phase",
        default="A,B,C,D",
        help="Comma-separated phases to run (default: A,B,C,D)",
    )
    args = parser.parse_args()

    if not args.dry and not args.live:
        sys.stderr.write("Pass --dry or --live\n")
        return 2
    if args.dry and args.live:
        sys.stderr.write("Pick one of --dry or --live\n")
        return 2

    phases = [p.strip().upper() for p in args.phase.split(",") if p.strip()]
    valid = {"A", "B", "C", "D"}
    if not all(p in valid for p in phases):
        sys.stderr.write(f"--phase must be subset of A,B,C,D (got {phases})\n")
        return 2

    api_base = os.environ.get("HOOKKA_API_BASE", DEFAULT_API_BASE)
    email = os.environ.get("HOOKKA_EMAIL", DEFAULT_EMAIL)
    password = os.environ.get("HOOKKA_PASSWORD", DEFAULT_PASSWORD)

    ensure_dir(OUT_DIR)

    print(f"Mode: {'DRY' if args.dry else 'LIVE'}")
    print(f"Target: {api_base}")
    print(f"Phases: {phases}")

    client = Client(api_base, email, password)
    if args.live:
        print(f"Logging in as {email}...")
        client.login()
        print("OK, got token.")

    summary: dict[str, Any] = {}
    bail = False

    if "A" in phases:
        summary["A"] = phase_a(client, args.dry)
    if "B" in phases:
        ins, skip, errs = phase_b(client, args.dry)
        summary["B"] = {"inserted": ins, "skipped": skip, "errors": errs}
        if errs:
            bail = True
    if "C" in phases:
        ins, upd, skip, errs = phase_c(client, args.dry)
        summary["C"] = {
            "inserted": ins,
            "updated": upd,
            "skipped": skip,
            "errors": errs,
        }
        if errs:
            bail = True
    if "D" in phases:
        totals, errs = phase_d(client, args.dry)
        summary["D"] = {"totals": totals, "errors": errs}
        if errs:
            bail = True

    out_summary = OUT_DIR / f"summary-{'dry' if args.dry else 'live'}.json"
    dump_json(out_summary, summary)
    print(f"\nFull summary written to {out_summary}")

    if bail:
        sys.stderr.write("\nERRORS detected in at least one phase. Exit 1.\n")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
